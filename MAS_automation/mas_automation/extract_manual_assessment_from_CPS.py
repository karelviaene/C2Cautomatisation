### IMPORT
### Loading libraries ###
import pandas as pd
from openpyxl import load_workbook
import sqlite3
from pathlib import Path

### Functions ###

# Function to extract manual assessment if present from the CPS excel file as a df
def extract_info_from_manual_assessment(ref, value_col, file_path):
    wanted_names = [
        "Carcinogenicity",
        "Disruption of endocrine system",
        "Mutagenicity/genotoxicity",
        "Reproductive toxicity",
        "Development toxicity",
        "Neurotoxicity",
        "Oral toxicity",
        "Inhalative toxicity",
        "Dermal toxicity",
        "Skin, Eye, Respiratory corrosion/irritation",
        "Sensitization",
        "Fish toxicity",
        "Invertebrate toxicity",
        "Algae toxicity",
        "Terrestrial toxicity",
        "Other species toxicity",
        "Persistence",
        "Bioaccumulation",
        "Combined PB risk flag",
        "Combined aquatic risk flag",
        "Climatic relevance/ozone depletion potential",
    ]

    def clean_col_name(name):
        return (
            name.strip()
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
            .replace(".", "_")
            .replace("/", "_")
            .replace(",", "")
        )

    def get_value_even_if_merged(ws, cell_address):
        cell = ws[cell_address]

        if cell.value is not None:
            return cell.value

        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                )
                return top_left_cell.value

        return None

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    result = {
        "ref": ref,
    }

    for row in range(1, ws.max_row + 1):
        value_a = get_value_even_if_merged(ws, f"A{row}")

        if value_a is None:
            continue

        value_a_text = str(value_a).strip()

        for wanted in wanted_names:
            if wanted.lower() in value_a_text.lower():
                sql_col_name = "C2C_assessment_" + clean_col_name(wanted)

                value_i = get_value_even_if_merged(ws, f"{value_col}{row}")

                result[sql_col_name] = value_i

    return pd.DataFrame([result])
# Function to put the df into a table in SQL (and update it if needed) for the manual assessment C2C
def save_or_replace_if_changed(df, db_path, table_name, ref, ref_col="ref"):
    df = df.copy()

    # Add ref if not already present
    df[ref_col] = ref

    # Put ref first
    first_cols = [ref_col]
    other_cols = [col for col in df.columns if col not in first_cols]
    df = df[first_cols + other_cols]

    print("Connecting to the database")
    print(f"Working on {ref}")
    conn = sqlite3.connect(db_path)

    # Check if table exists
    table_exists = pd.read_sql_query(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        conn,
        params=(table_name,)
    )

    if table_exists.empty:
        df.to_sql(table_name, conn, if_exists="replace", index=False)
        print(f"New table created {table_name}")
        conn.close()
        return

    # Read existing table columns
    existing_table = pd.read_sql_query(
        f'SELECT * FROM "{table_name}" LIMIT 0',
        conn
    )

    existing_cols = existing_table.columns.tolist()

    # Add missing columns to df
    for col in existing_cols:
        if col not in df.columns:
            df[col] = pd.NA

    # Add new df columns to SQL table if needed
    new_cols = [col for col in df.columns if col not in existing_cols]

    for col in new_cols:
        conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{col}" TEXT')
        existing_cols.append(col)

    # Reorder df to match SQL table
    df = df[existing_cols]

    # Check if this ref already exists
    old_df = pd.read_sql_query(
        f'SELECT * FROM "{table_name}" WHERE "{ref_col}" = ?',
        conn,
        params=(ref,)
    )

    if old_df.empty:
        df.to_sql(table_name, conn, if_exists="append", index=False)
        print(f"New record added {ref}")

    else:
        old_row = old_df.iloc[0]
        new_row = df.iloc[0]

        # Align and compare
        new_row = new_row[old_row.index]

        old_compare = old_row.fillna("").astype(str)
        new_compare = new_row.fillna("").astype(str)

        if old_compare.equals(new_compare):
            print("No change needed")

        else:
            conn.execute(
                f'DELETE FROM "{table_name}" WHERE "{ref_col}" = ?',
                (ref,)
            )

            df.to_sql(table_name, conn, if_exists="append", index=False)
            print(f"Data has changed, updating")

    conn.commit()
    conn.close()
    print("Closing the database")

### PROGRAM FUNCTION

# excel path
file_path = "/Users/juliakulpa/Desktop/test/CPS_CAS_000-00-221.xlsx"
db_path = '/Users/juliakulpa/Desktop/test/C2Cdatabase.db'

## Running the program

### This is for extraction of the manual assessment for each excel
cas_number = Path(file_path).stem.replace("CPS_CAS ", "")
ref = cas_number
value_col = "I"

df_1 = extract_info_from_manual_assessment(ref, value_col, file_path)
save_or_replace_if_changed(
    df=df_1,
    db_path=db_path,
    table_name="COLOUR ASSESSMENT C2C",
    ref=cas_number
)