
#### SET UP ####
import sqlite3
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


# Define the path to your SQLite database file
C2Cpath = "/Users/juliakulpa/Desktop/Test_excel_imports"
C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# C2Cfiles_path = os.path.join(C2Cpath)
db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")
# /Users/juliakulpa/Desktop/Test_excel_imports/Database/C2Cdatabase.db
print(db_path)
# # LOAD EXCEL CPS TEMPLATE
template_path = "/Users/juliakulpa/Desktop/Test_excel_imports/Template/CPS_CAS TEMPLATE V2.xlsm"
template_wb = load_workbook(template_path, read_only=False, keep_vba=True)
ws_template = template_wb["C2Coverview"]


### CUSTOM FUNCTIONS ###


def db_get_non_empty_columns(linked_db, lookup_column, lookup_value):
    """
    For the table `linked_db`, find all rows where `lookup_column = lookup_value`
    and return a comma-separated string of column names that have at least one
    non-empty value in those rows.

    Parameters are kept similar to db_to_excel_multiple_below for interface consistency.
    Only the following are actually used:
      - linked_db:   name of the table to inspect
      - lookup_column: column used for the WHERE condition
      - lookup_value: value to look up

    Returns:
      str: comma-separated column names with at least one non-empty value,
           or an empty string if nothing is found.
    """

    try:
        # Query all columns for the matching rows
        query = f"""
            SELECT *
            FROM {linked_db}
            WHERE {lookup_column} = ?
        """
        cursor.execute(query, (lookup_value,))
        rows = cursor.fetchall()

        if not rows:
            print(f"No rows found in {linked_db} where {lookup_column} = {lookup_value}")
            return ""

        # Get column names from cursor.description
        column_names = [desc[0] for desc in cursor.description]

        non_empty_columns = []

        # For each column, check if ANY row has a non-empty value
        for col_idx, col_name in enumerate(column_names):
            has_value = False
            for row in rows:
                value = row[col_idx]

                if value is None:
                    continue

                if isinstance(value, str):
                    if value.strip() == "":
                        continue

                # If we get here, the value is considered non-empty
                has_value = True
                break

            if has_value:
                non_empty_columns.append(col_name)

        result = ", ".join(non_empty_columns)
        return result

    except sqlite3.Error as e:
        print("SQLite error:", e)
        return ""

def db_get_non_empty_columns_1(maindb, main_ref, linked_db, link_ref, column_to_get,
                             lookup_column, lookup_value, label_excel):
    """
    Uses the same JOIN structure as db_to_excel_multiple_below, but instead of
    writing to Excel, it returns a string of column names from `linked_db` where
    at least one returned row contains a non-empty value in that column.

    Returns:
        str: comma-separated names of non-empty SQL columns.
    """

    try:
        # First get the rows using your exact JOIN pattern
        query = f"""
            SELECT a.[{column_to_get}]
            FROM {linked_db} a
            JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
        """

        cursor.execute(query, (lookup_value,))
        rows = cursor.fetchall()

        if not rows:
            print(f"No results found for {lookup_column} = {lookup_value}")
            return ""

        # Second: retrieve all columns from linked_db (structure inspection)
        schema_query = f"PRAGMA table_info({linked_db})"
        cursor.execute(schema_query)
        schema = cursor.fetchall()

        # schema[i][1] = column name
        all_columns = [col[1] for col in schema]

        non_empty_columns = []

        # For each SQL column in linked_db, check if ANY matched row has a value in that column
        for col_name in all_columns:
            test_query = f"""
                SELECT a.[{col_name}]
                FROM {linked_db} a
                JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
            """
            cursor.execute(test_query, (lookup_value,))
            col_rows = cursor.fetchall()

            # Determine if the column is non-empty
            has_value = False
            for (value,) in col_rows:
                if value is None:
                    continue
                if isinstance(value, str) and value.strip() == "":
                    continue
                has_value = True
                break

            if has_value:
                non_empty_columns.append(col_name)

        return ", ".join(non_empty_columns)

    except sqlite3.Error as e:
        print("SQLite error:", e)
        return ""

def refdb_to_excel_two_markers_OECD(
    maindb,
    main_ref,
    linked_db,
    link_ref,
    label1,             # e.g. "Hazard classification"
    label2,             # e.g. "Eye Irrit. 2"
    lookup_column,
    lookup_value,
    label_excel,        # anchor text in Excel, analogous to refdb_to_excel_source_right
    offset,
    include_resource=True
):
    """
    Reverse of add_info_right_two_markers_OECD, following the same style as
    refdb_to_excel_source_right.

    From SQL (linked_db):
      - Reads column named f"{label1}{label2}".
      - Optionally reads column 'resource-<sanitized label2>'.

    In Excel (ws_template):
      1) Finds a cell whose value contains label_excel (case-insensitive).
      2) Uses that cell's row as target_row.
      3) Starting from (that cell's column + 1 + offset), finds the first empty
         cell to the right and writes the SQL value there.
      4) If resource exists, writes it in the "Resource" column in that same row.
    """

    # helper: sanitize label like in add_info_* functions
    def sanitize_label(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace(" ", "-")
        s = re.sub(r"[^a-z0-9_\-]", "", s)
        s = re.sub(r"-{2,}", "-", s)
        return s or "unnamed"

    try:
        # 1) Get list of all columns in linked_db
        pragma_sql = f"PRAGMA table_info([{linked_db}])"
        cursor.execute(pragma_sql)
        cols_info = cursor.fetchall()
        all_cols = [row[1] for row in cols_info]  # row[1] is column name

        # Column used by the original Excel→SQL function
        col_name = f"{label1}{label2}"

        if col_name not in all_cols:
            print(f"Column '{col_name}' not found in table '{linked_db}'")
            return

        # Resource column name in SQL: resource-<sanitized label2>
        resource_sql_col = None
        if include_resource:
            safe_label2 = sanitize_label(label2)
            candidate_resource_col = f"resource-{safe_label2}"
            if candidate_resource_col in all_cols:
                resource_sql_col = candidate_resource_col

        # 2) Build SELECT list
        select_parts = [f"a.[{col_name}]"]
        if resource_sql_col:
            select_parts.append(f"a.[{resource_sql_col}]")
        select_list = ", ".join(select_parts)

        query = f"""
            SELECT {select_list}
            FROM {linked_db} a
            JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
            LIMIT 1
        """

        cursor.execute(query, (lookup_value,))
        row = cursor.fetchone()
        if not row:
            print(f"No results found for {lookup_column} = {lookup_value}")
            return

        main_val = row[0]
        resource_val = row[1] if (resource_sql_col and len(row) > 1) else None

        if main_val in (None, "") and (resource_val in (None, "")):
            print("No data in SQL to write back (main value and resource empty).")
            return

        # 3) Find target row in Excel via label_excel (analogous to refdb_to_excel_source_right)
        max_row = ws_template.max_row
        max_col = ws_template.max_column

        target_row = None
        anchor_col = None

        for excel_row in ws_template.iter_rows():
            for cell in excel_row:
                if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                    target_row = cell.row
                    anchor_col = cell.column
                    break
            if target_row is not None:
                break

        if target_row is None or anchor_col is None:
            print(f"Label '{label_excel}' not found in worksheet.")
            return

        # 4) Write main value to the first empty cell to the right of anchor_col (+ offset)
        if main_val not in (None, ""):
            # starting search column
            start_col = anchor_col + 1 + offset
            if start_col < anchor_col + 1:
                start_col = anchor_col + 1  # safety

            write_col = None
            scan_max_col = max(max_col, start_col)

            for cc in range(start_col, scan_max_col + 1):
                cell_val = ws_template.cell(row=target_row, column=cc).value
                # treat None or empty/whitespace string as empty
                if cell_val is None or (isinstance(cell_val, str) and cell_val.strip() == ""):
                    write_col = cc
                    break

            if write_col is None:
                write_col = scan_max_col + 1

            ws_template.cell(row=target_row, column=write_col).value = main_val
            print(
                f"Inserted '{main_val}' into cell "
                f"{ws_template.cell(row=target_row, column=write_col).coordinate}"
            )

        # 5) Optionally write resource to Excel "Resource" column
        if include_resource and resource_val not in (None, ""):
            resource_col_idx = None
            # Find the "Resource" header column in the template
            for hdr_row in ws_template.iter_rows():
                for hdr_cell in hdr_row:
                    if (
                        isinstance(hdr_cell.value, str)
                        and hdr_cell.value.strip().lower() == "resource"
                    ):
                        resource_col_idx = hdr_cell.column
                        break
                if resource_col_idx is not None:
                    break

            if resource_col_idx is not None:
                ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                print(
                    f"Inserted resource '{resource_val}' into cell "
                    f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                )
            else:
                print("Could not find 'Resource' column in Excel to write resource value.")

    except sqlite3.Error as e:
        print("SQLite error:", e)


def refdb_to_excel_two_markers_OECD_test(
    maindb,
    main_ref,
    linked_db,
    link_ref,
    label1,            # e.g. "Hazard classification"
    label2,            # e.g. "Eye Irrit. 2"
    lookup_column,
    lookup_value,
    offset,
    include_resource=True
):
    """
    Reverse of add_info_right_two_markers_OECD, following the argument style
    of refdb_to_excel_source_right.

    From SQL (linked_db):
      - Reads column named f"{label1}{label2}".
      - Optionally reads column 'resource-<sanitized label2>'.

    In Excel (ws_template):
      1) Find a row where two adjacent cells match (label1, label2) left→right.
      2) Starting from (label2 cell column + 1 + offset), find the first empty
         cell to the right and write the SQL value there.
      3) If resource exists, write it in the "Resource" column in that same row.
    """

    # helper: sanitize label like in add_info_* functions
    def sanitize_label(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace(" ", "-")
        s = re.sub(r"[^a-z0-9_\-]", "", s)
        s = re.sub(r"-{2,}", "-", s)
        return s or "unnamed"

    try:
        # 1) Get list of all columns in linked_db
        pragma_sql = f"PRAGMA table_info([{linked_db}])"
        cursor.execute(pragma_sql)
        cols_info = cursor.fetchall()
        all_cols = [row[1] for row in cols_info]  # row[1] is column name

        # Column used by the original Excel→SQL function
        col_name = f"{label1}{label2}"

        if col_name not in all_cols:
            print(f"Column '{col_name}' not found in table '{linked_db}'")
            return

        # Resource column name in SQL: resource-<sanitized label2>
        resource_sql_col = None
        if include_resource:
            safe_label2 = sanitize_label(label2)
            candidate_resource_col = f"resource-{safe_label2}"
            if candidate_resource_col in all_cols:
                resource_sql_col = candidate_resource_col

        # 2) Build SELECT list
        select_parts = [f"a.[{col_name}]"]
        if resource_sql_col:
            select_parts.append(f"a.[{resource_sql_col}]")
        select_list = ", ".join(select_parts)

        query = f"""
            SELECT {select_list}
            FROM {linked_db} a
            JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
            LIMIT 1
        """

        cursor.execute(query, (lookup_value,))
        row = cursor.fetchone()
        if not row:
            print(f"No results found for {lookup_column} = {lookup_value}")
            return

        main_val = row[0]
        resource_val = row[1] if (resource_sql_col and len(row) > 1) else None

        if main_val in (None, "") and (resource_val in (None, "")):
            print("No data in SQL to write back (main value and resource empty).")
            return

        # 3) Find target row in Excel via adjacent (label1, label2)
        max_row = ws_template.max_row
        max_col = ws_template.max_column

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        target_row = None
        label2_col = None

        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = ws_template.cell(row=r, column=c).value
                v2 = ws_template.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    label2_col = c + 1
                    break
            if target_row is not None:
                break

        if target_row is None or label2_col is None:
            print("Target row with adjacent labels not found in Excel.")
            return

        # 4) Write main value to the first empty cell to the right of label2 (+ offset)
        if main_val not in (None, ""):
            # starting search column
            start_col = label2_col + 1 + offset
            if start_col < label2_col + 1:
                start_col = label2_col + 1  # safety

            write_col = None
            # allow writing beyond current max_col
            scan_max_col = max(max_col, start_col)

            for cc in range(start_col, scan_max_col + 1):
                cell_val = ws_template.cell(row=target_row, column=cc).value
                # treat None or empty/whitespace string as empty
                if cell_val is None or (isinstance(cell_val, str) and cell_val.strip() == ""):
                    write_col = cc
                    break

            if write_col is None:
                write_col = scan_max_col + 1

            ws_template.cell(row=target_row, column=write_col).value = main_val
            print(
                f"Inserted '{main_val}' into cell "
                f"{ws_template.cell(row=target_row, column=write_col).coordinate}"
            )

        # 5) Optionally write resource to Excel "Resource" column
        if include_resource and resource_val not in (None, ""):
            resource_col_idx = None
            # Find the "Resource" header column in the template
            for hdr_row in ws_template.iter_rows():
                for hdr_cell in hdr_row:
                    if (
                        isinstance(hdr_cell.value, str)
                        and hdr_cell.value.strip().lower() == "resource"
                    ):
                        resource_col_idx = hdr_cell.column
                        break
                if resource_col_idx is not None:
                    break

            if resource_col_idx is not None:
                ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                print(
                    f"Inserted resource '{resource_val}' into cell "
                    f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                )
            else:
                print("Could not find 'Resource' column in Excel to write resource value.")

    except sqlite3.Error as e:
        print("SQLite error:", e)

def refdb_to_excel_two_markers_OECD_test_3(
    maindb,
    main_ref,
    linked_db,
    link_ref,
    label1,             # e.g. "Hazard classification"
    label2,             # e.g. "Eye Irrit. 2"
    lookup_column,
    lookup_value,
    label_excel,        # anchor text in Excel
    include_resource=True
):
    """
    From SQL (linked_db):
      - Reads column named f"{label1}{label2}".
      - Optionally reads column 'resource-<sanitized label2>'.

    In Excel (ws_template):
      - Finds cell containing label_excel.
      - Writes `label2` in the first column to the right of label_excel.
      - Writes the SQL value in the *next empty cell* to the right of that.
      - Optionally writes resource in the 'Resource' column.
    """

    def sanitize_label(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace(" ", "-")
        s = re.sub(r"[^a-z0-9_\-]", "", s)
        s = re.sub(r"-{2,}", "-", s)
        return s or "unnamed"

    try:
        # 1) Get list of all columns in linked_db
        pragma_sql = f"PRAGMA table_info([{linked_db}])"
        cursor.execute(pragma_sql)
        cols_info = cursor.fetchall()
        all_cols = [row[1] for row in cols_info]

        col_name = f"{label1}{label2}"
        if col_name not in all_cols:
            print(f"Column '{col_name}' not found in table '{linked_db}'")
            return

        # Resource column name in SQL: resource-<sanitized label2>
        resource_sql_col = None
        if include_resource:
            safe_label2 = sanitize_label(label2)
            candidate_resource_col = f"resource-{safe_label2}"
            if candidate_resource_col in all_cols:
                resource_sql_col = candidate_resource_col

        # 2) Build SELECT list
        select_parts = [f"a.[{col_name}]"]
        if resource_sql_col:
            select_parts.append(f"a.[{resource_sql_col}]")
        select_list = ", ".join(select_parts)

        query = f"""
            SELECT {select_list}
            FROM {linked_db} a
            JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
            LIMIT 1
        """

        cursor.execute(query, (lookup_value,))
        row = cursor.fetchone()
        if not row:
            print(f"No results found for {lookup_column} = {lookup_value}")
            return

        main_val = row[0]
        resource_val = row[1] if (resource_sql_col and len(row) > 1) else None

        # 3) Find target row in Excel via label_excel
        max_row = ws_template.max_row
        max_col = ws_template.max_column

        target_row = None
        anchor_col = None

        for excel_row in ws_template.iter_rows():
            for cell in excel_row:
                if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                    target_row = cell.row
                    anchor_col = cell.column
                    break
            if target_row is not None:
                break

        if target_row is None or anchor_col is None:
            print(f"Label '{label_excel}' not found in worksheet.")
            return

        # 4) Write label2 in the first column to the right of label_excel
        label2_col = anchor_col + 1
        ws_template.cell(row=target_row, column=label2_col).value = label2
        print(
            f"Inserted label2 '{label2}' into cell "
            f"{ws_template.cell(row=target_row, column=label2_col).coordinate}"
        )

        # 5) Write main value in the next empty cell to the right of label2
        if main_val not in (None, ""):
            write_col = None
            # allow scanning one beyond current max_col to append if needed
            scan_max = max_col + 1
            for cc in range(label2_col + 1, scan_max + 1):
                cell_val = ws_template.cell(row=target_row, column=cc).value
                if cell_val is None or (isinstance(cell_val, str) and cell_val.strip() == ""):
                    write_col = cc
                    break

            if write_col is None:
                write_col = scan_max + 1

            ws_template.cell(row=target_row, column=write_col).value = main_val
            print(
                f"Inserted '{main_val}' into cell "
                f"{ws_template.cell(row=target_row, column=write_col).coordinate}"
            )

        # 6) Optionally write resource to Excel "Resource" column
        if include_resource and resource_val not in (None, ""):
            resource_col_idx = None
            for hdr_row in ws_template.iter_rows():
                for hdr_cell in hdr_row:
                    if (
                        isinstance(hdr_cell.value, str)
                        and hdr_cell.value.strip().lower() == "resource"
                    ):
                        resource_col_idx = hdr_cell.column
                        break
                if resource_col_idx is not None:
                    break

            if resource_col_idx is not None:
                ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                print(
                    f"Inserted resource '{resource_val}' into cell "
                    f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                )
            else:
                print("Could not find 'Resource' column in Excel to write resource value.")

    except sqlite3.Error as e:
        print("SQLite error:", e)


def add_info_right_two_markers_OECD(sheet, label1: str, label2: str, maindatabase, newdatabase, mainID, include_resource: bool = True):
    """
    label 1 - first row to match (e.g. Hazard classification)
    label 2 - second row to match (e.g. Eye Irrit. 2)
    1) Find a row where two adjacent cells match (label1, label2) left→right.
    2) Capture first non-empty cell to the right of label2.
    3) Write to SQL columns named: {label1}{label2}

    Optional behavior (when include_resource=True):
      - Captures the sheet's 'Resource' column value (if present and not empty)
        into SQL column 'resource-<sanitized label2>' for the same row.
      - If no 'Resource' column exists or the cell is empty, skips creating that column.
    """

    # --- helpers ---
    def q(name: str) -> str:
        return f'"{name}"'

    def matches(val, needle: str) -> bool:
        if val is None:
            return False
        return needle.lower() in str(val).lower()

    # Sanitize label for safe SQL column naming (for resource column)
    def sanitize_label(s: str) -> str:
        s = (s or "").strip().lower()
        s = s.replace(" ", "-")
        s = re.sub(r"[^a-z0-9_\-]", "", s)
        s = re.sub(r"-{2,}", "-", s)
        return s or "unnamed"

    max_row = sheet.max_row
    max_col = sheet.max_column

    safe_label2 = sanitize_label(label2)
    resource_colname = f"resource-{safe_label2}"

    # --- 1) Find target row via adjacent (label1, label2) ---
    target_row = None
    for r in range(1, max_row + 1):
        for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
            v1 = sheet.cell(row=r, column=c).value
            v2 = sheet.cell(row=r, column=c + 1).value
            if matches(v1, label1) and matches(v2, label2):
                target_row = r
                break
        if target_row is not None:
            #print("Target row:", target_row)
            break

    if target_row is None:
        print("Target row not found")
        return  # no matching row → nothing to insert

    # --- Optionally find the column index for "Resource" ---
    resource_col = None
    if include_resource:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "resource":
                    resource_col = getattr(cell, "col_idx", cell.column)
                    break
            if resource_col:
                break

    # --- 2) Scan the row to find targets; capture right-hand values ---
    extracted_data = {}

    # label → col name
    col_name = f"{label1}{label2}"

    # move to look for the first value to the right
    def capture_right_of_label(row: int, label: str):
        # Find the column containing the label
        for c in range(1, max_col):
            cell_value = sheet.cell(row=row, column=c).value
            if matches(cell_value, label):

                # Start searching to the right of this column
                for cc in range(c + 1, max_col + 1):
                    right_val = sheet.cell(row=row, column=cc).value

                    # Skip empty or whitespace-only
                    if right_val is None:
                        continue

                    if isinstance(right_val, str):
                        rv = right_val.strip()
                        if rv == "":
                            continue
                        return rv  # return first non-empty string

                    # Non-string, non-None → return immediately
                    return right_val

                # If no value was found to the right
                return None

        # Label not found at all
        return None

    val = capture_right_of_label(target_row, label2)
    #print("Val", val)

    if val is not None:
        extracted_data[col_name] = val

    # Optionally add the Resource value, but only if it's not empty
    if include_resource and resource_col:
        resource_value = sheet.cell(row=target_row, column=resource_col).value
        if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
            extracted_data[resource_colname] = resource_value

    if not extracted_data:
        print("Target extracted not found")
        return  # nothing to insert

    # --- 3) Ensure table/columns exist ---
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
    table_exists = cursor.fetchone()

    needed_columns = list(extracted_data.keys())

    if not table_exists:
        cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
        fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
        cursor.execute(f'''
            CREATE TABLE {q(newdatabase)} (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                ref TEXT
                {"," if cols_def else ""} {cols_def}
                {fk_clause}
            )
        ''')
    else:
        cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
        existing_cols = [col[1] for col in cursor.fetchall()]
        if "ref" not in existing_cols and newdatabase != maindatabase:
            cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
        for col in needed_columns:
            if col not in existing_cols:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

    # --- 4) Upsert (same rules as your working pattern) ---
    if newdatabase != maindatabase:
        cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
        exists = cursor.fetchone()
        if exists:
            set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
            cursor.execute(
                f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                list(extracted_data.values()) + [mainID]
            )
        else:
            cols = ["ref"] + list(extracted_data.keys())
            placeholders = ", ".join(["?"] * len(cols))
            cursor.execute(
                f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                [mainID] + list(extracted_data.values())
            )
    else:
        cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
        exists = cursor.fetchone()
        if exists:
            set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
            cursor.execute(
                f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                list(extracted_data.values()) + [mainID]
            )
        else:
            cols = ["ID"] + list(extracted_data.keys())
            placeholders = ", ".join(["?"] * len(cols))
            cursor.execute(
                f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                [mainID] + list(extracted_data.values())
            )

##

def refdb_to_pandas(maindb, main_ref, linked_db, link_ref, column_to_get, lookup_column, lookup_value):
    """
    Query the database for all matching values in `column_to_get`
    (from `linked_db`) for rows where `maindb.lookup_column == lookup_value`,
    joined via `link_ref` (linked_db) and `main_ref` (maindb),
    and return the result as a pandas DataFrame.
    """
    try:
        query = f"""
            SELECT a.[{column_to_get}]
            FROM {linked_db} a
            JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
        """

        cursor.execute(query, (lookup_value,))
        rows = cursor.fetchall()

        if not rows:
            print(f"No results found for {lookup_column} = {lookup_value}")
            # Return empty DataFrame with the correct column name
            return pd.DataFrame(columns=[column_to_get])

        # Make DataFrame with a single column named as requested
        df = pd.DataFrame(rows, columns=[column_to_get])
        return df

    except sqlite3.Error as e:
        print("SQLite error:", e)
        # On error, return empty DataFrame with the correct column name
        return pd.DataFrame(columns=[column_to_get])

def refdb_to_pandas_multi(
    maindb, main_ref,
    linked_db, link_ref,
    columns_to_get,           # list of column names
    lookup_column, lookup_value
):
    """
    Query multiple columns from `linked_db` joined to `maindb` and
    return the result as a pandas DataFrame.

    columns_to_get must be a list, e.g.: ["col1", "col2", "col3"]
    """

    # Build SELECT a.[col1], a.[col2], ...
    select_clause = ", ".join([f"a.[{col}]" for col in columns_to_get])

    try:
        query = f"""
            SELECT {select_clause}
            FROM {linked_db} a
            JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
        """

        cursor.execute(query, (lookup_value,))
        rows = cursor.fetchall()

        if not rows:
            print(f"No results found for {lookup_column} = {lookup_value}")
            return pd.DataFrame(columns=columns_to_get)

        # Return DataFrame with proper column names
        return pd.DataFrame(rows, columns=columns_to_get)

    except sqlite3.Error as e:
        print("SQLite error:", e)
        return pd.DataFrame(columns=columns_to_get)


def refdb_to_column_names_unique(maindb, main_ref,linked_db, link_ref,lookup_column, lookup_value):
    """
    Returns ALL columns from `linked_db` for rows matched through the join
    with `maindb` where maindb.lookup_column == lookup_value.

    Output: pandas DataFrame with all columns from linked_db.
    """

    try:
        # Query EVERYTHING (*) from linked_db
        query = f"""
            SELECT a.*
            FROM {linked_db} a
            JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
            WHERE c.{lookup_column} = ?
        """

        cursor.execute(query, (lookup_value,))
        rows = cursor.fetchall()

        # If nothing found → return empty DataFrame (still safe)
        if not rows:
            print(f"No results found for {lookup_column} = {lookup_value}")
            return pd.DataFrame()

        # Extract column names automatically from cursor.description
        colnames = [desc[0] for desc in cursor.description]

        dataframe = pd.DataFrame(rows, columns=colnames)

        # cutting columns with NULL values
        dataframe_cut = dataframe.dropna(axis=1, how='all')
        # dropping columns with ID and ref (not needed here)
        dataframe_cut = dataframe_cut.drop(columns=["ID", 'ref'])
        column_name = list(dataframe_cut.columns)
        # takes away from the string the resources names
        result = [c_name for c_name in column_name if "resource" not in c_name.lower() ]
        return result

    except sqlite3.Error as e:
        print("SQLite error:", e)
        return pd.DataFrame()

def remove_text_from_string(string, target_name):
    result = []
    for s in string:
        name = s.replace(target_name, "").strip()
        result.append(name)
    return result

def fill_right_multiple_below_name(ws_template, target_name, values_string, label_excel,
                                   delimiter=",", allow_contains_label=True):
    """
    Looks for `label_excel` in ws_template. Starting one row below that label,
    while the cell in that column matches `target_name`, fill the cell to the right
    with successive items from `values_string`.

    Stops when:
      - target_name stops appearing in consecutive rows below the label, OR
      - values are exhausted.

    Parameters
    ----------
    ws_template : openpyxl worksheet
    target_name : str
        The specific name to look for in the cells below the label.
    values_string : str or list/tuple
        Values to write. If str, split on `delimiter`.
    label_excel : str
        The label to find in the worksheet.
    delimiter : str
        Used if values_string is a single string.
    allow_contains_label : bool
        If True, match label by "contains" (like your function).
        If False, match label by exact equality (case-insensitive).
    """

    try:
        # Normalize values into a list
        if values_string is None:
            print("No values provided.")
            return

        if isinstance(values_string, str):
            values = [v.strip() for v in values_string.split(delimiter) if v.strip() != ""]
        else:
            values = list(values_string)

        if not values:
            print("Values list is empty after parsing.")
            return

        # Find the label in the worksheet
        for row in ws_template.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = cell.value.strip().lower()
                    label_text = label_excel.strip().lower()

                    label_match = (
                        (label_text in cell_text) if allow_contains_label
                        else (label_text == cell_text)
                    )

                    if label_match:
                        start_row = cell.row + 1
                        col = cell.column

                        # Now walk downward while the target_name continues
                        r = start_row
                        i = 0  # index for values

                        while i < len(values):
                            current_val = ws_template.cell(row=r, column=col).value

                            # stop if the name sequence ends
                            if current_val is None:
                                break

                            if not isinstance(current_val, str):
                                break

                            if current_val.strip().lower() != target_name.strip().lower():
                                break

                            # write next value to the right
                            ws_template.cell(row=r, column=col + 1).value = values[i]
                            print(
                                f"Inserted '{values[i]}' into cell "
                                f"{ws_template.cell(row=r, column=col + 1).coordinate}"
                            )

                            # move down + advance values
                            r += 1
                            i += 1

                        if i == 0:
                            print(f"No consecutive '{target_name}' cells found below '{label_excel}'.")
                        elif i < len(values):
                            print(f"Stopped early: '{target_name}' ended after {i} row(s).")
                        else:
                            print(f"All {i} value(s) inserted.")

                        return

        print(f"Label '{label_excel}' not found in worksheet.")

    except Exception as e:
        print("Error:", e)


try:
    ### SQL SET-UP
    connection = sqlite3.connect('/Users/juliakulpa/Desktop/Test_excel_imports/Database /C2Cdatabase.db')
    cursor = connection.cursor()

    print("Connected to SQLite database at:", db_path)

    CAS = "10-00-0"

    # MUTAGENICITY
    #SCL = []
    #SCL_df_1 = refdb_to_pandas_multi(maindb="C2C_DATABASE", main_ref="ID", linked_db="MUTAGENICITY", link_ref="ref",columns_to_get=SCL, lookup_column="ID",lookup_value =CAS)
    #print(SCL_df_1_1)


    # SCL_all_df = refdb_to_pandas_all(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM", link_ref="ref", lookup_column="ID",lookup_value =CAS)
    # #print(SCL_all_df)
    # #cutting columns with NULL values
    # SCL_adjust_df = SCL_all_df.dropna(axis=1, how='all')
    # #dropping columns with ID and ref (not needed here)
    # SCL_adjust_df = SCL_adjust_df.drop(columns=["ID",'ref'])
    # #print(SCL_adjust_df)
    # names = list(SCL_adjust_df.columns)
    # print(names)

    point_mut_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT", link_ref="ref",
                                     lookup_column="ID", lookup_value=CAS)
    print(point_mut_names)
    point_mut_names= remove_text_from_string(point_mut_names, "Point mutations:")
    print(point_mut_names)

    ch_dam_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM",
                                                   link_ref="ref",
                                                   lookup_column="ID", lookup_value=CAS)
    ch_dam_names = remove_text_from_string(ch_dam_names, "Chromosome damaging:")
    print(ch_dam_names)

    SCL_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                link_ref="ref",
                                                lookup_column="ID", lookup_value=CAS)
    SCL_names = remove_text_from_string(SCL_names, " - Lower Limit: (%)")
    SCL_names = remove_text_from_string(SCL_names, " - Upper Limit: (%)")
    SCL_names_dist = list(dict.fromkeys(SCL_names))
    print(SCL_names_dist)

    #fill_right_multiple_below_name(template_wb, "Hazard classification:", names,"Hazard classification:" )




    #### SAVE THE FILLED IN CPS EXCEL ####
    template_wb.save('/Users/juliakulpa/Desktop/Test_excel_imports/Testing/Test-export-1.xlsm')

except sqlite3.Error as e:
    print("SQLite error:", e)



# for row in ws_template.iter_rows():
#     for cell in row:
#         if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#             start_row = cell.row + 1
#             col = cell.column
#             print(f"First test on{start_row, col}")
#
#             # Place each value in the first empty cell below the starting row
#             for result in results:
#                 # Start searching from start_row downward
#                 target_row = start_row
#
#                 # Keep moving down until we find an empty cell in the target column
#                 while ws_template.cell(row=target_row, column=col).value not in (None, ''):
#                     target_row += 1
#
#                 print(f"target row{target_row}")
#
#                 # Write the value in the first empty cell found
#                 ws_template.cell(row=target_row, column=col).value = result[0]
#
#                 print(f"Inserted '{result[0]}' into cell {ws_template.cell(row=target_row, column=col).coordinate}")
#
#
#
#
#
# # Insert each value below the label, adding rows if needed
#                     for i, result in enumerate(results):
#                         target_row = start_row + i
#                         row_cells = ws_template[target_row+1]
#
#                         # Check if all cells in the row are empty
#                         if all(cell.value in (None, '') for cell in row_cells):
#                             # Row exists and is empty — reuse it
#                             ws_template.cell(row=target_row, column=col).value = result[0]
#                         else:
#                             # Row has data — insert a new row
#                             ws_template.insert_rows(target_row)
#                             ws_template.cell(row=target_row, column=col).value = result[0]
#
#
#
#
#
# import os
# import zipfile
#
# #import Pillow
#
# # Change depending on where is the excel and where you want to save the images
# #output_dir = r"/Users/juliakulpa/Desktop/Imag_test/Photos"  # <-- put your directory here
# #excel_path = "/Users/juliakulpa/Desktop/Imag_test/Image.xlsx"
#
# C2Cpath = "/Users/juliakulpa/Desktop/test"
# C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# images_output = "/Users/juliakulpa/Desktop/test/Chem_image"
#
# import os
# import zipfile
#
# import os
# import zipfile
#
# def extract_all_images_from_excel(excel_path, output_dir):
#     """
#     Extract all embedded images from an Excel .xlsx or .xlsm file and save them to output_dir.
#     Images are renamed to "<excel_filename>-01.<ext>", "<excel_filename>-02.<ext>", etc.
#     Returns a list of saved file paths.
#     """
#     # Basic validations
#     # Skip non-files
#     if not os.path.isfile(excel_path):
#         print(f"Skipped (not a file): {excel_path}")
#         return []
#
#     # Get extension safely
#     _, ext_in = os.path.splitext(excel_path)
#     ext_in = ext_in.lower()
#
#     # Skip unsupported or extensionless files
#     if ext_in not in (".xlsx", ".xlsm"):
#         print(f"Skipped (unsupported or missing extension): {excel_path}")
#         return []
#     if not zipfile.is_zipfile(excel_path):
#         raise ValueError(f"The file doesn't look like a valid Excel Open XML package: {excel_path}")
#
#     os.makedirs(output_dir, exist_ok=True)
#
#     excel_name = os.path.splitext(os.path.basename(excel_path))[0]
#     saved_paths = []
#
#     with zipfile.ZipFile(excel_path, 'r') as z:
#         # Images live under xl/media in OOXML workbooks (both .xlsx and .xlsm)
#         image_files = [f for f in z.namelist() if f.startswith('xl/media/')]
#
#         if not image_files:
#             print(f"No images found in {excel_path}.")
#             return saved_paths
#
#         # Sort for deterministic ordering
#         image_files.sort()
#
#         for idx, img_name in enumerate(image_files, start=1):
#             img_data = z.read(img_name)
#             img_ext = os.path.splitext(img_name)[1]  # keep original extension from the package
#
#             #saves each time a new image
#             if idx == 1:
#                 filename = f"{excel_name}{img_ext}"
#             else:
#                 filename = f"{excel_name}-{idx - 1}{img_ext}"
#             output_path = os.path.join(output_dir, filename)
#
#             # If the same name exists, bump a counter
#             # if os.path.exists(output_path):
#             #     bump = 1
#             #     base, ext = os.path.splitext(filename)
#             #     while os.path.exists(output_path):
#             #         output_path = os.path.join(output_dir, f"{base}({bump}){ext}")
#             #         bump += 1
#             # Skip if file already exists
#             # if os.path.exists(output_path):
#             #     print(f"Skipped (already exists): {output_path}")
#             #     continue
#
#             with open(output_path, 'wb') as f:
#                 f.write(img_data)
#
#             saved_paths.append(output_path)
#             #print(f"Saved: {output_path}") # check-point
#
#     #print(f"\nExtracted {len(image_files)} image(s) from '{os.path.basename(excel_path)}' to: {output_dir}") # check-point
#     return saved_paths
#
#
#
# for filename in os.listdir(C2Cfiles_path):
#     full_path = os.path.join(C2Cfiles_path, filename)
#     #print([full_path]) # check-point
#     extract_all_images_from_excel(full_path, images_output)

#extract_all_images_from_excel("/Users/juliakulpa/Desktop/test/CPS/CPS_CAS 10-00-1.xlsx", images_output)

# def extract_all_images_from_excel(excel_path, output_dir):
#     """
#     Extracts all embedded images from an Excel (.xlsx) file
#     and saves them into a chosen folder.
#     """
#     # Create the folder if it doesn’t exist
#     os.makedirs(output_dir, exist_ok=True)
#
#     # Extract images from excel
#     with zipfile.ZipFile(excel_path, 'r') as z:
#         image_files = [f for f in z.namelist() if f.startswith('xl/media/')]
#
#         if not image_files:
#             print(f"No images found in {excel_path}.")
#             return
#
#         for img_name in image_files:
#             img_data = z.read(img_name)
#             filename = os.path.basename(img_name)
#             output_path = os.path.join(output_dir, filename)
#             with open(output_path, 'wb') as f:
#                 f.write(img_data)
#             print(f"Saved: {output_path}")
#
#     print(f"\nExtracted {len(image_files)} image(s) to: {output_dir}")

#extract_all_images_from_excel(excel_path, output_dir)
#
# EXTRACTING IMAGE TO EXCEL FILE
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image as XLImage
# from openpyxl.utils import get_column_letter
# import os
#
# def insert_image_below_label(excel_path, image_path, sheet_name=None, label_text="Image"):
#     """
#     Finds the cell with text `label_text` (case-insensitive) in an Excel file,
#     and inserts the given image directly below it (same column, next row).
#     """
#
#     if not os.path.exists(excel_path):
#         raise FileNotFoundError(f"Excel file not found: {excel_path}")
#     if not os.path.exists(image_path):
#         raise FileNotFoundError(f"Image file not found: {image_path}")
#
#     # Load workbook (must NOT be read_only to add images)
#     wb = load_workbook(excel_path)
#     ws = wb[sheet_name] if sheet_name else wb.active
#
#     # Find the cell labeled 'Image' (case-insensitive)
#     image_label_row = None
#     image_label_col = None
#     wanted = label_text.strip().lower()
#
#     for row in ws.iter_rows():
#         for cell in row:
#             val = cell.value
#             if isinstance(val, str) and val.strip().lower() == wanted:
#                 image_label_row = cell.row
#                 image_label_col = cell.column  # 1-based index
#                 break
#         if image_label_row is not None:
#             break
#
#     if image_label_row is None:
#         raise ValueError(f"No cell labeled '{label_text}' found.")
#
#     # Cell below the label
#     target_row = image_label_row + 1
#     target_col_letter = get_column_letter(image_label_col)
#     target_cell = f"{target_col_letter}{target_row}"
#
#     # Create and add the image
#     img = XLImage(image_path)
#     # Optional: resize the image
#     img.width = 100
#     img.height = 100
#
#     ws.add_image(img, target_cell)
#
#     # Optional: adjust row height to avoid clipping
#     ws.row_dimensions[target_row].height = max(ws.row_dimensions[target_row].height or 15, 80)
#
#     wb.save(excel_path)
#     print(f"Image inserted below '{label_text}' at {target_cell} in {excel_path}")
#
#
# # Example
#
# insert_image_below_label(excel_path=r"/Users/juliakulpa/Desktop/Imag_test/Image copy.xlsx", image_path=r"/Users/juliakulpa/Desktop/Imag_test/image2.png", sheet_name=None, label_text="Image")
#



# string = ["a", "c", "d"]
#
# name = []
# for i in string:
#     name.append(i + "-a")
# print(name)


# import pandas as pd
# # sheet = pd.read_excel("/Users/juliakulpa/Desktop/function-test.xlsx")
# # print(sheet)
#
# from openpyxl import load_workbook
#
# testing_excel = '/Users/juliakulpa/Desktop/test/CPS/CPS_CAS 10-00-0.xlsx'
# #testing_excel = "/Users/juliakulpa/Desktop/function-test.xlsx"
#
# wb = load_workbook(testing_excel, data_only=True)
# sheet = wb.active  # or wb["SheetName"]
#
# def extract_notifiers_resources_wide(sheet):
#     """
#     Find 'Notifiers' and 'Resources' headers (case-insensitive).
#     For each header, read up to 250 rows below (skip blanks).
#       - For 'Notifiers':   name = value 6 columns left of the header (col-6) (sensitive, you need to change it inside the function for it to work)
#       - For 'Resources':   name = value 7 columns left of the header (col-7)
#       - Section value = the cell under the header
#     Returns list of dicts, merged wide by name:
#       {'name': 'Canc', 'Notifiers': '54', 'Resources': 'ECHA'}
#     """
#
#     TARGETS = {"notifiers": "Notifiers", "resources": "Resources"}
#     NAME_OFFSETS = {"Notifiers": 6, "Resources": 7}  # excel sensitive, if columns change needs adjusting
#
#     max_row = sheet.max_row
#     max_col = sheet.max_column
#     rows_by_name = {}
#
#     def coerce(v):
#         if v is None:
#             return None
#         if isinstance(v, str):
#             s = v.strip()
#             try:
#                 if "." in s:
#                     return float(s)
#                 return int(s)
#             except ValueError:
#                 return s
#         return v
#
#     def clean_name(raw):
#         if raw is None:
#             return ""
#         name = str(raw).strip()
#         if name.endswith(":"):
#             name = name[:-1].strip()
#         return name
#
#     def process_column(header_row, header_col, section_label):
#         name_offset = NAME_OFFSETS.get(section_label, 6)
#         for r in range(header_row + 1, min(header_row + 251, max_row + 1)):
#             val = sheet.cell(row=r, column=header_col).value
#             if val is None or (isinstance(val, str) and val.strip() == ""):
#                 continue  # skip blanks, don't stop
#
#             name_col = max(1, header_col - name_offset)
#             raw_name = sheet.cell(row=r, column=name_col).value
#             name = clean_name(raw_name)
#             if not name:
#                 continue
#
#             if name not in rows_by_name:
#                 rows_by_name[name] = {"name": name}
#             rows_by_name[name][section_label] = coerce(val)
#
#     # Find headers and process
#     for row in sheet.iter_rows():
#         for cell in row:
#             v = cell.value
#             if v is None:
#                 continue
#             txt = str(v).strip().lower()
#             if txt in TARGETS:
#                 section = TARGETS[txt]
#                 header_row = cell.row
#                 header_col = getattr(cell, "col_idx", getattr(cell, "column", None))
#                 if isinstance(header_col, int) and 1 <= header_col <= max_col:
#                     process_column(header_row, header_col, section)
#
#     return [rows_by_name[k] for k in rows_by_name]
#
#
#
# data = extract_notifiers_resources_wide(sheet)
#
# #print(data)
#
# df = pd.DataFrame(data)
# print(df)
#
#
# ### Test functions:
#
# # def add_info_CPS_right_until_empty_res(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
# #     """
# #     Like add_info_CPS_one_cell_right, but starting from column_offsets[0] to the right,
# #     keeps reading consecutive cells until it finds the first empty cell.
# #     Column naming:
# #       - first value uses column_names[0] (base)
# #       - next values use column_names[1:], if present
# #       - beyond that, auto-name as base-1, base-2, ...
# #       - also captures the sheet's 'Resource' column value (if present on that row)
# #         into SQL column 'resource-<sanitized rowlabel>'.
# #     """
# #     if len(column_offsets) != len(column_names):
# #         raise ValueError("column_offsets and column_names must have the same length")
# #     if not column_offsets:
# #         return
# #
# #     # Quote identifiers for SQL safety
# #     def q(name: str) -> str:
# #         return f'"{name}"'
# #
# #     # Sanitize rowlabel for safe SQL column naming
# #     def sanitize_label(s: str) -> str:
# #         s = (s or "").strip().lower()
# #         s = s.replace(" ", "-")
# #         s = re.sub(r"[^a-z0-9_\-]", "", s)
# #         s = re.sub(r"-{2,}", "-", s)
# #         return s or "unnamed"
# #
# #     safe_rowlabel = sanitize_label(rowlabel)
# #     resource_colname = f"resource-{safe_rowlabel}"
# #
# #     # --- locate the cell containing rowlabel ---
# #     match_row_idx = None
# #     match_col_idx = None
# #     for row in sheet.iter_rows():
# #         for cell in row:
# #             if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
# #                 match_row_idx = cell.row
# #                 # prefer numeric index (openpyxl)
# #                 match_col_idx = getattr(cell, "col_idx", cell.column)
# #                 break
# #         if match_row_idx is not None:
# #             break
# #
# #     if match_row_idx is None:
# #         return  # nothing to insert
# #
# #     # Find the column index for "Resource" (sheet header named 'Resource')
# #     resource_col = None
# #     for row in sheet.iter_rows():
# #         for cell in row:
# #             if cell.value and str(cell.value).strip().lower() == "resource":
# #                 resource_col = getattr(cell, "col_idx", cell.column)
# #                 break
# #         if resource_col:
# #             break
# #
# #     # Determine start offset and base name
# #     start_offset = column_offsets[0]
# #     base_name = column_names[0]
# #
# #     # --- read to the right until the first empty cell ---
# #     extracted_data = {}
# #     k = 0
# #     max_col = sheet.max_column
# #     while (match_col_idx + start_offset + k) <= max_col:
# #         target = sheet.cell(row=match_row_idx, column=match_col_idx + start_offset + k)
# #         tv = target.value
# #         # stop at first empty/blank
# #         if tv is None or (isinstance(tv, str) and tv.strip() == ""):
# #             break
# #
# #         # choose column name
# #         if k < len(column_names):
# #             col_name = column_names[k]
# #         else:
# #             col_name = f"{base_name}-{k - (len(column_names) - 1)}" if len(column_names) > 0 else f"col-{k}"
# #
# #         extracted_data[col_name] = tv
# #         k += 1
# #
# #     # Add the Resource value (may be None if column not found); always include the column
# #     extracted_data[resource_colname] = sheet.cell(row=match_row_idx, column=resource_col).value if resource_col else None
# #
# #     if not extracted_data:
# #         return  # nothing to insert
# #
# #     # --- ensure table exists and has needed columns ---
# #     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
# #     table_exists = cursor.fetchone()
# #
# #     # set of all columns we might write this time
# #     needed_columns = list(extracted_data.keys())
# #
# #     if not table_exists:
# #         cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
# #         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
# #         cursor.execute(f'''
# #             CREATE TABLE {q(newdatabase)} (
# #                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
# #                 ref TEXT
# #                 {"," if cols_def else ""} {cols_def}
# #                 {fk_clause}
# #             )
# #         ''')
# #     else:
# #         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
# #         existing_cols = [col[1] for col in cursor.fetchall()]
# #         if "ref" not in existing_cols and newdatabase != maindatabase:
# #             cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
# #         for col in needed_columns:
# #             if col not in existing_cols:
# #                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
# #
# #     # --- upsert (same keying rules as your working function) ---
# #     if newdatabase != maindatabase:
# #         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
# #         exists = cursor.fetchone()
# #         if exists:
# #             update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
# #             cursor.execute(
# #                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
# #                 [extracted_data[col] for col in needed_columns] + [mainID]
# #             )
# #         else:
# #             all_cols = ['ref'] + needed_columns
# #             placeholders = ", ".join(["?"] * len(all_cols))
# #             cursor.execute(
# #                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
# #                 [mainID] + [extracted_data[col] for col in needed_columns]
# #             )
# #     else:
# #         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
# #         exists = cursor.fetchone()
# #         if exists:
# #             update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
# #             cursor.execute(
# #                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
# #                 [extracted_data[col] for col in needed_columns] + [mainID]
# #             )
# #         else:
# #             all_cols = ['ID'] + needed_columns
# #             placeholders = ", ".join(["?"] * len(all_cols))
# #             cursor.execute(
# #                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
# #                 [mainID] + [extracted_data[col] for col in needed_columns]
# #             )
#
#
# # This function works but puts the resource column for all the sheets, so even if its not given (=empty row), the resource column will appear
# # def add_info_CPS_right_until_empty(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID, include_resource=True):
# #     """
# #     Like add_info_CPS_one_cell_right, but starting from column_offsets[0] to the right,
# #     keeps reading consecutive cells until it finds the first empty cell.
# #     Column naming:
# #       - first value uses column_names[0] (base)
# #       - next values use column_names[1:], if present
# #       - beyond that, auto-name as base-1, base-2, ...
# #       - If include_resource=True, also captures the sheet's 'Resource' column value (if present on that row)
# #         into SQL column 'resource-<sanitized rowlabel>'.
# #     """
# #     if len(column_offsets) != len(column_names):
# #         raise ValueError("column_offsets and column_names must have the same length")
# #     if not column_offsets:
# #         return
# #
# #     # Quote identifiers for SQL safety
# #     def q(name: str) -> str:
# #         return f'"{name}"'
# #
# #     # Sanitize rowlabel for safe SQL column naming
# #     def sanitize_label(s: str) -> str:
# #         s = (s or "").strip().lower()
# #         s = s.replace(" ", "-")
# #         s = re.sub(r"[^a-z0-9_\-]", "", s)
# #         s = re.sub(r"-{2,}", "-", s)
# #         return s or "unnamed"
# #
# #     safe_rowlabel = sanitize_label(rowlabel)
# #     resource_colname = f"resource-{safe_rowlabel}"
# #
# #     # --- locate the cell containing rowlabel ---
# #     match_row_idx = None
# #     match_col_idx = None
# #     for row in sheet.iter_rows():
# #         for cell in row:
# #             if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
# #                 match_row_idx = cell.row
# #                 # prefer numeric index (openpyxl)
# #                 match_col_idx = getattr(cell, "col_idx", cell.column)
# #                 break
# #         if match_row_idx is not None:
# #             break
# #
# #     if match_row_idx is None:
# #         return  # nothing to insert
# #
# #     # Optionally find the column index for "Resource"
# #     resource_col = None
# #     if include_resource:
# #         for row in sheet.iter_rows():
# #             for cell in row:
# #                 if cell.value and str(cell.value).strip().lower() == "resource":
# #                     resource_col = getattr(cell, "col_idx", cell.column)
# #                     break
# #             if resource_col:
# #                 break
# #
# #     # Determine start offset and base name
# #     start_offset = column_offsets[0]
# #     base_name = column_names[0]
# #
# #     # --- read to the right until the first empty cell ---
# #     extracted_data = {}
# #     k = 0
# #     max_col = sheet.max_column
# #     while (match_col_idx + start_offset + k) <= max_col:
# #         target = sheet.cell(row=match_row_idx, column=match_col_idx + start_offset + k)
# #         tv = target.value
# #         # stop at first empty/blank
# #         if tv is None or (isinstance(tv, str) and tv.strip() == ""):
# #             break
# #
# #         # choose column name
# #         if k < len(column_names):
# #             col_name = column_names[k]
# #         else:
# #             col_name = f"{base_name}-{k - (len(column_names) - 1)}" if len(column_names) > 0 else f"col-{k}"
# #
# #         extracted_data[col_name] = tv
# #         k += 1
# #
# #     # Optionally add the Resource value
# #     if include_resource:
# #         extracted_data[resource_colname] = (
# #             sheet.cell(row=match_row_idx, column=resource_col).value if resource_col else None
# #         )
# #
# #     if not extracted_data:
# #         return  # nothing to insert
# #
# #     # --- ensure table exists and has needed columns ---
# #     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
# #     table_exists = cursor.fetchone()
# #
# #     needed_columns = list(extracted_data.keys())
# #
# #     if not table_exists:
# #         cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
# #         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
# #         cursor.execute(f'''
# #             CREATE TABLE {q(newdatabase)} (
# #                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
# #                 ref TEXT
# #                 {"," if cols_def else ""} {cols_def}
# #                 {fk_clause}
# #             )
# #         ''')
# #     else:
# #         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
# #         existing_cols = [col[1] for col in cursor.fetchall()]
# #         if "ref" not in existing_cols and newdatabase != maindatabase:
# #             cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
# #         for col in needed_columns:
# #             if col not in existing_cols:
# #                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
# #
# #     # --- upsert logic (same keying as your previous working function) ---
# #     if newdatabase != maindatabase:
# #         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
# #         exists = cursor.fetchone()
# #         if exists:
# #             update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
# #             cursor.execute(
# #                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
# #                 [extracted_data[col] for col in needed_columns] + [mainID]
# #             )
# #         else:
# #             all_cols = ['ref'] + needed_columns
# #             placeholders = ", ".join(["?"] * len(all_cols))
# #             cursor.execute(
# #                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
# #                 [mainID] + [extracted_data[col] for col in needed_columns]
# #             )
# #     else:
# #         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
# #         exists = cursor.fetchone()
# #         if exists:
# #             update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
# #             cursor.execute(
# #                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
# #                 [extracted_data[col] for col in needed_columns] + [mainID]
# #             )
# #         else:
# #             all_cols = ['ID'] + needed_columns
# #             placeholders = ", ".join(["?"] * len(all_cols))
# #             cursor.execute(
# #                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
# #                 [mainID] + [extracted_data[col] for col in needed_columns]
# #             )

## Function for OECD Muta test. Really nice one:)
# def add_info_right_two_markers_OECD(sheet, label1: str, label2: str, maindatabase, newdatabase, mainID):
#     """
#     label 1 - first row to match (e.g. Hazard classification)
#     label 2 - second row to match (e.g. Eye Irrit. 2)
#     1) Find a row where two adjacent cells match (label1, label2) left→right.
#     3) Write to SQL columns named:{label2}
#     Matching is case-insensitive 'contains'.
#     Requires a DB cursor in outer scope named `cursor`.
#     """
#
#     # --- helpers ---
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#     def matches(val, needle: str) -> bool:
#         if val is None:
#             return False
#         return needle.lower() in str(val).lower()
#
#     max_row = sheet.max_row
#     max_col = sheet.max_column
#
#     # --- 1) Find target row via adjacent (label1, label2) ---
#     target_row = None
#     for r in range(1, max_row + 1):
#         for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
#             v1 = sheet.cell(row=r, column=c).value
#             v2 = sheet.cell(row=r, column=c + 1).value
#             if matches(v1, label1) and matches(v2, label2):
#                 target_row = r
#                 break
#         if target_row is not None:
#             print("Target row:",target_row)
#             break
#
#     if target_row is None:
#         print("Target row not found")
#         return  # no matching row → nothing to insert
#
#     # --- 2) Scan the row to find targets; capture right-hand values ---
#     extracted_data = {}
#
#     # label → col name
#     col_name = f"{label1}{label2}"
#     # move to look for the first value to the right
#     def capture_right_of_label(row: int, label: str):
#         for c in range(1, max_col):  # Find the column containing the label
#             cell_value = sheet.cell(row=row, column=c).value
#             if matches(cell_value, label):
#
#                 # Start searching to the right of this column
#                 for cc in range(c + 1, max_col + 1):
#                     right_val = sheet.cell(row=row, column=cc).value
#
#                     # Skip empty or whitespace-only
#                     if right_val is None:
#                         continue
#
#                     if isinstance(right_val, str):
#                         rv = right_val.strip()
#                         if rv == "":
#                             continue
#                         return rv  # return first non-empty string
#
#                     # Non-string, non-None → return immediately
#                     return right_val
#
#                 # If no value was found to the right
#                 return None
#
#         # Label not found at all
#         return None
#
#     val = capture_right_of_label(target_row, label2)
#     print("Val",val)
#
#
#     if val is not None:
#         extracted_data[col_name] = val
#
#     if not extracted_data:
#         print("Target extracted not found")
#         return  # neither target produced a value → nothing to insert
#
#     # --- 3) Ensure table/columns exist ---
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     needed_columns = list(extracted_data.keys())
#
#     if not table_exists:
#         cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
#                 ref TEXT
#                 {"," if cols_def else ""} {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         if "ref" not in existing_cols and newdatabase != maindatabase:
#             cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
#         for col in needed_columns:
#             if col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
#
#     # --- 4) Upsert (same rules as your working pattern) ---
#     if newdatabase != maindatabase:
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
#                 list(extracted_data.values()) + [mainID]
#             )
#         else:
#             cols = ["ref"] + list(extracted_data.keys())
#             placeholders = ", ".join(["?"] * len(cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
#                 [mainID] + list(extracted_data.values())
#             )
#     else:
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
#                 list(extracted_data.values()) + [mainID]
#             )
#         else:
#             cols = ["ID"] + list(extracted_data.keys())
#             placeholders = ", ".join(["?"] * len(cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
#                 [mainID] + list(extracted_data.values())
#             )

# def refdb_to_excel_source_right_test(
#     maindb,
#     main_ref,
#     linked_db,
#     link_ref,
#     column_to_get,    # base name, e.g. "source"
#     lookup_column,
#     lookup_value,
#     label_excel,
#     offset,
#     max_suffix=5     # how far to look for -1, -2, ..., -20
# ):
#     try:
#         # 1) Determine which columns exist: base + suffixes (-1, -2, ...)
#         # e.g. source, source-1, source-2 ...
#         pragma_sql = f"PRAGMA table_info([{linked_db}])"
#         cursor.execute(pragma_sql)
#         cols_info = cursor.fetchall()
#         all_cols = [row[1] for row in cols_info]  # row[1] is the column name in PRAGMA table_info
#
#         # Collect matching columns
#         matching_cols = []
#         if column_to_get in all_cols:
#             matching_cols.append(column_to_get)
#
#         # Check for column_to_get-1, column_to_get-2, ... up to max_suffix
#         for i in range(1, max_suffix + 1):
#             candidate = f"{column_to_get}-{i}"
#             if candidate in all_cols:
#                 matching_cols.append(candidate)
#
#         if not matching_cols:
#             print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
#             return
#
#         # 2) Build SELECT with all matching columns in order
#         select_list = ", ".join(f"a.[{col}]" for col in matching_cols)
#         query = f"""
#             SELECT {select_list}
#             FROM {linked_db} a
#             JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
#             WHERE c.{lookup_column} = ?
#             LIMIT 1
#         """
#
#         cursor.execute(query, (lookup_value,))
#         row = cursor.fetchone()
#         if not row:
#             print(f"No results found for {lookup_column} = {lookup_value}")
#             return
#
#         # 3) Find the label in the worksheet
#         for excel_row in ws_template.iter_rows():
#             for cell in excel_row:
#                 if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#                     start_row = cell.row
#                     start_col = cell.column + offset  # first cell to the right where we start writing
#
#                     # 4) Write SQL values horizontally to the right
#                     current_col = start_col
#                     for value in row:
#                         if value is not None and value != "":
#                             ws_template.cell(row=start_row, column=current_col).value = value
#                             print(
#                                 f"Inserted '{value}' into cell "
#                                 f"{ws_template.cell(row=start_row, column=current_col).coordinate}"
#                             )
#                         current_col += 1
#
#                     return  # done after first matching label
#
#         print(f"Label '{label_excel}' not found in worksheet.")
#
#     except sqlite3.Error as e:
#         print("SQLite error:", e)
#
#
# def refdb_to_excel_source_right(maindb, main_ref, linked_db, link_ref, column_to_get, lookup_column, lookup_value, label_excel,offset):
#
#     # Query the database for all matching values
#     try:
#         query = f"""
#              SELECT a.[{column_to_get}]
#              FROM {linked_db} a
#              JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
#              WHERE c.{lookup_column} = ?
#          """
#         cursor.execute(query, (lookup_value,))
#         results = cursor.fetchall()
#         if not results:
#             print(f"No results found for {lookup_column} = {lookup_value}")
#             return
#
#         # Find the label in the worksheet
#         for row in ws_template.iter_rows():
#             for cell in row:
#                 if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#                     start_row = cell.row
#                     col = cell.column + offset
#                     for i,result in enumerate(results):
#                         target_row = start_row  # insert to the right
#                         if result[0] != None:
#                             ws_template.cell(row=target_row, column=col).value = result[0]
#                             print(f"Inserted '{result[0]}' into cell {ws_template.cell(row=start_row, column=col).coordinate}")
#                     return
#
#         print(f"Label '{label_excel}' not found in worksheet.")
#     except sqlite3.Error as e:
#         print("SQLite error:", e)

#
# def add_info_right_two_markers_OECD_old(sheet, label1: str, label2: str, maindatabase, newdatabase, mainID, include_resource: bool = True):
#     """
#     label 1 - first row to match
#     label 2 - second row to match
#     1) Find a row where two adjacent cells match (label1, label2) left→right.
#     2) Capture first non-empty cell to the right of label2.
#     3) Write to SQL columns named: {label1}{label2}
#
#     Optional behavior (when include_resource=True):
#       - Captures the sheet's 'Resource' column value (if present and not empty)
#         into SQL column 'resource-<sanitized label2>' for the same row.
#       - If no 'Resource' column exists or the cell is empty, skips creating that column.
#     """
#
#     # --- helpers ---
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#     def matches(val, needle: str) -> bool:
#         if val is None:
#             return False
#         return needle.lower() in str(val).lower()
#
#     # Sanitize label for safe SQL column naming (for resource column)
#     def sanitize_label(s: str) -> str:
#         s = (s or "").strip().lower()
#         s = s.replace(" ", "-")
#         s = re.sub(r"[^a-z0-9_\-]", "", s)
#         s = re.sub(r"-{2,}", "-", s)
#         return s or "unnamed"
#
#     max_row = sheet.max_row
#     max_col = sheet.max_column
#
#     safe_label2 = sanitize_label(label2)
#     resource_colname = f"resource-{safe_label2}"
#
#     # --- 1) Find target row via adjacent (label1, label2) ---
#     target_row = None
#     for r in range(1, max_row + 1):
#         for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
#             v1 = sheet.cell(row=r, column=c).value
#             v2 = sheet.cell(row=r, column=c + 1).value
#             if matches(v1, label1) and matches(v2, label2):
#                 target_row = r
#                 break
#         if target_row is not None:
#             #print("Target row:", target_row)
#             break
#
#     if target_row is None:
#         print("Target row not found")
#         return  # no matching row → nothing to insert
#
#     # --- Optionally find the column index for "Resource" ---
#     resource_col = None
#     if include_resource:
#         for row in sheet.iter_rows():
#             for cell in row:
#                 if cell.value and str(cell.value).strip().lower() == "resource":
#                     resource_col = getattr(cell, "col_idx", cell.column)
#                     break
#             if resource_col:
#                 break
#
#     # --- 2) Scan the row to find targets; capture right-hand values ---
#     extracted_data = {}
#
#     # label → col name
#     col_name = f"{label1}{label2}"
#
#     # move to look for the first value to the right
#     def capture_right_of_label(row: int, label: str):
#         # Find the column containing the label
#         for c in range(1, max_col):
#             cell_value = sheet.cell(row=row, column=c).value
#             if matches(cell_value, label):
#
#                 # Start searching to the right of this column
#                 for cc in range(c + 1, max_col + 1):
#                     right_val = sheet.cell(row=row, column=cc).value
#
#                     # Skip empty or whitespace-only
#                     if right_val is None:
#                         continue
#
#                     if isinstance(right_val, str):
#                         rv = right_val.strip()
#                         if rv == "":
#                             continue
#                         return rv  # return first non-empty string
#
#                     # Non-string, non-None → return immediately
#                     return right_val
#
#                 # If no value was found to the right
#                 return None
#
#         # Label not found at all
#         return None
#
#     val = capture_right_of_label(target_row, label2)
#     #print("Val", val)
#
#     if val is not None:
#         extracted_data[col_name] = val
#
#     # Optionally add the Resource value, but only if it's not empty
#     if include_resource and resource_col:
#         resource_value = sheet.cell(row=target_row, column=resource_col).value
#         if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
#             extracted_data[resource_colname] = resource_value
#
#     if not extracted_data:
#         print("Target extracted not found")
#         return  # nothing to insert
#
#     # --- 3) Ensure table/columns exist ---
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     needed_columns = list(extracted_data.keys())
#
#     if not table_exists:
#         cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
#                 ref TEXT
#                 {"," if cols_def else ""} {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         if "ref" not in existing_cols and newdatabase != maindatabase:
#             cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
#         for col in needed_columns:
#             if col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
#
#     # --- 4) Upsert (same rules as your working pattern) ---
#     if newdatabase != maindatabase:
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
#                 list(extracted_data.values()) + [mainID]
#             )
#         else:
#             cols = ["ref"] + list(extracted_data.keys())
#             placeholders = ", ".join(["?"] * len(cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
#                 [mainID] + list(extracted_data.values())
#             )
#     else:
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
#                 list(extracted_data.values()) + [mainID]
#             )
#         else:
#             cols = ["ID"] + list(extracted_data.keys())
#             placeholders = ", ".join(["?"] * len(cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
#                 [mainID] + list(extracted_data.values())
#             )