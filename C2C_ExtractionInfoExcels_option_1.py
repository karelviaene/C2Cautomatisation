### This script goes to the CPS directory, checks all CAS files present and adds the info to the SQLite database.

#### SET UP ####
import sqlite3
import os
import re
import pandas as pd
from datetime import date
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from datetime import datetime
from copy import copy
from openpyxl import load_workbook, Workbook
from pathlib import Path
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.cell_range import CellRange

# Define the path to your SQLite database file
C2Cpath = "/Users/juliakulpa/Desktop/Test_excel_imports"
C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# C2Cfiles_path = os.path.join(C2Cpath)
db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")
# /Users/juliakulpa/Desktop/Test_excel_imports/Database/C2Cdatabase.db
print(db_path)
# # LOAD EXCEL CPS TEMPLATE
template_path = "/Users/juliakulpa/Desktop/Test_excel_imports/Template/CPS_CAS TEMPLATE V2.xlsm"
#template_wb = load_workbook(template_path, read_only=False, keep_vba=True)
#ws_template = template_wb["C2Coverview"]

database = '/Users/juliakulpa/Desktop/Test_excel_imports/Database /C2Cdatabase.db'
#CAS = ("108-31-6")
folder = '/Users/juliakulpa/Desktop/Test_excel_imports/Testing/'
image_dir = '/Users/juliakulpa/Desktop/test/Chem_image'


### CUSTOM FUNCTIONS ###

# def db_to_excel_one_below(table_name, column_to_get,lookup_column,lookup_value,label_excel):
#     # Query the database
#     query = f"SELECT [{column_to_get}] FROM {table_name} WHERE {lookup_column} = ?"
#     cursor.execute(query, (lookup_value,))
#     result = cursor.fetchone()
#     if not result:
#         print(f"No result found for {lookup_column} = {lookup_value}")
#         return
#     value_to_insert = result[0]
#     # Find the label in the worksheet and insert the value below it
#     for row in ws_template.iter_rows():
#         for cell in row:
#             if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#                 ws_template.cell(row=cell.row + 1, column=cell.column).value = value_to_insert
#                 print(f"Inserted '{value_to_insert}' below '{label_excel}' in cell {cell.coordinate}")
#                 return
#     print(f"Label '{label_excel}' not found in worksheet.")
#
# def db_to_excel_x_right(table_name, columns_to_get, lookup_column, lookup_value, labels_excel,offset):
#     # Ensure both inputs are lists
#     if isinstance(columns_to_get, str):
#         columns_to_get = [columns_to_get]
#     if isinstance(labels_excel, str):
#         labels_excel = [labels_excel]
#
#     # Query the database for multiple columns
#     query = f"SELECT {', '.join(f'[{col}]' for col in columns_to_get)} FROM {table_name} WHERE {lookup_column} = ?"
#     cursor.execute(query, (lookup_value,))
#     result = cursor.fetchone()
#     if not result:
#         print(f"No result found for {lookup_column} = {lookup_value}")
#         return
#     # Map column -> value from database
#     col_value_map = dict(zip(columns_to_get, result))
#
#     # Loop over each label and insert corresponding column value
#     for col_name, label_excel in zip(columns_to_get, labels_excel):
#         value_to_insert = col_value_map[col_name]
#         inserted = False
#         for row in ws_template.iter_rows():
#             for cell in row:
#                 if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#                     # Add the value in the column x (= offset) columns to the right
#                     ws_template.cell(row=cell.row , column=cell.column + offset).value = value_to_insert
#                     print(f"Inserted '{value_to_insert}' to the right of '{label_excel}' in cell {cell.coordinate}")
#                     inserted = True
#                     break
#             if inserted:
#                 break
#         if not inserted:
#             print(f"Label '{label_excel}' not found in worksheet.")
#


### Function

def extraction_info_excels(database, CAS, folder, image_dir):
    def db_to_excel_multiple_below(maindb, main_ref, linked_db, link_ref, column_to_get, lookup_column, lookup_value,
                                   label_excel):

        # Query the database for all matching values
        try:
            query = f"""
                 SELECT a.[{column_to_get}]
                 FROM {linked_db} a
                 JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                 WHERE c.{lookup_column} = ?
             """
            cursor.execute(query, (lookup_value,))
            results = cursor.fetchall()
            if not results:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            # Find the label in the worksheet
            for row in ws_template.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                        start_row = cell.row + 1
                        col = cell.column
                        # print(f"First test on{start_row, col}")

                        # Place each value in the first empty cell below the starting row
                        for result in results:
                            # If result is not none:

                            if result[0] != None:
                                # Start searching from start_row downward
                                target_row = start_row

                                # Keep moving down until we find an empty cell in the target column
                                while ws_template.cell(row=target_row, column=col).value not in (None, ''):
                                    target_row += 1

                                # print(f"target row{target_row}")

                                # Write the value in the first empty cell found
                                ws_template.cell(row=target_row, column=col).value = result[0]

                                print(
                                    f"Inserted '{result[0]}' into cell {ws_template.cell(row=target_row, column=col).coordinate}")

                        return

            print(f"Label '{label_excel}' not found in worksheet.")
        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_excel_source_right(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,    # base name, e.g. "source"
        lookup_column,
        lookup_value,
        label_excel,
        offset,
        max_suffix=5,    # how far to look for -1, -2 for the additional data
        include_resource=True # does it look for resources
    ):
        # helper: sanitize label like in add_info_CPS_right_until_empty
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]  # row[1] is column name

            # 2) Collect base + suffix columns for the main data (source, source-1, source-2, ...)
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                candidate = f"{column_to_get}-{i}"
                if candidate in all_cols:
                    matching_cols.append(candidate)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column name in SQL: resource-<sanitized label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(label_excel)
                candidate_resource_col = f"resource-{safe_label}"
                if candidate_resource_col in all_cols:
                    resource_sql_col = candidate_resource_col
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")

            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            # Split row into value columns and (optional) resource
            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # 5) Find the label in the worksheet
            for excel_row in ws_template.iter_rows():
                for cell in excel_row:
                    if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                        start_row = cell.row
                        start_col = cell.column + offset  # first column to write values

                        # 6) Write the main values horizontally to the right
                        current_col = start_col
                        for val in value_cols:
                            if val is not None and val != "":
                                ws_template.cell(row=start_row, column=current_col).value = val
                                print(
                                    f"Inserted '{val}' into cell "
                                    f"{ws_template.cell(row=start_row, column=current_col).coordinate}"
                                )
                            current_col += 1

                        # 7) If resource is present in SQL, put it in Excel "Resource" column
                        if include_resource and resource_val not in (None, ""):
                            resource_col_idx = None
                            # Find the "Resource" header column in the template
                            for hdr_row in ws_template.iter_rows():
                                for hdr_cell in hdr_row:
                                    if (
                                        isinstance(hdr_cell.value, str)
                                        and hdr_cell.value.strip().lower() == "resource"
                                    ):
                                        resource_col_idx = hdr_cell.column
                                        break
                                if resource_col_idx is not None:
                                    break

                            if resource_col_idx is not None:
                                ws_template.cell(row=start_row, column=resource_col_idx).value = resource_val
                                print(
                                    f"Inserted resource '{resource_val}' into cell "
                                    f"{ws_template.cell(row=start_row, column=resource_col_idx).coordinate}"
                                )
                            else:
                                print("Could not find 'Resource' column in Excel to write resource value.")

                        return  # done after first matching label

            print(f"Label '{label_excel}' not found in worksheet.")

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_column_names_unique(maindb, main_ref,linked_db, link_ref,lookup_column, lookup_value):
        """
        Returns a string with column names unique for each CAS, as a string
        """

        try:
            # Query EVERYTHING (*) from linked_db
            query = f"""
                SELECT a.*
                FROM {linked_db} a
                JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
            """

            cursor.execute(query, (lookup_value,))
            rows = cursor.fetchall()

            # If nothing found → return empty DataFrame (still safe)
            if not rows:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return pd.DataFrame()

            # Extract column names automatically from cursor.description
            colnames = [desc[0] for desc in cursor.description]

            dataframe = pd.DataFrame(rows, columns=colnames)

            # cutting columns with NULL values
            dataframe_cut = dataframe.dropna(axis=1, how='all')
            # dropping columns with ID and ref (not needed here)
            dataframe_cut = dataframe_cut.drop(columns=["ID", 'ref'])
            column_name = list(dataframe_cut.columns)
            # takes away from the string the resources names
            result = [c_name for c_name in column_name if "resource" not in c_name.lower() ]
            return result

        except sqlite3.Error as e:
            print("SQLite error:", e)
            return pd.DataFrame()

    def remove_text_from_string(string, target_name):
        '''removes text from string, used for Muta tests and SCL'''
        result = []
        for s in string:
            name = s.replace(target_name, "").strip()
            result.append(name)
        return result

    def write_list_right_of_label(ws_template: Worksheet, label_excel: str, offset: int, values_list: list):
        """
        Find the first cell whose value EXACTLY matches 'label_excel'
        (after stripping whitespace, case-insensitive),
        then write values from values_list to the right, moving downwards
        as long as the label cell below also exactly matches the label.
        """

        if not values_list:
            print("Value list is empty — nothing to write.")
            return

        # Normalize the target label once
        normalized_label = label_excel.strip().lower()

        # 1) Find the first exact match
        label_cell = None
        for row in ws_template.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == normalized_label:
                    label_cell = cell
                    break
            if label_cell:
                break

        if not label_cell:
            print(f"Exact label '{label_excel}' not found.")
            return

        start_row = label_cell.row
        label_col = label_cell.column
        target_col = label_col + offset
        max_row = ws_template.max_row

        current_row = start_row
        value_index = 0

        # 2) Write downward while exact match continues
        while value_index < len(values_list) and current_row <= max_row:
            current_cell_value = ws_template.cell(row=current_row, column=label_col).value

            # Check if current row still has EXACT match
            if not isinstance(current_cell_value, str) or \
               current_cell_value.strip().lower() != normalized_label:
                break

            ws_template.cell(row=current_row, column=target_col).value = values_list[value_index]

            print(
                f"Inserted '{values_list[value_index]}' into "
                f"{ws_template.cell(row=current_row, column=target_col).coordinate}"
            )

            value_index += 1
            current_row += 1

        if value_index < len(values_list):
            print(f"Warning: {len(values_list) - value_index} values not written — no more exact matching label rows.")

    def refdb_to_excel_source_after_two_targets(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,      # base name, e.g. "source"
        lookup_column,
        lookup_value,
        first_label_excel,  # label of first target cell
        second_label_excel, # label of second target cell (same row as first)
        max_suffix=5,       # how far to look for -1, -2, ... columns
        include_resource=True
    ):
        """
        Finds a row where both first_label_excel and second_label_excel are exact matches.
        Uses the rightmost of those as the "second target" and writes values starting at
        the first unmerged empty cell to its right.
        SQL resource column is: resource-<sanitized second_label_excel>.
        In Excel, resource always goes to the 'Resource' column.
        """
        #pre-step sanitizing labels
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]

            # 2) Collect base + suffix columns for the main data
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                cand = f"{column_to_get}-{i}"
                if cand in all_cols:
                    matching_cols.append(cand)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column in SQL: resource-<sanitized second_label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(second_label_excel)
                cand_res = f"resource-{safe_label}"
                if cand_res in all_cols:
                    resource_sql_col = cand_res
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")
            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # 5) Find a row where BOTH labels exist as exact cell matches
            row_found = False
            for excel_row in ws_template.iter_rows():
                first_col_idx = None
                second_col_idx = None

                for cell in excel_row:
                    if isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == str(first_label_excel).strip():
                            first_col_idx = cell.column
                        elif cell_val == str(second_label_excel).strip():
                            second_col_idx = cell.column

                if first_col_idx is not None and second_col_idx is not None:
                    row_found = True
                    target_row = excel_row[0].row
                    second_target_col = max(first_col_idx, second_col_idx)

                    # 6) Find the first *unmerged* empty cell to the right
                    current_col = second_target_col + 1
                    while True:
                        cell = ws_template.cell(row=target_row, column=current_col)

                        # Skip merged cells (non top-left)
                        if isinstance(cell, MergedCell):
                            current_col += 1
                            continue

                        if cell.value in (None, ""):
                            break

                        current_col += 1

                    # 7) Write values, skipping merged/occupied cells
                    for val in value_cols:
                        if val in (None, ""):
                            continue

                        while True:
                            cell = ws_template.cell(row=target_row, column=current_col)

                            if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                                current_col += 1
                                continue

                            cell.value = val
                            print(
                                f"Inserted '{val}' into cell {cell.coordinate}"
                            )
                            current_col += 1
                            break

                    # 8) Put resource into Excel "Resource" column (header cell == 'Resource')
                    if include_resource and resource_val not in (None, ""):
                        resource_col_idx = None
                        for hdr_row in ws_template.iter_rows():
                            for hdr_cell in hdr_row:
                                if (
                                    isinstance(hdr_cell.value, str)
                                    and hdr_cell.value.strip().lower() == "resource"
                                ):
                                    resource_col_idx = hdr_cell.column
                                    break
                            if resource_col_idx is not None:
                                break

                        if resource_col_idx is not None:
                            ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                            print(
                                f"Inserted resource '{resource_val}' into cell "
                                f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                            )
                        else:
                            print("Could not find 'Resource' column in Excel to write resource value.")
                    break

            if not row_found:
                print(
                    f"No row found where both '{first_label_excel}' "
                    f"and '{second_label_excel}' are present as exact matches."
                )

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_excel_source_after_two_targets_OECD(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,      # base name, e.g. "source"
        lookup_column,
        lookup_value,
        first_label_excel,  # label of first target cell
        second_label_excel, # label of second target cell (same row as first)
        max_suffix=5,       # how far to look for -1, -2, ... columns
        include_resource=True
    ):
        """
        Finds a row where both first_label_excel and second_label_excel are exact matches.
        Uses the rightmost of those as the "second target" and writes values starting at
        the first unmerged empty cell to its right.

        To work with the function getting info from SQL database, the No data could not be left as an empty cell as was before
        so it is written as "no data", however then the program to get the same looking template just has to skip it, not to
        write "do data" twice
        If second_label_excel == "No data" (case-insensitive):
          - Skip writing the SQL values for column_to_get and its suffixes.
          - Still write the SQL resource value (if available) into the Excel "Resource" column.
        """
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]

            # 2) Collect base + suffix columns for the main data
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                cand = f"{column_to_get}-{i}"
                if cand in all_cols:
                    matching_cols.append(cand)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column in SQL: resource-<sanitized second_label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(second_label_excel)
                cand_res = f"resource-{safe_label}"
                if cand_res in all_cols:
                    resource_sql_col = cand_res
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")
            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # Skipping no data: decide whether to skip writing the SQL values
            skip_value_write = str(second_label_excel).strip().lower() == "no data"

            # 5) Find a row where BOTH labels exist as exact cell matches
            row_found = False
            for excel_row in ws_template.iter_rows():
                first_col_idx = None
                second_col_idx = None

                for cell in excel_row:
                    if isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == str(first_label_excel).strip():
                            first_col_idx = cell.column
                        elif cell_val == str(second_label_excel).strip():
                            second_col_idx = cell.column

                if first_col_idx is not None and second_col_idx is not None:
                    row_found = True
                    target_row = excel_row[0].row
                    second_target_col = max(first_col_idx, second_col_idx)

                    # 6) Find the first *unmerged* empty cell to the right
                    current_col = second_target_col + 1
                    while True:
                        cell = ws_template.cell(row=target_row, column=current_col)

                        # Skip merged cells (non top-left)
                        if isinstance(cell, MergedCell):
                            current_col += 1
                            continue

                        if cell.value in (None, ""):
                            break

                        current_col += 1

                    # 7) Write values, skipping merged/occupied cells
                    # Only write values if NOT "No data"
                    if not skip_value_write:
                        for val in value_cols:
                            if val in (None, ""):
                                continue

                            while True:
                                cell = ws_template.cell(row=target_row, column=current_col)

                                if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                                    current_col += 1
                                    continue

                                cell.value = val
                                print(
                                    f"Inserted '{val}' into cell {cell.coordinate}"
                                )
                                current_col += 1
                                break
                    else:
                        print("Second label is 'No data' → skipping SQL values, but still handling resource if present.")

                    # 8) Put resource into Excel "Resource" column (header cell == 'Resource')
                    if include_resource and resource_val not in (None, ""):
                        resource_col_idx = None
                        for hdr_row in ws_template.iter_rows():
                            for hdr_cell in hdr_row:
                                if (
                                    isinstance(hdr_cell.value, str)
                                    and hdr_cell.value.strip().lower() == "resource"
                                ):
                                    resource_col_idx = hdr_cell.column
                                    break
                            if resource_col_idx is not None:
                                break

                        if resource_col_idx is not None:
                            ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                            print(
                                f"Inserted resource '{resource_val}' into cell "
                                f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                            )
                        else:
                            print("Could not find 'Resource' column in Excel to write resource value.")
                    break

            if not row_found:
                print(
                    f"No row found where both '{first_label_excel}' "
                    f"and '{second_label_excel}' are present as exact matches."
                )

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def insert_image_under_label(ws_template: Worksheet, label_excel: str, image_name: str,image_dir: str):
        """
        Find the first cell whose value exactly matches 'label_excel'
        then insert the image from image_dir/image_name into the cell
        directly below that label cell.
        """

        # Build image path
        image_path = Path(image_dir) / image_name

        if not image_path.exists():
            print(f"Image not found: {image_path}")
            return

        # Normalize the target label once
        normalized_label = label_excel.strip().lower()

        # 1) Find the first exact match for the label
        label_cell = None
        for row in ws_template.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == normalized_label:
                    label_cell = cell
                    break
            if label_cell:
                break

        if not label_cell:
            print(f"Exact label '{label_excel}' not found.")
            return

        # Cell directly below the label cell
        target_row = label_cell.row + 1
        target_col = label_cell.column
        anchor_coord = ws_template.cell(row=target_row, column=target_col).coordinate

        # 2) Create and place the image
        img = XLImage(str(image_path))
        img.anchor = anchor_coord
        ws_template.add_image(img)

        print(
            f"Inserted image '{image_name}' at {anchor_coord} "
            f"under label '{label_excel}' in sheet '{ws_template.title}'."
        )

    def put_template_into_CPS(filepath, template_path):
        '''Puts the template into the CPS excel that exists or creates a new one'''

        def open_or_create_xlsm(filepath):
            """
            Opens an existing .xlsm file with macros preserved.
            If it doesn't exist, creates a new one.
            Returns the workbook object.
            """
            if os.path.exists(filepath):
                # Load workbook and preserve macros
                wb = load_workbook(filepath, keep_vba=True)
                print(f"Opened existing file: {filepath}")
            else:
                # Create new workbook and save as xlsm
                wb = Workbook()
                wb.save(filepath)
                print(f"Created new xlsm file: {filepath}")
            return wb

        def add_new_sheet(filepath, new_sheet_name):
            """
            Opens (or creates) an xlsm file and adds a new sheet.
            Saves the file afterwards.
            """
            wb = open_or_create_xlsm(filepath)

            # If sheet already exists, create a unique name
            if new_sheet_name in wb.sheetnames:
                base = new_sheet_name
                i = 1
                while f"{base}_{i}" in wb.sheetnames:
                    i += 1
                new_sheet_name = f"{base}_{i}"

            # Create the sheet as first sheet
            ws = wb.create_sheet(new_sheet_name, 0)

            wb.save(filepath)
            print(f"Added sheet '{new_sheet_name}' to {filepath}")

            return ws

        def rename_with_date_and_move_to_back(filepath, sheet_name, date_format="%Y_%m_%d"):
            """
            Renames 'sheet_name' to 'sheet_name_YYYY_MM_DD'
            and moves it to the back of the workbook.
            """
            # Load workbook safely
            if filepath.lower().endswith(".xlsm"):
                wb = load_workbook(filepath, keep_vba=True)
            else:
                wb = load_workbook(filepath)

            # Ensure sheet exists
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' does not exist in workbook.")

            ws = wb[sheet_name]

            # Create date suffix
            date_str = datetime.now().strftime(date_format)

            # Build new name
            new_name = f"{sheet_name}_{date_str}"

            # Apply the new title
            ws.title = new_name

            # Move to the back
            wb._sheets.remove(ws)
            wb._sheets.append(ws)

            # Save file
            wb.save(filepath)

            print(f"Renamed '{sheet_name}' → '{new_name}' and moved to back.")

        def load_wb_any(path):
            """Load xlsx/xlsm, preserving VBA if present."""
            if path.lower().endswith(".xlsm"):
                return load_workbook(path, keep_vba=True)
            return load_workbook(path)

        def copy_sheet_to_other_workbook(
                src_path,
                src_sheet_name,
                dest_path,
                dest_sheet_name
        ):
            """
            Copy sheet `src_sheet_name` from src_path into dest_path
            as a new sheet called `dest_sheet_name`.

            Copies:
              - cell values
              - styles
              - merged cells
              - column widths
              - row heights
              - existing data validation (drop-down menus)
              - existing conditional formatting that applies to column I
            """

            # --- Load source workbook ---
            if not os.path.exists(src_path):
                raise FileNotFoundError(f"Source file not found: {src_path}")
            src_wb = load_wb_any(src_path)

            if src_sheet_name not in src_wb.sheetnames:
                raise ValueError(f"Sheet '{src_sheet_name}' not found in source workbook.")

            src_ws = src_wb[src_sheet_name]

            # --- Load or create destination workbook ---
            if os.path.exists(dest_path):
                dest_wb = load_wb_any(dest_path)
            else:
                from openpyxl import Workbook
                dest_wb = Workbook()
                # Clear the default sheet
                default_sheet = dest_wb.active
                dest_wb.remove(default_sheet)

            # If sheet with that name already exists in dest, remove it or rename first
            if dest_sheet_name in dest_wb.sheetnames:
                dest_wb.remove(dest_wb[dest_sheet_name])

            # Create destination sheet (as first sheet)
            dest_ws = dest_wb.create_sheet(title=dest_sheet_name, index=0)

            # --- Copy column dimensions (width, hidden, etc.) ---
            for col_letter, col_dim in src_ws.column_dimensions.items():
                new_dim = dest_ws.column_dimensions[col_letter]
                new_dim.width = col_dim.width
                new_dim.hidden = col_dim.hidden
                new_dim.outlineLevel = col_dim.outlineLevel
                new_dim.bestFit = col_dim.bestFit

            # --- Copy row dimensions (height, hidden, etc.) ---
            for row_idx, row_dim in src_ws.row_dimensions.items():
                new_row_dim = dest_ws.row_dimensions[row_idx]
                new_row_dim.height = row_dim.height
                new_row_dim.hidden = row_dim.hidden
                new_row_dim.outlineLevel = row_dim.outlineLevel

            # --- Copy cell values and styles ---
            for row in src_ws.iter_rows():
                for cell in row:
                    dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                    if cell.has_style:
                        dest_cell._style = copy(cell._style)

                    dest_cell.data_type = cell.data_type
                    dest_cell.number_format = cell.number_format
                    dest_cell.protection = copy(cell.protection)
                    dest_cell.alignment = copy(cell.alignment)

            # --- Copy merged cells ---
            if src_ws.merged_cells.ranges:
                for merged_range in src_ws.merged_cells.ranges:
                    dest_ws.merge_cells(str(merged_range.coord))

            # --- Copy existing data validation (drop-down menus) ---
            if src_ws.data_validations is not None:
                for dv in src_ws.data_validations.dataValidation:
                    new_dv = copy(dv)
                    dest_ws.add_data_validation(new_dv)

            # --- Copy conditional formatting only for column I ---
            col_I_idx = 9  # column I

            # Iterate internal CF structures similar to the SO snippet
            for cf in src_ws.conditional_formatting._cf_rules:
                for rng in cf.cells.ranges:
                    cr = CellRange(rng.coord)

                    # If the CF range covers column I
                    if cr.min_col <= col_I_idx <= cr.max_col:
                        # Intersect with column I: keep row range, force column I
                        dst_range = f"I{cr.min_row}:I{cr.max_row}"

                        for rule in cf.cfRule:
                            dest_ws.conditional_formatting.add(dst_range, copy(rule))

            # --- Save destination workbook ---
            dest_wb.save(dest_path)
            print(
                f"Copied sheet '{src_sheet_name}' from '{src_path}' "
                f"to '{dest_path}' as '{dest_sheet_name}' "
                f"(with data validation + CF for column I)."
            )

        # move the old C2C to the back and add a new C2C sheet in the front
        rename_with_date_and_move_to_back(filepath, "C2Coverview")
        add_new_sheet(filepath, "C2Coverview")
        # copy the template from the template
        copy_sheet_to_other_workbook(
            src_path=template_path,
            src_sheet_name="C2Coverview",
            dest_path=filepath,
            dest_sheet_name="C2Coverview"
        )

    ### Start with extracting
    # make a new sheet in the CAS specific folder
    print(f"Working on {CAS}")
    filepath = f"{folder}/Test CPS_CAS {CAS}.xlsm"
    put_template_into_CPS(filepath, template_path)

    template_wb = load_workbook(filepath, read_only=False, keep_vba=True)
    ws_template = template_wb["C2Coverview"]
    try:
        ### SQL SET-UP
        connection = sqlite3.connect(database)
        cursor = connection.cursor()

        print("Connected to SQLite database at:", db_path)

        # GENERAL INFO
        #Add general info
        namesDBcol_gen = ["Chemical name", "Common name", "CAS number", "EC number", "Linked CAS Read across",
                       "Linked CAS Monomers", "Linked CAS Degradation Products"]
        namesExcel_gen = ["Chemical name", "Common name", "CAS number", "EC number", "Linked CAS Read across",
                       "Linked CAS Monomers", "Linked CAS Degradation Products"]
        for namesDBcol_gen, namesExcel_gen in zip(namesDBcol_gen,namesExcel_gen):
            db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="GENERALINFO", link_ref="ref",
                                   column_to_get=namesDBcol_gen, lookup_column="ID",lookup_value =CAS, label_excel=namesExcel_gen)
        # ADD ASSESSORS
        db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="ASSESSORS", link_ref="ref",
                                   column_to_get="Name assessor", lookup_column="ID", lookup_value=CAS,
                                   label_excel="Name assessor")
        db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="ASSESSORS", link_ref="ref",
                                   column_to_get="Date assessed", lookup_column="ID", lookup_value=CAS,
                                   label_excel="Date created/updated")
        ## Add various info CHEMICAL CLASS
        namesDBcol_CC = ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]
        namesExcel_CC = ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]
        for names_DB, name_EX in zip(namesDBcol_CC, namesExcel_CC):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHEMICALCLASS", link_ref="ref",
                                        column_to_get=names_DB, lookup_column="ID", lookup_value=CAS,
                                        label_excel=name_EX, offset=2)

        # Adding other info
        namesDBcol_OTHER = ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]
        namesExcel_OTHER = ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]
        for names_DB_OTHER, name_EX_OTHER in zip(namesDBcol_OTHER, namesExcel_OTHER):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="OTHERINFO", link_ref="ref",
                                        column_to_get=names_DB_OTHER, lookup_column="ID", lookup_value=CAS,
                                        label_excel=name_EX_OTHER, offset=2)
        #  OTHER CRITERIA
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="OCRIT", link_ref="ref",
                                    column_to_get="Other comments", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Other comments", offset=1)
        # CARCINOGENICITY
        namesDBcols = ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK", "Carcinogenicity Classified IARC",
                        "Carcinogenicity Classified TLV","Carcinogenicity experimental evidence","Carcinogenicity Comments"]
        namesExcel = ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK", "Carcinogenicity Classified IARC",
                        "Carcinogenicity Classified TLV","Carcinogenicity experimental evidence","Carcinogenicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcols,namesExcel):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CARCINOGENICITY", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1,max_suffix=5,include_resource=True)

        # ED
        namesDBcols = ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]
        namesExcel = ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]
        for namesDBcol, nameExcel in zip(namesDBcols,namesExcel):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ENDOCRINE", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # MUTAGENICITY
        namesDBcol_MUT = ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]
        namesExcel_MUT = ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_MUT,namesExcel_MUT):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="MUTAGENICITY", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # MUTAGENICITY OECD TESTS
        #Point mutations
        point_mut_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT",
                                                       link_ref="ref",
                                                       lookup_column="ID", lookup_value=CAS)
        point_mut_names_cleared = remove_text_from_string(point_mut_names, "Point mutations:")
        #print(point_mut_names)

        write_list_right_of_label(ws_template, "Point mutations:", 1, point_mut_names_cleared)

        for namesDB, nameExcel in zip(point_mut_names, point_mut_names_cleared):
            refdb_to_excel_source_after_two_targets_OECD(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel="Point mutations:", second_label_excel=nameExcel, max_suffix=5, include_resource=True )

        # Chromosome damaging
        ch_dam_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM",
                                                    link_ref="ref",
                                                    lookup_column="ID", lookup_value=CAS)
        ch_dam_names_cleared = remove_text_from_string(ch_dam_names, "Chromosome damaging:")
        #print(ch_dam_names)
        write_list_right_of_label(ws_template, "Chromosome damaging:", 1, ch_dam_names_cleared)

        for namesDB, nameExcel in zip(ch_dam_names, ch_dam_names_cleared):
            refdb_to_excel_source_after_two_targets_OECD(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel="Chromosome damaging:", second_label_excel=nameExcel, max_suffix=5, include_resource=True )


        # REPROTOX
        namesDBcol_REP = ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]
        namesExcel_REP = ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_REP,namesExcel_REP):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="REPROTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # DEVELOPMENTAL TOX
        namesDBcol_DEV = ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]
        namesExcel_DEV = ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_DEV,namesExcel_DEV):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="DEVELOPTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # NEUROTOX
        namesDBcol_NETOX = ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]
        namesExcel_NETOX = ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_NETOX,namesExcel_NETOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="NEUROTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # ORAL TOX
        namesDBcol_ORTOX = ["Oral toxicity Acute Tox classified","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]
        namesExcel_ORTOX = ["Oral toxicity Acute Tox classified:","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_ORTOX,namesExcel_ORTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ORALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # INHALE TOX

        namesDBcol_INHTOX = ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]
        namesExcel_INHTOX = ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_INHTOX,namesExcel_INHTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="INHALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # DERMAL TOX
        namesDBcol_DERMTOX = ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]
        namesExcel_DERMTOX = ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_DERMTOX,namesExcel_DERMTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="DERMALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SKIN/EYE IRRIT/COR
        namesDBcol_IRR = ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]
        namesExcel_IRR = ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_IRR,namesExcel_IRR):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="IRRITCOR", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SENSITISATION
        namesDBcol_SENS = ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]
        namesExcel_SENS = ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_SENS,namesExcel_SENS):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="SENSITISATION", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SPECIFIC CONCENTRATION LIMITS
        SCL_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                 link_ref="ref",
                                                 lookup_column="ID", lookup_value=CAS)
        # cleaning the names so there is only distinct SCL names
        SCL_names_clean = remove_text_from_string(SCL_names, " - Lower Limit: (%)")
        SCL_names_clean = remove_text_from_string(SCL_names_clean, " - Upper Limit: (%)")
        SCL_names_dist = list(dict.fromkeys(SCL_names_clean))
        #print(SCL_names_dist)
        write_list_right_of_label(ws_template, "Hazard classification:", 1, SCL_names_dist)
        # # Lower limit
        # choosing SQL with only Lower limit data
        SCL_DB_names_lower = [s for s in SCL_names if "Lower Limit:" in s]
        print(SCL_DB_names_lower)
        SCL_EX_names_lower = remove_text_from_string(SCL_DB_names_lower, " - Lower Limit: (%)")
        print(SCL_EX_names_lower)
        # extracting Lower limit data
        for namesDB, nameExcel in zip(SCL_DB_names_lower, SCL_EX_names_lower):
            refdb_to_excel_source_after_two_targets(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel=nameExcel, second_label_excel="Lower Limit: (%)", max_suffix=5, include_resource=True )

        # # Upper limit
        # choosing SQL with only Lower limit data
        SCL_DB_names_upper = [s for s in SCL_names if "Upper Limit:" in s]
        SCL_EX_names_upper = remove_text_from_string(SCL_DB_names_upper, " - Upper Limit: (%)")
        # extracting Lower limit data
        for namesDB, nameExcel in zip(SCL_DB_names_upper, SCL_EX_names_upper):
            refdb_to_excel_source_after_two_targets(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                    link_ref="ref",
                                                    column_to_get=namesDB, lookup_column="ID", lookup_value=CAS,
                                                    first_label_excel=nameExcel, second_label_excel="Upper Limit: (%)",
                                                    max_suffix=5, include_resource=True)

        # AQUATIC TOXICITY

        namesDBcol_AQTOX = ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]
        namesExcel_AQTOX = ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]
        for namesDBcol, nameExcel in zip(namesDBcol_AQTOX,namesExcel_AQTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="AQUATOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # VERTEBRATE FISH
        namesDBcol_FISHTOX = ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]
        namesExcel_FISHTOX = ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_FISHTOX,namesExcel_FISHTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="FISHTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # INVERTEBRATE TOX
        namesDBcol_INVTOX = ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]
        namesExcel_INVTOX = ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_INVTOX,namesExcel_INVTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="INVTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # ALGAE TOX
        namesDBcol_ALGTOX = ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]
        namesExcel_ALGTOX = ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]
        for namesDBcol, nameExcel in zip(namesDBcol_ALGTOX, namesExcel_ALGTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ALGAETOX", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)

        # TERRESTRIAL TOX
        namesDBcol_TERTOX =  ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]
        namesExcel_TERTOX =  ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_TERTOX, namesExcel_TERTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="TERTOX", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)

        # OTHER SPECIES TOX
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="SPECTOX", link_ref="ref",
                                    column_to_get="Any other CLP species classification", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Any other CLP species classification", offset=1)
        # PERSISTENCE
        namesDBcol_PERS =  ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]
        namesExcel_PERS =  ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_PERS, namesExcel_PERS):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="PERSISTENCE", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        # BIOACCUMULATION
        namesDBcol_BIOAC =  ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]
        namesExcel_BIOAC =  ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_BIOAC, namesExcel_BIOAC):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="BIOACCUMULATION", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        # CLIMATIC RELEVANCE
        namesDBcol_CLIMREL =  ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]
        namesExcel_CLIMREL =  ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_CLIMREL, namesExcel_CLIMREL):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CLIMATICRELEVANCE", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        #  ADDITIONAL SOURCES
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ADDSOURCE", link_ref="ref",
                                    column_to_get="Additional sources", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Additional sources", offset=1)

        ### Image
        ex_label = "Molecular Formula or chemical picture"
        image_name = f"CPS_CAS {CAS}.png"

        insert_image_under_label(ws_template, ex_label, image_name, image_dir)

        ### SAVE THE FILLED IN CPS EXCEL ####
        name = f"Test CPS_CAS {CAS}.xlsm"
        saving_path = os.path.join(folder,name)
        template_wb.save(saving_path)

    except sqlite3.Error as e:
        print("SQLite error:", e)

CAS_numbers= ["10-00-0","50-00-0","108-31-6","110-54-3","1592-23-0"]
for CAS in CAS_numbers:
    extraction_info_excels(database, CAS, folder, image_dir)

# extraction_info_excels(database, "50-00-0", ws_template, folder, image_dir)
extraction_info_excels(database, "108-31-6", folder, image_dir)
# extraction_info_excels(database, "50-00-0", ws_template, folder, image_dir)