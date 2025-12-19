### This script goes to the CPS directory, checks all CAS files present and adds the info to the SQLite database.

#### SET UP ####
import sqlite3
import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
import traceback

# Define the path to your SQLite database file
C2Cpath = "/Users/juliakulpa/Desktop/test"
C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# C2Cfiles_path = os.path.join(C2Cpath)
db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")

# Format today’s date
today = datetime.today().strftime("%Y%m%d")  # e.g., 20250903
# Specify whether you want to load-in the CPS files or whether you want to start from a preloaded database
READ_IN_CPS = True

#### CUSTOM FUNCTIONS ####
# Custom function for extracting info from Excel: it takes all the info below a certain cell until an empty string is reached
# It then adds the info to a new SQL database connected to the main database
def add_info_CPS_below(sheet, search_strings, maindatabase, newdatabase, mainID):
    """
    Finds each search string somewhere in the sheet (not necessarily in the first row),
    then reads the values directly below it column-wise.
    Collects one "row" of data across all search strings,
    and inserts into the SQLite database.

    Stops when *all* searched columns are empty in the same row.

    search_strings:
        - list: Excel labels are used as SQL column names.
        - dict: {Excel label: SQL column name}
    """

    # Handle both list and dict
    if isinstance(search_strings, str):
        search_strings = [search_strings]
    if isinstance(search_strings, list):
        mapping = {s: s for s in search_strings}  # Excel label -> same SQL name
    elif isinstance(search_strings, dict):
        mapping = search_strings  # Excel label -> custom SQL name
    else:
        raise TypeError("search_strings must be a str, list, or dict")

    # Locate the target cells (positions of the labels in Excel)
    col_positions = {}
    for excel_label in mapping.keys():
        found = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and excel_label.lower() in str(cell.value).lower():
                    found = cell
                    break
            if found:
                break
        if found:
            col_positions[excel_label] = (found.column, found.row + 1)  # start below label

    # Quote identifiers for SQL safety
    def q(name: str) -> str:
        return f'"{name}"'

    # Check if the table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                   (newdatabase,))
    table_exists = cursor.fetchone()

    if not table_exists:
        # Create table with auto-increment ID and extracted columns
        cols_def = ", ".join([f"{q(sql_col)} TEXT" for sql_col in mapping.values()])
        fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
        cursor.execute(f'''
            CREATE TABLE {q(newdatabase)} (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                ref TEXT
                {"," if cols_def else ""} {cols_def}
                {fk_clause}
            )
        ''')
    else:
        # Add missing columns if needed
        cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
        existing_cols = [col[1] for col in cursor.fetchall()]
        for sql_col in mapping.values():
            if sql_col not in existing_cols:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(sql_col)} TEXT")

    # Iterate row by row until all searched columns are empty
    row_offset = 0
    while True:
        row_data = {}
        all_empty = True
        for excel_label, (col, start_row) in col_positions.items():
            cell = sheet.cell(row=start_row + row_offset, column=col)
            if cell.value is not None and str(cell.value).strip() != "":
                sql_col = mapping[excel_label]
                row_data[sql_col] = cell.value
                all_empty = False
        if all_empty:
            break  # stop when all searched columns are empty in the same row

        if newdatabase == maindatabase:
            # --- Case 1: Append into the main database table ---
            # Check if the row for this ID exists
            cursor.execute(
                f"SELECT 1 FROM {q(maindatabase)} WHERE ID = ?",
                (mainID,)
            )
            exists = cursor.fetchone()

            if exists:
                # Update only the new columns (append data)
                update_clause = ", ".join([f"{q(col)} = ?" for col in row_data.keys()])
                cursor.execute(
                    f"UPDATE {q(maindatabase)} SET {update_clause} WHERE ID = ?",
                    list(row_data.values()) + [mainID]
                )
            else:
                # Insert new row with ID and these values
                all_cols = ["ID"] + list(row_data.keys())
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(maindatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + list(row_data.values())
                )

        else:
            # --- Case 2: Insert into child table ---
            row_data["ref"] = mainID

            # Build WHERE clause dynamically based on all row_data keys
            where_clause = " AND ".join([f"{q(col)} = ?" for col in row_data.keys()])
            params = list(row_data.values())

            cursor.execute(
                f"SELECT 1 FROM {q(newdatabase)} WHERE {where_clause}",
                params
            )
            exists = cursor.fetchone()

            if not exists:
                # Insert new row
                all_cols = list(row_data.keys())
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    list(row_data.values())
                )

        row_offset += 1

def add_info_CPS_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
    """
    Finds rows containing `rowlabel`, extracts specified columns to the right,
    plus the value in the "Resource" column of the same row,
    and inserts or updates the data in the SQLite database.

    Parameters:
        sheet: openpyxl worksheet
        rowlabel: string to search for in any row
        column_offsets: list of integers (e.g., [2, 3]) for columns to the right
        column_names: list of strings for custom SQL column names
        maindatabase: name of the main database (for foreign key reference)
        newdatabase: name of the table to update
        mainID: unique identifier for the row
    """

    if len(column_offsets) != len(column_names):
        raise ValueError("column_offsets and column_names must have the same length")

    # Quote identifiers for SQL safety
    def q(name: str) -> str:
        return f'"{name}"'

    # Step 1a: Find the column index for "Resource"
    resource_col = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "resource":
                resource_col = cell.column
                break
        if resource_col:
            break

    # Step 1b: Find the row containing the rowlabel
    extracted_data = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                row_idx = cell.row
                col_idx = cell.column
                for offset, col_name in zip(column_offsets, column_names):
                    target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                    extracted_data[col_name] = target_cell.value
                # Always try to grab Resource from the same row
                extracted_data["Resource"] = (
                    sheet.cell(row=row_idx, column=resource_col).value if resource_col else None
                )
                break
        if extracted_data:
            break

    if not extracted_data:
        return  # nothing to insert

    # Step 2: Check if table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
    table_exists = cursor.fetchone()

    if not table_exists:
        # Create table with ID, ref, and extracted columns
        cols_def = ", ".join([f"{q(col)} TEXT" for col in column_names + ["Resource"]])
        fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
        cursor.execute(f'''
            CREATE TABLE {q(newdatabase)} (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                ref TEXT
                {"," if cols_def else ""} {cols_def}
                {fk_clause}
            )
        ''')
    else:
        # Add missing columns
        cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
        existing_cols = [col[1] for col in cursor.fetchall()]
        if "ref" not in existing_cols and newdatabase != maindatabase:
            cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
        for col in column_names + ["Resource"]:
            if col not in existing_cols:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

    # Step 3: Insert or update
    if newdatabase != maindatabase:
        cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
        exists = cursor.fetchone()
        if exists:
            update_clause = ", ".join([f"{q(col)} = ?" for col in column_names + ["Resource"]])
            cursor.execute(
                f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                [extracted_data[col] for col in column_names + ["Resource"]] + [mainID]
            )
        else:
            all_cols = ['ref'] + column_names + ["Resource"]
            placeholders = ", ".join(["?"] * len(all_cols))
            cursor.execute(
                f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                [mainID] + [extracted_data[col] for col in column_names + ["Resource"]]
            )

    else:  # when newdatabase == maindatabase
        cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
        exists = cursor.fetchone()
        if exists:
            update_clause = ", ".join([f"{q(col)} = ?" for col in column_names + ["Resource"]])
            cursor.execute(
                f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                [extracted_data[col] for col in column_names + ["Resource"]] + [mainID]
            )
        else:
            all_cols = ['ID'] + column_names + ["Resource"]
            placeholders = ", ".join(["?"] * len(all_cols))
            cursor.execute(
                f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                [mainID] + [extracted_data[col] for col in column_names + ["Resource"]]
            )

def add_all_info_CPS_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
    """
    Scans all cells in the Excel sheet for `rowlabel`, extracts specified columns to the right,
    plus the 'Resource' column (if present), and inserts or updates each match into the SQLite database.

    Parameters:
        sheet: openpyxl worksheet
        rowlabel: string to search for in any cell
        column_offsets: list of integers (e.g., [2, 3]) for columns to the right
        column_names: list of strings for custom SQL column names
        maindatabase: name of the main database (for foreign key reference)
        newdatabase: name of the table to update
        mainID: foreign key reference to maindatabase
    """

    if len(column_offsets) != len(column_names):
        raise ValueError("column_offsets and column_names must have the same length")

    def q(name: str) -> str:
        return f'"{name}"'

    # Step 1: Find the column index for "Resource"
    resource_col = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "resource":
                resource_col = cell.column
                break
        if resource_col:
            break

    # Step 2: Prepare the table
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
    table_exists = cursor.fetchone()

    all_columns = column_names + ["Resource"]

    if not table_exists:
        cols_def = ", ".join([f"{q(col)} TEXT" for col in all_columns])
        fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
        cursor.execute(f'''
            CREATE TABLE {q(newdatabase)} (
                ID TEXT PRIMARY KEY,
                ref TEXT,
                {cols_def}
                {fk_clause}
            )
        ''')
    else:
        cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
        existing_cols = [col[1] for col in cursor.fetchall()]
        for col in all_columns:
            if col not in existing_cols:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

    # Step 3: Scan all cells and process matches
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                row_idx = cell.row
                col_idx = cell.column

                # Extract custom columns
                extracted_data = {}
                for offset, col_name in zip(column_offsets, column_names):
                    target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                    extracted_data[col_name] = target_cell.value

                # Extract Resource if available
                extracted_data["Resource"] = (
                    sheet.cell(row=row_idx, column=resource_col).value if resource_col else None
                )

                # Build unique ID (row_idx_col_idx ensures uniqueness)
                newID = f"{row_idx}_{col_idx}"

                # Check if row already exists
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (newID,))
                exists = cursor.fetchone()

                if exists:
                    update_clause = ", ".join([f"{q(col)} = ?" for col in all_columns])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                        [extracted_data[col] for col in all_columns] + [newID]
                    )
                else:
                    all_cols = ['ID', 'ref'] + all_columns
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        [newID, mainID] + [extracted_data[col] for col in all_columns]
                    )

# def add_info_checkstring(table_name,id_column,mainID,search_column,search_string,update_column,update_string):
#     """
#     Checks if `search_string` exists in `search_column` for a given `mainID`.
#     If found, updates `update_column` with `update_string`.
#
#     Parameters:
#         db_path (str): Path to the SQLite database file
#         table_name (str): Name of the table to query
#         id_column (str): Name of the ID column (e.g., "ID")
#         mainID (str): The ID value to look for
#         search_column (str): Column to search for the string
#         search_string (str): String to search for
#         update_column (str): Column to update
#         update_string (str): String to insert if match is found
#     """
#
#     # Get existing columns
#     cursor.execute(f"PRAGMA table_info({table_name})")
#     existing_columns = [col[1] for col in cursor.fetchall()]
#
#     # Add missing columns if needed
#     if update_column not in existing_columns:
#             cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN '{update_column}' TEXT")
#
#     # Check if the string exists for the given ID
#     cursor.execute(f"""
#         SELECT 1 FROM {table_name}
#         WHERE {id_column} = ? AND {search_column} LIKE ?
#     """, (mainID, f"%{search_string}%"))
#
#     if cursor.fetchone():
#         # Update the target column
#         cursor.execute(f"""
#             UPDATE {table_name}
#             SET {update_column} = ?
#             WHERE {id_column} = ?
#         """, (update_string, mainID))
#

#### Create/update C2C database with CAS numbers from Excel files ####

if READ_IN_CPS == True:
    try:
        ### SQL SET-UP
        connection = sqlite3.connect(db_path)
        print("Connected to SQLite database at", db_path)

        # Create main C2C_DATABASE table if not existing
        cursor = connection.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                ID TEXT PRIMARY KEY,
                LastUpdate TEXT,
                FileName TEXT,
                Comments TEXT
            )
        ''')

        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.xlsx')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.xlsx', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        # Loop through Excel files with CAS number and add their info from the template
        for filename in os.listdir(C2Cfiles_path):
            full_path = os.path.join(C2Cfiles_path, filename)
            if os.path.isfile(full_path):
                match = file_pattern.search(filename)
                if match:
                    print(filename)
                    #### Update database: general info #####
                    # Get last modification time and format it as DD/MM/YYYY
                    mod_time = os.path.getmtime(full_path)
                    last_update = datetime.fromtimestamp(mod_time).strftime("%d/%m/%Y")

                    # Extract CAS number or EC number if applicable
                    match_inv = cas_pattern.search(filename)    # Check for CAS number
                    comments = "There should be something here. Please check."
                    if match_inv:     # If CAS
                        comments = "CAS"
                        if match_inv.group(2):  # If there is additional info after the CAS number, save it in comments
                            comments = "CAS, " +  match_inv.group(2)

                    else:   # if no CAS, then check for EC
                        match_inv = ec_pattern.search(filename)
                        comments = "EC"
                    if match_inv:   # If a standard format is found, save for use in the database
                        inv_number = match_inv.group(1)
                    else:   # Else print the file for which there is an issue
                        print(filename)


                    #### Update database: extract info on CAS from file if new/more recent #####
                    cursor.execute('SELECT 1 FROM C2C_DATABASE WHERE ID = ? AND LastUpdate > ?', (inv_number, last_update))
                    exists = cursor.fetchone()
                    if not exists:  # If there is no info or there is more recent info
                        # Update general entry in the database
                        cursor.execute(
                            'INSERT OR REPLACE INTO C2C_DATABASE (ID, LastUpdate, FileName , Comments) VALUES (?,?,?,?)',
                            (inv_number, last_update, filename, comments))

                        # Open the Excel file
                        CPS_wb_obj = openpyxl.load_workbook(full_path)
                        CPSsheet = CPS_wb_obj.active

                        # CHEMICAL NAME
                        add_info_CPS_below(CPSsheet, ["Chemical name"],"C2C_DATABASE","CHEMICALNAMES",inv_number)

                        # ASSESSOR
                        add_info_CPS_below(CPSsheet, {"Name assessor":"Name assessor","Date created/updated" : "Date assessed"},"C2C_DATABASE","ASSESSORS",inv_number)

                        # CHECKED status
                        add_info_CPS_below(CPSsheet, ["Checked"],"C2C_DATABASE","C2C_DATABASE",inv_number)

                        # Add various info
                        for info in ["Common name","Organohalogens","Toxic metal", "Colourant", "Geological", "Biological", "Polymer"]:
                            add_info_CPS_right(CPSsheet,info,[2],[info],
                                "C2C_DATABASE","C2C_DATABASE",inv_number)
                        # # Add SMILES
                        for info in ["SMILES"]:
                            add_info_CPS_right(CPSsheet,info,[1],[info],
                                "C2C_DATABASE","C2C_DATABASE",inv_number)
                        # # # Molecular formula
                        add_info_CPS_below(CPSsheet, ["Molecular Formula"], "C2C_DATABASE", "C2C_DATABASE", inv_number)

                        # CARCINOGENICITY
                        for carc_type in ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK","Carcinogenicity Classified IARC",
                            "Carcinogenicity Classified TLV", "Carcinogenicity Comments"]:
                            add_info_CPS_right(CPSsheet,carc_type,[1],[carc_type],
                                "C2C_DATABASE","CARCINOGENICITY",inv_number)

                        # ED
                        for ED_type in ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]:
                            add_info_CPS_right(CPSsheet,ED_type,[1],[ED_type],
                                "C2C_DATABASE","ENDOCRINE",inv_number)

                        # MUTAGENICITY/GENOTOXICITY
                        for muta_type in ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]:
                            add_all_info_CPS_right(CPSsheet,muta_type,[1],[muta_type],
                                "C2C_DATABASE","MUTAGENICITY",inv_number)
                        # For the strings with multiple possible hits
                        for muta_type in ["OECD 471", "OECD 473", "OECD 474", "OECD 475", "OECD 476", "OECD 478",
                            "OECD 483", "OECD 485", "OECD 486", "OECD 487", "OECD 488", "OECD 489", "OECD 490", "#No data"]:
                            add_all_info_CPS_right(CPSsheet,muta_type,[3],[muta_type],
                                "C2C_DATABASE","MUTAGENICITY",inv_number)

                        # REPRODUCTIVE TOXICITY
                        for repro_type in ["Reprotox Classified CLP", "Reprotox Classified MAK", "Oral NOAEL =",
                                           "Inhalation NOAEL =", "Reproductive Toxicity Comments"]:
                            add_info_CPS_right(CPSsheet,repro_type,[1],[repro_type],
                                "C2C_DATABASE","REPROTOX",inv_number)

                        # DEVELOPMENTAL TOXICITY
                        for devo_type in ["Developmental Classified CLP", "Developmental Classified MAK", "Oral NOAEL =",
                                           "Inhalation NOAEL =", "Developmental Toxicity Comments"]:
                            add_info_CPS_right(CPSsheet,devo_type,[1],[devo_type],
                                "C2C_DATABASE","DEVELOPTOX",inv_number)

                        # ORAL TOXICITY
                        for oral_type in ["Oral toxicity Acute Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                            "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]:
                            add_info_CPS_right(CPSsheet,oral_type,[1],[oral_type],
                                "C2C_DATABASE","ORALTOX",inv_number)

                        # INHALATIVE TOXICITY
                        for inhal_type in ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                            "Acute: LC50 (gas) =", "Acute: LC50 (vapor) =", "Acute: LC50 (dust/mist/aerosol) =", "Chronic: LOAEL (gas) =",
                            "Chronic: LOAEL (vapor) =", "Chronic: LOAEL (dust/mist/aerosol) =", "Boiling Point", "Inhalative Toxicity Comments"]:
                            add_info_CPS_right(CPSsheet,inhal_type,[1],[inhal_type],
                                "C2C_DATABASE","INHALTOX",inv_number)

                        # DERMAL TOXICITY
                        for dermal_type in ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                            "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]:
                            add_info_CPS_right(CPSsheet,dermal_type,[1],[dermal_type],
                                "C2C_DATABASE","DERMALTOX",inv_number)

                        # NEUROTOXICITY
                        for neuro_type in ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                            "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]:
                            add_info_CPS_right(CPSsheet,neuro_type,[1],[neuro_type],
                                "C2C_DATABASE","NEUROTOX",inv_number)

                        # SKIN/EYE IRRITATION/CORROSION
                        for irrit_type in ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                            "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion"]:
                            add_info_CPS_right(CPSsheet,irrit_type,[1],[irrit_type],
                                "C2C_DATABASE","IRRITCOR",inv_number)

                        # SENSITISATION
                        for sens_type in ["Skin sensitization classification", "Skin sensitization testing conclusion",
                            "Skin sensitization classified MAK", "Inhalation sensitization classification",
                            "Inhalation sensitization testing conclusion", "Inhalation sensitization classified MAK"]:
                            add_info_CPS_right(CPSsheet,sens_type,[1],[sens_type],
                                "C2C_DATABASE","SENSITISATION",inv_number)

                        # AQUATIC TOXICITY
                            # VERTEBRATE
                        for fish_type in ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC ="]:
                            add_info_CPS_right(CPSsheet,fish_type,[1],[fish_type],
                                "C2C_DATABASE","FISHTOX",inv_number)
                            # INVERTEBRATE
                        for inv_type in ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC ="]:
                            add_info_CPS_right(CPSsheet, inv_type, [1], [inv_type],
                                               "C2C_DATABASE", "INVTOX", inv_number)
                            # ALGAE
                        for algae_type in ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC ="]:
                            add_info_CPS_right(CPSsheet, algae_type, [1], [algae_type],
                                               "C2C_DATABASE", "ALGAETOX", inv_number)
                            # General aquatic tox
                            for aqtox_type in ["Aquatic toxicity CLP classification","Water solubility", "M factor: "]:
                                add_info_CPS_right(CPSsheet, aqtox_type, [1], [aqtox_type],
                                                   "C2C_DATABASE", "AQUATOX", inv_number)

                        # TERRESTRIAL TOXICITY
                        for tertox_type in ["Terrestial toxicity CLP classification"]:
                            add_info_CPS_right(CPSsheet, tertox_type, [1], [tertox_type],
                                               "C2C_DATABASE", "TERTOX", inv_number)

                        # PERSISTENCE
                        for pers_type in ["% DOC biodegradation after 28 days", "% ThOD biodegradation after 28 days",
                            "Degradation halflife time in air", "Degradation halflife time in water", "soil or sediment", "QSAR prediction"]:
                            add_info_CPS_right(CPSsheet,pers_type,[1],[pers_type],
                                "C2C_DATABASE","PERSISTENCE",inv_number)

                        # BIOACCUMULATION
                        for bioac_type in ["BCF/BAF (QSAR)", "BCF/BAF (experimental)", "Log kow", "Molecular weight"]:
                            add_info_CPS_right(CPSsheet,bioac_type,[1],[bioac_type],
                                "C2C_DATABASE","BIOACCUMULATION",inv_number)

                        # CLIMATIC RELEVANCE
                        for clima_type in ["Climatic relevance/ozone depletion potential"]:
                            add_info_CPS_right(CPSsheet,clima_type,[2],[clima_type],
                                "C2C_DATABASE","CLIMATICREL",inv_number)

                        # PHYSICAL PROPERTIES
                        for physchem_type in ["VOC designation", "Molecular weight", "Boiling point", "Log kow (octanol-water partition coefficient)",
                            "Vapor pressure", "Water solubility"]:
                            add_info_CPS_right(CPSsheet,physchem_type,[1],[physchem_type],
                                "C2C_DATABASE","PHYSCHEM",inv_number)

        connection.commit()
        print("SQL updated")

    except sqlite3.Error as e:
        print("SQLite error", e, inv_number)

    finally:
        if connection:

            # Create new database copy with date of today (master database is also updated)
            # Define new filename
            backup_path = db_path.replace(".db", f"_{today}.db")
            # Create backup connection
            backup_conn = sqlite3.connect(backup_path)
            # Perform the backup
            with backup_conn:
                connection.backup(backup_conn)
                print("Backup made: " + backup_path)
            backup_conn.close()

            connection.close()
            print("Connection closed.")



#### Save Database as Excel ####
try:
    ### SQL SET-UP
    connection = sqlite3.connect(db_path)
    print("Reconnected to database to export to Excel", db_path)

    # Get a list of all tables in the database
    tables_df = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';", connection)
    tables = tables_df['name'].tolist()
    tables = [t for t in tables if t not in ("C2C_DATABASE", "ECHACHEM_CL")] # Exclude the main table from the aggregation
    # Also excludes ECHA-CHEM as that datatable is not connected through ref
    cte_list = []
    join_list = []
    main_table = "C2C_DATABASE"

    # AGGREGATE ALL DEPENDENT TABLES SO THAT WE CAN EASILY USE LEFTJOIN
    for table in tables:
        cursor = connection.cursor()
        cursor.execute(f"PRAGMA table_info({table})")
        columns_info = cursor.fetchall()
        cols = [col[1] for col in columns_info if col[1] not in ('ref', 'ID')]

        if not cols or 'ref' not in [col[1] for col in columns_info]:
            continue  # skip tables without 'ref'

        # Build GROUP_CONCAT for each column
        group_concat_cols = ",\n       ".join([f"GROUP_CONCAT([{col}], CHAR(10)) AS [{col}]" for col in cols])

        # Define CTE
        cte_name = f"{table}Agg"
        cte = f"{cte_name} AS (\n    SELECT ref,\n           {group_concat_cols}\n    FROM {table}\n    GROUP BY ref\n)"
        cte_list.append(cte)

        # Prepare LEFT JOIN
        join_list.append(f"LEFT JOIN {cte_name} {table[:4].lower()} ON {main_table}.ID = {table[:4].lower()}.ref")

    # Combine all CTEs
    cte_sql = ",\n".join(cte_list)

    # Select all columns from main_table + aggregated tables
    cursor.execute(f"PRAGMA table_info({main_table})")
    main_cols = [f"{main_table}.[{col[1]}]" for col in cursor.fetchall()]

    select_cols = main_cols.copy()
    for table in tables:
        cursor.execute(f"PRAGMA table_info({table})")
        for col in cursor.fetchall():
            if col[1] not in ('ref',"ID"):
                select_cols.append(f"{table[:4].lower()}.[{col[1]}]")

    select_sql = ",\n    ".join(select_cols)
    join_sql = "\n".join(join_list)

    final_query = f"WITH {cte_sql}\nSELECT\n    {select_sql}\nFROM {main_table}\n{join_sql};"

    df = pd.read_sql_query(final_query, connection)

    # Make sure everything is readable for Excel
    def sanitize_excel_cells(df):
        def fix_value(val):
            if isinstance(val, str):
                # If it starts with '=', prefix with a single quote to force Excel to treat it as text
                if val.strip().startswith("="):
                    return "'" + val
                # Remove control characters that break XML
                return ''.join(ch for ch in val if ch.isprintable())
            return val

        for col in df.columns:
            df[col] = df[col].map(fix_value)
        return df


    # Clean both headers and cell values
    # df.columns = clean_excel_headers(df.columns)  # from earlier
    df = sanitize_excel_cells(df)

    # Write away the database as Excel
    with pd.ExcelWriter(C2Cpath + '/Database/C2Cdatabase ' + today + '.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="C2C DATABASE", index=False)
    print("Database exported to Excel")


except sqlite3.Error as e:
    # Catches SQLite-specific errors
    print("SQLite error:", e)
    traceback.print_exc()  # prints the full traceback

except Exception as e:
    # Catches other Python errors
    print("General error:", e)
    traceback.print_exc()

finally:
    # Always close connection if it was created
    try:
        connection.close()
        print("Connection closed.")
    except NameError:
        pass  # connection was never created