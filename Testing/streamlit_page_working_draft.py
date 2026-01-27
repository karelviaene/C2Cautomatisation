import streamlit as st
import traceback
import requests
import openpyxl
import time
import json
import copy
import sqlite3
import os
import re
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from copy import copy
from openpyxl import load_workbook, Workbook
from pathlib import Path
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.cell_range import CellRange
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from streamlit import empty

st.title("ARCHE CAS database")
# text
st.markdown(
    """ 
    Checks and updates the CAS database.
    But before anything:
    """)
# green color:)
st.markdown("""
    <style>
    .stApp {
        background-color: #90EE90;
    }
    </style>
""", unsafe_allow_html=True)
# useless button
if st.button(":green[Press this button if you need some cheering up]"):
    st.balloons()
    st.toast("You can do this!", icon="🎉")
st.markdown("""
<style>
div.stButton > button {
    background-color: #0BDA51;     
    color: white;                 
    border-radius: 8px;
    border: none;
    padding: 0.6em 1.2em;
}
/* HOVER */
div.stButton > button:hover {
    background-color: #45a049;
    color: white;
}
""", unsafe_allow_html=True)
# CAS file upload
st.header(
    """ 
    Upload excel file with CAS numbers
    """)
uploaded_file = st.file_uploader("Upload Excel file with CAS: A column with name CAS containing all CAS/EC numbers to screen in individual rows below. This should be on the first sheet.", type=["xlsx", "xlsm"])
#database_location = st.file_uploader("Upload a text file with database location in .txt", type=["txt"])

database_location = st.text_input(
    "Enter the full path to the SQLite database file",
    placeholder="/Users/Library/new_DB_tests/Database/C2Cdatabase.db"
)
def strip_outer_quotes(s: str) -> str:
    if not s:
        return s
    s = s.strip()
    if (s.startswith("'") and s.endswith("'")) or (s.startswith('"') and s.endswith('"')):
        return s[1:-1]
    return s
database_location = strip_outer_quotes(database_location)
#st.selectbox

# Uploading the excel with CAS numbers
if uploaded_file is not None:
    # write if the file was uploaded
    CASallpd = pd.read_excel(uploaded_file)
    if "CAS" not in CASallpd.columns:
        st.error("The 'CAS' column was not found in the Excel file.")
    else:
        # Clean CAS values
        raw_cas = (
            CASallpd["CAS"]
            .dropna()
            .astype(str)
            .str.strip()
            .tolist()
        )

        # Detect duplicates
        total_count = len(raw_cas)
        unique_cas = sorted(set(raw_cas))
        unique_count = len(unique_cas)
        duplicate_count = total_count - unique_count

        # Optional: list duplicated CAS explicitly
        duplicated_cas = (
            pd.Series(raw_cas)
            .value_counts()
            .loc[lambda x: x > 1]
            .index
            .tolist()
        )

        # Use only unique CAS downstream
        CASall = unique_cas

        # Format for API
        formatted_cas = [{"casNumber": cas, "ecNumber": ""} for cas in CASall]
        formatted_ec = [{"casNumber": "", "ecNumber": ec} for ec in CASall]

        # Streamlit feedback
        st.success(f"Uploaded Excel file with {unique_count} unique CAS numbers: {', '.join(CASall)}")

        # this is not necessary but informs about duplicates:
        if duplicate_count > 0:
            st.warning(f"The Excel file had duplicate CAS entries: {', '.join(duplicated_cas)}")
#starting with no path and then uploading it
db_path = 0
# Uploading the database location
if database_location is not None:
    if database_location:
        if not os.path.exists(database_location):
            st.error("The specified path does not exist.")
        elif not os.path.isfile(database_location):
            st.error("The specified path is not a file.")
        elif not database_location.lower().endswith(".db"):
            st.error("The selected file is not a .db SQLite database.")
        else:
            db_path = database_location
            st.success(f"Database found: {db_path}")
            st.write(db_path)

### FUNCTIONS
def make_a_backup(db_path, backup_dir):
    try:
        connection = sqlite3.connect(db_path)
        st.write("Connected to SQLite database:", db_path)

        # Create date-stamped backup filename
        today = datetime.now().strftime("%Y-%m-%d")
        db_name = os.path.basename(db_path)
        base, ext = os.path.splitext(db_name)
        backup_filename = f"{base}_backup_from_{today}{ext}"

        # Full backup path
        backup_path = os.path.join(backup_dir, backup_filename)

        # Create backup connection
        backup_conn = sqlite3.connect(backup_path)

        # Perform the backup
        with backup_conn:
            connection.backup(backup_conn)

        st.write("Backup made:", backup_path)

    finally:
        if 'backup_conn' in locals():
            backup_conn.close()
        if connection:
            connection.close()
            st.write("Connection closed.")
def check_json(CASall, API_key, save_json_dirr):
    st.write('Checking API')

    # Load the API key from file: It's on the dropbox under Science/Data searches/ED screener/input databases/NextSDS API key.txt
    with open(API_key) as creds:
        api_key = creds.readlines()[0]  # API key is on the first line

    # Split up list in smaller parts (chunks)
    def chunk_list(lst, chunk_size):
        for i in range(0, len(lst), chunk_size):
            yield lst[i:i + chunk_size]

    start_url = "https://api.nextsds.com/jobs/start"
    status_url = "https://api.nextsds.com/jobs/retrieve"
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    # Step 1: Submit all jobs
    jobs = []
    for idx, cas_chunk in enumerate(chunk_list(CASall, 250)):
        data = {
            "taskId": "echa-api",
            "payload": cas_chunk
        }
        try:
            response = requests.post(start_url, headers=headers, json=data)
            if response.status_code == 200:
                job_id = response.json().get("id")
                jobs.append({"id": job_id, "index": idx + 1, "done": False, "output": None})
                st.write(f"Chunk {idx + 1}: Job submitted successfully: {job_id}")
            else:
                st.write(f"Chunk {idx + 1}: Failed to submit job")
        except Exception as e:
            st.write(f"Chunk {idx + 1}: Exception during job submission: {str(e)}")

    # Step 2: Monitor all jobs in one loop
    while not all(job["done"] for job in jobs):
        time.sleep(10)
        for job in jobs:
            if job["done"]:
                continue
            try:
                status_response = requests.get(f"{status_url}/{job['id']}", headers=headers)
                if status_response.status_code == 200:
                    status_data = status_response.json()
                    job_status = status_data.get("status")
                    st.write(f"Chunk {job['index']}: Job status: {job_status}")
                    if job_status not in ["STARTED", "EXECUTING"]:
                        job["done"] = True
                        job["output"] = status_data.get("output", [])
                elif status_response.status_code in [400, 404]:
                    st.write(f"Chunk {job['index']}: Job error ({status_response.status_code})")
                    job["done"] = True
            except Exception as e:
                st.write(f"Chunk {job['index']}: Exception during status check: {str(e)}")

    # Step 3: Combine all outputs
    CnL_json = []
    for job in jobs:
        if job["output"]:
            CnL_json.extend(job["output"])

    # Save to a JSON file
    exportpath = os.path.join(save_json_dirr,"output")
    st.write("Save to:",exportpath)
    if not os.path.exists(exportpath):
        os.makedirs(exportpath)
    formatted_time = datetime.now().strftime("%Y-%m-%d %H-%M")  # Customize format as needed
    exportjson = os.path.join(exportpath, "CnLscreener exportJSON " + formatted_time +".json")
    with open(exportjson, "w") as json_file:
        json.dump(CnL_json, json_file, indent=2)
    return CnL_json
def checking_if_CAS_exists(CASall, db_path):
    found = []
    not_found = []

    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()

        for cas in CASall:
            # looks at ID and
            cursor.execute(
                """
                SELECT 1
                WHERE EXISTS (
                    SELECT 1 FROM C2C_DATABASE WHERE ID = ?
                )
                AND EXISTS (
                    SELECT 1 FROM GENERALINFO WHERE ref = ?
                )
                """,
                (cas, cas)
            )
            row = cursor.fetchone()

            if row:
                found.append(cas)
            else:
                not_found.append(cas)

    except sqlite3.Error as e:
        st.write("SQLite error:", e)

    finally:
        if 'connection' in locals():
            connection.close()

    return found, not_found
def check_if_excel_is_in_folder(folder_excels, CAS_list):
    CAS_in_folder = []
    CAS_not_in_folder = []

    file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
    cas_pattern = re.compile(r'CAS (\d{2,7}[-‐-–—]\d{2,3}[-‐-–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
    ec_pattern = re.compile(r'EC (\d{2,7}[-‐-–—]\d{3}[-‐-–—]\d{1})')

    # collect all inventory numbers (CAS) found in the folder
    inv_in_folder = set()

    for filename in os.listdir(folder_excels):
        full_path = os.path.join(folder_excels, filename)
        if os.path.isfile(full_path):
            match = file_pattern.search(filename)
            if match:
                # Extract CAS inventory number (inv_number)
                match_inv = cas_pattern.search(filename)
                if match_inv:
                    inv_number = match_inv.group(1)  # this is the CAS
                    inv_in_folder.add(inv_number)
                else:
                    # if no CAS, then check for EC (kept from your code)
                    match_inv = ec_pattern.search(filename)
                    if match_inv:
                        inv_number = match_inv.group(1)
                    else:
                        st.write(f"Issue with: {filename}")

    # compare input CAS_list against what was found in folder
    for cas in CAS_list:
        if cas in inv_in_folder:
            CAS_in_folder.append(cas)
        else:
            CAS_not_in_folder.append(cas)

    return CAS_in_folder, CAS_not_in_folder
def insert_json_info_to_DB(CnL_json, db_path, target_cas_list):
    data = CnL_json
    cas_hazards = {} # used later to create a list of things to update
    cas_with_no_json = []
    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
     #id INTEGER PRIMARY KEY AUTOINCREMENT,
        # Ensure ECHACHEM_CL table exists
        cursor.execute("PRAGMA foreign_keys = ON;")
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS ECHACHEM_CL (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL,
            on_cl TEXT,
            cas TEXT,
            ec TEXT,
            name_echachem TEXT,
            type_classification TEXT,
            hazards TEXT,
            date_checked TEXT,
            FOREIGN KEY (code) REFERENCES C2C_DATABASE(ID)
        )
        """)


        st.write("Connected to SQLite database at:", db_path)

        today = datetime.now().date()

        for target_cas in target_cas_list:

            # Find the entry for the CAS you want
            entry = next((e for e in data if e.get("casNumber") == target_cas), None)

            if entry is None:
                st.write(f"CAS {target_cas} not found in JSON.")
                cas_with_no_json.append(target_cas)
            else:
                # Set up dictionary to collect all relevant info
                sqlinfo = {
                    "code": entry.get("casNumber"),
                    "on_cl": "-",
                    "cas": "-",
                    "ec": "-",
                    "name_echachem": "-",
                    "type_classification": "-",
                    "hazards": "-"
                }

                st.write(f"Testing for: {entry.get('casNumber')}")

                #### ECHA-CHEM C&L from NEXTSDS-API ####
                if entry.get("found") is False:  # If the chemical was NOT found on C&L
                    sqlinfo["on_cl"] = "No"
                else:  # If the chemical was found on C&L
                    sqlinfo["on_cl"] = "Yes"
                    sqlinfo["cas"] = entry.get("cas")
                    sqlinfo["ec"] = entry.get("ecNumber")
                    sqlinfo["name_echachem"] = entry.get("name")

                    if entry.get("isHarmonized") is True:
                        sqlinfo["type_classification"] = "Harmonized"
                    else:
                        sqlinfo["type_classification"] = "Self-classification"

                    # Safe hazards extraction (prevents crashes if hazards is missing/not a dict)
                    hazards = entry.get("hazards", {})
                    if isinstance(hazards, dict):
                        sqlinfo["hazards"] = hazards.get("hazardClasses", "-")
                #st.write(sqlinfo)

        #     with open(CnL_json, "r", encoding="utf-8") as f:
        #         data = json.load(f)
        #
        #         # If JSON is a list of entries
        #         for entry in data:
        #             print(entry)
        #         # Set up dictionary to collect all relevant info
        #         sqlinfo = {"code": entry.get("casNumber"), "on_cl": "-", "cas": "-", "ec": "-", "name_echachem": "-",
        #                    "type_classification": "-", "hazards": "-"}
        #         print(f"Adding chemical: {entry.get("casNumber")}")
        #
        #         #### ECHA-CHEM C&L from NEXTSDS-API ####
        #         if entry.get("found") == False:  # If the chemical was NOT found on C&L
        #             sqlinfo["on_cl"] = "No"
        #         else:  # If the chemical was found on C&L (then there is no "found" entry)
        #             sqlinfo["on_cl"] = "Yes"
        #             sqlinfo["cas"] = entry.get("cas")
        #             sqlinfo["ec"] = entry.get("ecNumber")
        #             sqlinfo["name_echachem"] = entry.get("name")
        #             if entry.get("isHarmonized") == True:
        #                 sqlinfo["type_classification"] = "Harmonized"
        #             else:
        #                 sqlinfo["type_classification"] = "Self-classification"
        #             sqlinfo["hazards"] = entry.get("hazards")["hazardClasses"]
        #
                # Check if CAS already exists
                cursor.execute("SELECT 1 FROM ECHACHEM_CL WHERE cas = ?", (sqlinfo["cas"],))
                exists = cursor.fetchone()
                if exists:
                    st.write(f"CAS {sqlinfo['cas']} already in database")
                    cursor.execute("SELECT 1 FROM ECHACHEM_CL WHERE cas = ? AND hazards = ?", (sqlinfo["cas"],sqlinfo["hazards"]))
                    same_hazard = cursor.fetchone()
                    if same_hazard:
                        st.write(f"Hazards for {sqlinfo['cas']} are the same as for the last update. NO ACTION NEEDED.")
                    else:
                        st.write(f"Inserting CAS {sqlinfo['cas']}...")
                        cursor.execute("""
                            UPDATE ECHACHEM_CL
                            SET hazards = ?, date_checked = ?
                            WHERE cas = ?
                        """, (
                            sqlinfo["hazards"],
                            today,
                            sqlinfo["cas"]
                        ))
                        connection.commit()
                        st.write(f"Hazards for {sqlinfo['cas']} are DIFFERENT as for the last update. INFO IN TABLE CnL UPDATED. ACTION REQUIRED.")

                        ### Needed for the next step to gather info and update
                        cursor.execute("SELECT hazards FROM ECHACHEM_CL WHERE code = ?",
                                       (sqlinfo["code"],))
                        # needed if info was added
                        row = cursor.fetchone()
                        hazards_list = row[0].split(",") if row and row[0] else []
                        cas_hazards[target_cas] = hazards_list
                        st.write(hazards_list)


                else:
                    st.write(f"CAS not in CnL database: {sqlinfo['cas']}")
                    cursor.execute(
                        "SELECT 1 FROM C2C_DATABASE WHERE ID = ?",
                        (sqlinfo["code"],)
                    )
                    exists = cursor.fetchone()
                    if not exists:
                        cursor.execute(
                            "INSERT INTO C2C_DATABASE (ID) VALUES (?)",
                            (sqlinfo["code"],)
                        )
                        st.write(f"CAS was not in the main C2C database. CAS added to the main C2C database: {sqlinfo['cas']}")
                    else:
                        st.write(f"CAS already exists in the main C2C database: {sqlinfo['cas']}")

                    cursor.execute("""
                        INSERT INTO ECHACHEM_CL (code, on_cl, cas, ec, name_echachem, type_classification, hazards, date_checked)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        sqlinfo["code"],
                        sqlinfo["on_cl"],
                        sqlinfo["cas"],
                        sqlinfo["ec"],
                        sqlinfo["name_echachem"],
                        sqlinfo["type_classification"],
                        sqlinfo["hazards"],
                        today
                    ))
                    connection.commit()
                    st.write(f"Information inserted to CnL database: {sqlinfo['cas']}")
                    ### Needed for the next step to gather info and update
                    cursor.execute("SELECT hazards FROM ECHACHEM_CL WHERE code = ?",
                                                 (sqlinfo["code"],))
                    #needed if info was added
                    row = cursor.fetchone()
                    hazards_list = row[0].split(",") if row and row[0] else []
                    cas_hazards[target_cas] = hazards_list
                    st.write(hazards_list)


        st.write(f"SQL checked, in case hazards were changed, they are here: {cas_hazards}")
        return cas_hazards, cas_with_no_json
    #
    finally:
        if connection:
            connection.commit()
            connection.close()
            st.write("Connection closed.")
def is_DB_data_up_to_date_with_excel(db_path, folder_excels, CAS_list):
    excel_files_that_need_updating = []
    CAS_older_than_3_years = []
    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
        st.write("Connected to SQLite database at", db_path)

        cursor.execute('''
                CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                    ID TEXT PRIMARY KEY,
                    LastUpdate TEXT,
                    FileName TEXT,
                    Comments TEXT
                )
            ''')
        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        for filename in os.listdir(folder_excels):
            full_path = os.path.join(folder_excels, filename)
            if not os.path.isfile(full_path):
                continue
            if not file_pattern.search(filename):
                continue
            # File date
            mod_time = os.path.getmtime(full_path)
            last_update = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d")

            # Extract CAS or EC
            match_inv = cas_pattern.search(filename)
            if match_inv is None:
                st.write("There should be something here. Please check.")

            if match_inv:
                inv_number = match_inv.group(1)
                comments = "CAS"
                if match_inv.group(2):
                    comments = "CAS, " + match_inv.group(2)
                #check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_list:
                    continue
            else:
                match_inv = ec_pattern.search(filename)
                if not match_inv:
                    st.write(f"Cannot extract CAS/EC from: {filename}")
                    continue
                inv_number = match_inv.group(1)
                comments = "EC"
                # check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_list:
                    continue

            cursor.execute("""
                    UPDATE C2C_DATABASE
                    SET LastUpdate = NULL
                    WHERE LastUpdate NOT LIKE '____-__-__'
                """)

            # Check what DB currently has
            cursor.execute("SELECT LastUpdate FROM C2C_DATABASE WHERE ID = ?", (inv_number,))
            row = cursor.fetchone()

            if row is None:
                # ID not in DB -> insert
                cursor.execute(
                    "INSERT INTO C2C_DATABASE (ID, LastUpdate, FileName, Comments) VALUES (?,?,?,?)",
                    (inv_number, last_update, filename, comments)
                )
                st.write(f"CHANGED (was not there before): inserted {inv_number}: set LastUpdate={last_update}")
                excel_files_that_need_updating.append(inv_number)

            else:
                db_last_update = row[0]  # ISO YYYY-MM-DD (string) or None

                # CHECKING: if the CAS is 3-year or older (DB date) ---
                three_years_ago = datetime.now() - timedelta(days=3 * 365)

                if db_last_update is None:
                    st.write(f"{inv_number}: No LastUpdate date available in DB")
                else:
                    try:
                        db_last_update_date = datetime.strptime(db_last_update, "%Y-%m-%d")
                        if db_last_update_date < three_years_ago:
                            st.write(f"{inv_number}: OLD CPS: DB LastUpdate is older than 3 years ({db_last_update})")
                            CAS_older_than_3_years.append(inv_number)
                        else:
                            st.write(
                                f"{inv_number}: NO ACTION NEEDED. DB LastUpdate is not older than 3 years {db_last_update_date} – good to go")

                            ## Checks for the modifications if the CAS is not older than 3 years

                            if db_last_update is None or db_last_update < last_update:
                                # File is newer -> update
                                cursor.execute(
                                    "UPDATE C2C_DATABASE SET LastUpdate = ?, FileName = ?, Comments = ? WHERE ID = ?",
                                    (last_update, filename, comments, inv_number)
                                )
                                st.write(
                                    f"{inv_number}: CHANGED: inserted {inv_number}: {db_last_update} -> {last_update}")
                                excel_files_that_need_updating.append(inv_number)

                            else:
                                # DB is newer or same -> skip
                                st.write(
                                    f"{inv_number}: NO ACTION NEEDED. CAS is up to date: DB {db_last_update} >= file {last_update}.")
                    except ValueError:
                        st.write(f"{inv_number}: WRONG DATE FORMAT: Invalid DB LastUpdate format ({db_last_update})")


        return excel_files_that_need_updating, CAS_older_than_3_years

    finally:
        if connection:
            connection.commit()
            connection.close()
        st.write("Connection closed.")
def extract_info_form_excel_to_DB(db_path, folder_excels, CAS_needing_DB_update):
    #### CUSTOM FUNCTIONS ####
    def add_info_CPS_below(sheet, search_strings, maindatabase, newdatabase, mainID):
        """
        Finds each search string somewhere in the sheet (not necessarily in the first row),
        then reads the values directly below it column-wise.
        Collects one "row" of data across all search strings,
        and inserts into the SQLite database.

        Stops when *all* searched columns are empty in the same row.

        search_strings:
            - list: Excel labels are used as SQL column names.
            - dict: {Excel label: SQL column name}
        """

        # Handle both list and dict
        if isinstance(search_strings, str):
            search_strings = [search_strings]
        if isinstance(search_strings, list):
            mapping = {s: s for s in search_strings}  # Excel label -> same SQL name
        elif isinstance(search_strings, dict):
            mapping = search_strings  # Excel label -> custom SQL name
        else:
            raise TypeError("search_strings must be a str, list, or dict")

        # Locate the target cells (positions of the labels in Excel)
        col_positions = {}
        for excel_label in mapping.keys():
            found = None
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and excel_label.lower() in str(cell.value).lower():
                        found = cell
                        break
                if found:
                    break
            if found:
                col_positions[excel_label] = (found.column, found.row + 1)  # start below label

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'

        # Check if the table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                       (newdatabase,))
        table_exists = cursor.fetchone()

        if not table_exists:
            # Create table with auto-increment ID and extracted columns
            cols_def = ", ".join([f"{q(sql_col)} TEXT" for sql_col in mapping.values()])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            # Add missing columns if needed
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            for sql_col in mapping.values():
                if sql_col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(sql_col)} TEXT")

        # Iterate row by row until all searched columns are empty
        row_offset = 0
        while True:
            row_data = {}
            all_empty = True
            for excel_label, (col, start_row) in col_positions.items():
                cell = sheet.cell(row=start_row + row_offset, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    sql_col = mapping[excel_label]
                    row_data[sql_col] = cell.value
                    all_empty = False
            if all_empty:
                break  # stop when all searched columns are empty in the same row

            if newdatabase == maindatabase:
                # --- Case 1: Append into the main database table ---
                # Check if the row for this ID exists
                cursor.execute(
                    f"SELECT 1 FROM {q(maindatabase)} WHERE ID = ?",
                    (mainID,)
                )
                exists = cursor.fetchone()

                if exists:
                    # Update only the new columns (append data)
                    update_clause = ", ".join([f"{q(col)} = ?" for col in row_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(maindatabase)} SET {update_clause} WHERE ID = ?",
                        list(row_data.values()) + [mainID]
                    )
                else:
                    # Insert new row with ID and these values
                    all_cols = ["ID"] + list(row_data.keys())
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(maindatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        [mainID] + list(row_data.values())
                    )

            else:
                # --- Case 2: Insert into child table ---
                row_data["ref"] = mainID

                # Build WHERE clause dynamically based on all row_data keys
                where_clause = " AND ".join([f"{q(col)} = ?" for col in row_data.keys()])
                params = list(row_data.values())

                cursor.execute(
                    f"SELECT 1 FROM {q(newdatabase)} WHERE {where_clause}",
                    params
                )
                exists = cursor.fetchone()

                if not exists:
                    # Insert new row
                    all_cols = list(row_data.keys())
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        list(row_data.values())
                    )

            row_offset += 1
    def add_info_CPS_one_cell_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
        """
        Finds rows containing `rowlabel`, extracts specified columns to the right,
        and inserts or updates the data in the SQLite database.

        Parameters:
            sheet: openpyxl worksheet
            rowlabel: string to search for in any row
            column_offsets: list of integers (e.g., [2, 3]) for columns to the right
            column_names: list of strings for custom SQL column names
            maindatabase: name of the main database (for foreign key reference)
            newdatabase: name of the table to update
            mainID: unique identifier for the row
        """

        if len(column_offsets) != len(column_names):
            raise ValueError("column_offsets and column_names must have the same length")

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'


        # Step 1: Find the row containing the rowlabel
        extracted_data = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                    row_idx = cell.row
                    col_idx = cell.column
                    for offset, col_name in zip(column_offsets, column_names):
                        target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                        extracted_data[col_name] = target_cell.value
                    break
            if extracted_data:
                break

        if not extracted_data:
            return  # nothing to insert

        # Step 2: Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        if not table_exists:
            # Create table with ID, ref, and extracted columns
            cols_def = ", ".join([f"{q(col)} TEXT" for col in column_names])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            # Add missing columns
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in column_names:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # Step 3: Insert or update
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                    [extracted_data[col] for col in column_names] + [mainID]
                )
            else:
                all_cols = ['ref'] + column_names
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in column_names]
                )

        else:  # when newdatabase == maindatabase
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                    [extracted_data[col] for col in column_names] + [mainID]
                )
            else:
                all_cols = ['ID'] + column_names
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in column_names]
                )
    def add_info_CPS_right_until_empty(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID, include_resource=True):
        """
        Like add_info_CPS_one_cell_right, but starting from column_offsets[0] to the right,
        keeps reading consecutive cells until it finds the first empty cell.
        Column naming:
          - first value uses column_names[0] (base)
          - next values use column_names[1:], if present
          - beyond that, auto-name as base-1, base-2, ...

        Optional behavior (when True):
          - If include_resource=True, captures the sheet's 'Resource' column value (if present and not empty)
            into SQL column 'resource-<sanitized rowlabel>'.
          - If no 'Resource' column exists or the cell is empty, skips creating that column.
        """

        if len(column_offsets) != len(column_names):
            raise ValueError("column_offsets and column_names must have the same length")
        if not column_offsets:
            return

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'

        # Sanitize rowlabel for safe SQL column naming
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        safe_rowlabel = sanitize_label(rowlabel)
        resource_colname = f"resource-{safe_rowlabel}"

        # --- locate the cell containing rowlabel ---
        match_row_idx = None
        match_col_idx = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                    match_row_idx = cell.row
                    match_col_idx = getattr(cell, "col_idx", cell.column)
                    break
            if match_row_idx is not None:
                break

        if match_row_idx is None:
            return  # nothing to insert

        # Optionally find the column index for "Resource"
        resource_col = None
        if include_resource:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == "resource":
                        resource_col = getattr(cell, "col_idx", cell.column)
                        break
                if resource_col:
                    break

        # Determine start offset and base name
        start_offset = column_offsets[0]
        base_name = column_names[0]

        # --- read to the right until the first empty cell ---
        extracted_data = {}
        k = 0
        max_col = sheet.max_column
        while (match_col_idx + start_offset + k) <= max_col:
            target = sheet.cell(row=match_row_idx, column=match_col_idx + start_offset + k)
            tv = target.value
            # stop at first empty/blank
            if tv is None or (isinstance(tv, str) and tv.strip() == ""):
                break

            # choose column name
            if k < len(column_names):
                col_name = column_names[k]
            else:
                col_name = f"{base_name}-{k - (len(column_names) - 1)}" if len(column_names) > 0 else f"col-{k}"

            extracted_data[col_name] = tv
            k += 1

        # Optionally add the Resource value, but only if it's not empty
        if include_resource and resource_col:
            resource_value = sheet.cell(row=match_row_idx, column=resource_col).value
            if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
                extracted_data[resource_colname] = resource_value

        if not extracted_data:
            return  # nothing to insert

        # --- ensure table exists and has needed columns ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- upsert logic (same keying as your working function) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                    [extracted_data[col] for col in needed_columns] + [mainID]
                )
            else:
                all_cols = ['ref'] + needed_columns
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in needed_columns]
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                    [extracted_data[col] for col in needed_columns] + [mainID]
                )
            else:
                all_cols = ['ID'] + needed_columns
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in needed_columns]
                )
    def loop_over_to_collect_right_values(sheet, rowlabel: str):
        """
        Finds the first cell whose value contains `rowlabel`, then collects the value
        in the cell to the right for that row and all consecutive rows below
        where the same label also appears in the same column.
        Returns a single string joining all collected right-hand values separated with commas.
        """

        # Case-insensitive: looking for upper or lower case
        def matches(value):
            if value is None:
                return False
            return str(value).strip().lower() == rowlabel.strip().lower()

        # Step 1: Find the first occurrence of the label
        start_row = None
        start_col = None
        for row in sheet.iter_rows():
            for cell in row:
                if matches(cell.value):
                    start_row = cell.row
                    start_col = getattr(cell, "col_idx", cell.column)
                    break
            if start_row is not None:
                break

        if start_row is None:
            return "No match with row label"

        # Step 2: Walk down while the label repeats and collect right-hand values
        collected = []
        r = start_row
        while r <= sheet.max_row:
            left_val = sheet.cell(row=r, column=start_col).value
            if not matches(left_val):
                break

            right_val = sheet.cell(row=r, column=start_col + 1).value
            if right_val is not None and str(right_val).strip() != "":
                collected.append(str(right_val).strip())

            r += 1

        # Step 2: Return as a single string
        return collected
    def add_info_CPS_from_row_with_two_markers(sheet, label1: str, label2: str, label3: str, label4: str, maindatabase, newdatabase, mainID):
        """
        label 1 - first row to match (e.g. Hazard classification)
        label 2 - second row to match (e.g. Eye Irrit. 2)
        label 3 & 4 looking in the row for them and taking values to the right
        1) Find a row where two adjacent cells match (label1, label2) left→right.
        2) In that row, find label3 and label4 (anywhere in the row), and for each:
           - take the value from the cell immediately to the right.
        3) Write to SQL columns named:
             - "{label3}-{label2}"  (for value next to label3)
             - "{label4}-{label2}"  (for value next to label4)
        Matching is case-insensitive 'contains'.
        Requires a DB cursor in outer scope named `cursor`.
        """

        # --- helpers ---
        def q(name: str) -> str:
            return f'"{name}"'

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        max_row = sheet.max_row
        max_col = sheet.max_column

        # --- 1) Find target row via adjacent (label1, label2) ---
        target_row = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = sheet.cell(row=r, column=c).value
                v2 = sheet.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    break
            if target_row is not None:
                break

        if target_row is None:
            return  # no matching row → nothing to insert

        # --- 2) Scan the row to find label3 and label4 targets; capture right-hand values ---
        extracted_data = {}

        # label3 → col name "{label3}-{label2}"
        col_name_3 = f"{label2} - {label3}"
        # label4 → col name "{label2}-{label4}"
        col_name_4 = f"{label2} - {label4}"

        # We search the entire row for each label, reading the cell to the right when found
        def capture_right_of_label(row: int, label: str):
            for c in range(1, max_col):  # up to max_col-1 to read c+1
                if matches(sheet.cell(row=row, column=c).value, label):
                    right_val = sheet.cell(row=row, column=c + 1).value
                    if right_val is None:
                        return None
                    if isinstance(right_val, str):
                        rv = right_val.strip()
                        return rv if rv != "" else None
                    return right_val
            return None

        val3 = capture_right_of_label(target_row, label3)
        val4 = capture_right_of_label(target_row, label4)

        if val3 is not None:
            extracted_data[col_name_3] = val3
        if val4 is not None:
            extracted_data[col_name_4] = val4

        if not extracted_data:
            return  # neither target produced a value → nothing to insert

        # --- 3) Ensure table/columns exist ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- 4) Upsert (same rules as your working pattern) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ref"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ID"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
    def add_info_right_two_markers_OECD(sheet, label1: str, label2: str, maindatabase, newdatabase, mainID, include_resource: bool = True):
        """
        label 1 - first row to match
        label 2 - second row to match
        1) Find a row where two adjacent cells match (label1, label2) left→right.
        2) Capture first non-empty cell to the right of label2.
        3) Write to SQL columns named: {label1}{label2}

        Optional behavior (when include_resource=True):
          - Captures the sheet's 'Resource' column value (if present and not empty)
            into SQL column 'resource-<sanitized label2>' for the same row.
          - If no 'Resource' column exists or the cell is empty, skips creating that column.
        """

        # --- helpers ---
        def q(name: str) -> str:
            return f'"{name}"'

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        # Sanitize label for safe SQL column naming (for resource column)
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        # --- NEW: if label2 is "no data", skip Excel and just write "no data" to SQL ---
        if label2.strip().lower() == "no data":
            safe_label2 = sanitize_label(label2)
            col_name = f"{label1}{label2}"   # same pattern as normal case
            extracted_data = {col_name: "no data"}

            # Ensure table/columns exist
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
            table_exists = cursor.fetchone()

            needed_columns = list(extracted_data.keys())

            if not table_exists:
                cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
                fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
                cursor.execute(f'''
                    CREATE TABLE {q(newdatabase)} (
                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                        ref TEXT
                        {"," if cols_def else ""} {cols_def}
                        {fk_clause}
                    )
                ''')
            else:
                cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
                existing_cols = [col[1] for col in cursor.fetchall()]
                if "ref" not in existing_cols and newdatabase != maindatabase:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
                for col in needed_columns:
                    if col not in existing_cols:
                        cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

            # Upsert
            if newdatabase != maindatabase:
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
                exists = cursor.fetchone()
                if exists:
                    set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                        list(extracted_data.values()) + [mainID]
                    )
                else:
                    cols = ["ref"] + list(extracted_data.keys())
                    placeholders = ", ".join(["?"] * len(cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                        [mainID] + list(extracted_data.values())
                    )
            else:
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
                exists = cursor.fetchone()
                if exists:
                    set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                        list(extracted_data.values()) + [mainID]
                    )
                else:
                    cols = ["ID"] + list(extracted_data.keys())
                    placeholders = ", ".join(["?"] * len(cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                        [mainID] + list(extracted_data.values())
                    )

            return  # done, no Excel lookup

        # --- Normal behavior below (when label2 is NOT "no data") ---

        max_row = sheet.max_row
        max_col = sheet.max_column

        safe_label2 = sanitize_label(label2)
        resource_colname = f"resource-{safe_label2}"

        # --- 1) Find target row via adjacent (label1, label2) ---
        target_row = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = sheet.cell(row=r, column=c).value
                v2 = sheet.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    break
            if target_row is not None:
                break

        if target_row is None:
            print("Target row not found")
            return  # no matching row → nothing to insert

        # --- Optionally find the column index for "Resource" ---
        resource_col = None
        if include_resource:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == "resource":
                        resource_col = getattr(cell, "col_idx", cell.column)
                        break
                if resource_col:
                    break

        # --- 2) Scan the row to find targets; capture right-hand values ---
        extracted_data = {}

        # label → col name
        col_name = f"{label1}{label2}"

        # move to look for the first value to the right
        def capture_right_of_label(row: int, label: str):
            # Find the column containing the label
            for c in range(1, max_col):
                cell_value = sheet.cell(row=row, column=c).value
                if matches(cell_value, label):

                    # Start searching to the right of this column
                    for cc in range(c + 1, max_col + 1):
                        right_val = sheet.cell(row=row, column=cc).value

                        # Skip empty or whitespace-only
                        if right_val is None:
                            continue

                        if isinstance(right_val, str):
                            rv = right_val.strip()
                            if rv == "":
                                continue
                            return rv  # return first non-empty string

                        # Non-string, non-None → return immediately
                        return right_val

                    # If no value was found to the right
                    return None

            # Label not found at all
            return None

        val = capture_right_of_label(target_row, label2)

        if val is not None:
            extracted_data[col_name] = val

        # Optionally add the Resource value, but only if it's not empty
        if include_resource and resource_col:
            resource_value = sheet.cell(row=target_row, column=resource_col).value
            if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
                extracted_data[resource_colname] = resource_value

        if not extracted_data:
            print("Target extracted not found")
            return  # nothing to insert

        # --- 3) Ensure table/columns exist ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- 4) Upsert (same rules as your working pattern) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ref"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ID"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )

    try:
        ### SQL SET-UP
        connection = sqlite3.connect(db_path)
        st.write("Connected to SQLite database at", db_path)

        # Create main C2C_DATABASE table if not existing
        cursor = connection.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                ID TEXT PRIMARY KEY,
                LastUpdate TEXT,
                FileName TEXT,
                Comments TEXT
            )
        ''')

        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        # Loop through Excel files with CAS number and add their info from the template
        for filename in os.listdir(folder_excels):
            full_path = os.path.join(folder_excels, filename)
            if not os.path.isfile(full_path):
                continue
            if not file_pattern.search(filename):
                continue
            # File date
            mod_time = os.path.getmtime(full_path)
            last_update = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d")

            # Extract CAS number or EC number if applicable
            match_inv = cas_pattern.search(filename)    # Check for CAS number
            comments = "There should be something here. Please check."
            if match_inv:
                inv_number = match_inv.group(1)
                comments = "CAS"
                if match_inv.group(2):
                    comments = "CAS, " + match_inv.group(2)
                #check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_needing_DB_update:
                    continue
            else:
                match_inv = ec_pattern.search(filename)
                if not match_inv:
                    st.write(f"Cannot extract CAS/EC from: {filename}")
                    continue
                inv_number = match_inv.group(1)
                comments = "EC"
                # check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_needing_DB_update:
                    continue
            # Check which inv you work with
            st.write(f"Updating CAS: {inv_number}")
            cursor.execute(
                'INSERT INTO C2C_DATABASE (ID, LastUpdate, FileName , Comments) VALUES (?,?,?,?)',
                (inv_number, last_update, filename, comments))

            # Open the Excel file
            CPS_wb_obj = openpyxl.load_workbook(full_path)
            CPSsheet = CPS_wb_obj.active

            # Add general info
            for g_info in ["Chemical name","Common name","CAS number", "EC number", "Linked CAS Read across", "Linked CAS Monomers", "Linked CAS Degradation Products"]:
                add_info_CPS_below(CPSsheet, g_info,"C2C_DATABASE","GENERALINFO",inv_number )

            # ASSESSOR
            add_info_CPS_below(CPSsheet, {"Name assessor":"Name assessor","Date created/updated" : "Date assessed"},"C2C_DATABASE","ASSESSORS",inv_number)

            # Add various info
            for info in ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]:
                   add_info_CPS_one_cell_right(CPSsheet,info,[2],[info],
                       "C2C_DATABASE","CHEMICALCLASS",inv_number)

            # Adding other info
            for o_info in ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]:
                   add_info_CPS_right_until_empty(CPSsheet,o_info,[2],[o_info],
                       "C2C_DATABASE","OTHERINFO",inv_number, include_resource=False)

            # CARCINOGENICITY
            for carc_type in ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK","Carcinogenicity Classified IARC",
                "Carcinogenicity Classified TLV", "Carcinogenicity experimental evidence", "Carcinogenicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,carc_type,[1],[carc_type],
                    "C2C_DATABASE","CARCINOGENICITY",inv_number)

            # ED
            for ED_type in ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,ED_type,[1],[ED_type],
                    "C2C_DATABASE","ENDOCRINE",inv_number)

            # MUTAGENICITY/GENOTOXICITY
            # General information
            for muta_type in ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,muta_type,[1],[muta_type],
                    "C2C_DATABASE","MUTAGENICITY",inv_number)
            # Point mutations
            P_mut = loop_over_to_collect_right_values(CPSsheet,"Point mutations:")  # make a string with all point mut
            #print("Point mut",P_mut) #print to see if it makes a string with the point mut
            for p_mut in P_mut:
                 add_info_right_two_markers_OECD(CPSsheet,"Point mutations:",p_mut, "C2C_DATABASE","POINTMUT",inv_number)


            # Chromosomal mutations
            Ch_mut = loop_over_to_collect_right_values(CPSsheet,
                                                      "Chromosome damaging:")  # make a string with all point mut
            #print("Ch mut",Ch_mut)  # print to see if it makes a string with the point mut
            for ch_mut in Ch_mut:
                add_info_right_two_markers_OECD(CPSsheet,"Chromosome damaging:",ch_mut,"C2C_DATABASE","CHROMDAM",inv_number)

            # REPRODUCTIVE TOXICITY
            for repro_type in ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,repro_type,[1],[repro_type],
                    "C2C_DATABASE","REPROTOX",inv_number)

            # DEVELOPMENTAL TOXICITY
            for devo_type in ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,devo_type,[1],[devo_type],
                    "C2C_DATABASE","DEVELOPTOX",inv_number)

            # NEUROTOXICITY
            for neuro_type in ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,neuro_type,[1],[neuro_type],
                    "C2C_DATABASE","NEUROTOX",inv_number)

            # ORAL TOXICITY
            for oral_type in ["Oral toxicity Acute Tox classified","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,oral_type,[1],[oral_type],
                    "C2C_DATABASE","ORALTOX",inv_number)

            # INHALATIVE TOXICITY
            for inhal_type in ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,inhal_type,[1],[inhal_type],
                    "C2C_DATABASE","INHALTOX",inv_number)

            # DERMAL TOXICITY
            for dermal_type in ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,dermal_type,[1],[dermal_type],
                    "C2C_DATABASE","DERMALTOX",inv_number)

            # SKIN/EYE IRRITATION/CORROSION
            for irrit_type in ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]:
                add_info_CPS_right_until_empty(CPSsheet,irrit_type,[1],[irrit_type],
                    "C2C_DATABASE","IRRITCOR",inv_number)

            # SENSITISATION
            for sens_type in ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]:
                add_info_CPS_right_until_empty(CPSsheet,sens_type,[1],[sens_type],
                    "C2C_DATABASE","SENSITISATION",inv_number)

            # ADD Specific concentration limits

            SCL = loop_over_to_collect_right_values(CPSsheet, "Hazard classification:") # make a string with all SCL for each CAS
            #print(SCL) #print to see if it makes a string with the specific conc limits
            for spec_conc_lim in SCL:
                add_info_CPS_from_row_with_two_markers(CPSsheet,"Hazard classification:", spec_conc_lim, "Lower Limit: (%)", "Upper Limit: (%)" , "C2C_DATABASE","SCONCLIM",inv_number)

            #  OTHER CRITERIA
            for other_criteria in ["Other comments"]:
                add_info_CPS_right_until_empty(CPSsheet, other_criteria, [1], [other_criteria],
                                   "C2C_DATABASE", "OCRIT", inv_number)

            # AQUATIC TOXICITY
            for aqtox_type in ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]:
                add_info_CPS_right_until_empty(CPSsheet, aqtox_type, [1], [aqtox_type],
                                    "C2C_DATABASE", "AQUATOX", inv_number)
                # VERTEBRATE FISH
            for fish_type in ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet,fish_type,[1],[fish_type],
                    "C2C_DATABASE","FISHTOX",inv_number)
                # INVERTEBRATE
            for inv_type in ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet, inv_type, [1], [inv_type],
                                   "C2C_DATABASE", "INVTOX", inv_number)
                # ALGAE
            for algae_type in ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]:
                add_info_CPS_right_until_empty(CPSsheet, algae_type, [1], [algae_type],
                                   "C2C_DATABASE", "ALGAETOX", inv_number)

            # TERRESTRIAL TOXICITY
            for tertox_type in ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet, tertox_type, [1], [tertox_type],
                                   "C2C_DATABASE", "TERTOX", inv_number)
            # Other species toxicity
            for spec_tox_type in ["Any other CLP species classification"]:
                add_info_CPS_right_until_empty(CPSsheet, spec_tox_type, [1], [spec_tox_type],
                                   "C2C_DATABASE", "SPECTOX", inv_number)

            # PERSISTENCE
            for pers_type in ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]:
                add_info_CPS_right_until_empty(CPSsheet,pers_type,[1],[pers_type],
                    "C2C_DATABASE","PERSISTENCE",inv_number)

            # BIOACCUMULATION
            for bioac_type in ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]:
                add_info_CPS_right_until_empty(CPSsheet,bioac_type,[1],[bioac_type],
                    "C2C_DATABASE","BIOACCUMULATION",inv_number)

            # CLIMATIC RELEVANCE
            for clima_type in ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]:
                add_info_CPS_right_until_empty(CPSsheet,clima_type,[1],[clima_type],
                    "C2C_DATABASE","CLIMATICRELEVANCE",inv_number)

            #  ADDITIONAL SOURCES
            for add_sources in ["Additional sources"]:
                add_info_CPS_right_until_empty(CPSsheet, add_sources, [1], [add_sources],
                                   "C2C_DATABASE", "ADDSOURCE", inv_number)

        connection.commit()
        st.write("SQL updated")
    except sqlite3.Error as e:
        st.write("SQLite error", e, inv_number)

    finally:
        if connection:
            connection.commit()
            connection.close()
        st.write("Connection closed.")
def save_DB_to_excel(db_path, DB_excel_saving_path):
    #### Save Database as Excel ####
    today = datetime.today().strftime("%Y%m%d")
    try:
        ### SQL SET-UP
        connection = sqlite3.connect(db_path)
        st.write("Reconnected to database to export to Excel", db_path)

        # Get a list of all tables in the database
        tables_df = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';", connection)
        tables = tables_df['name'].tolist()
        tables = [t for t in tables if t not in ("C2C_DATABASE", "ECHACHEM_CL")] # Exclude the main table from the aggregation
        # Also excludes ECHA-CHEM as that datatable is not connected through ref
        cte_list = []
        join_list = []
        main_table = "C2C_DATABASE"

        # AGGREGATE ALL DEPENDENT TABLES SO THAT WE CAN EASILY USE LEFTJOIN
        for table in tables:
            cursor = connection.cursor()
            cursor.execute(f"PRAGMA table_info({table})")
            columns_info = cursor.fetchall()
            cols = [col[1] for col in columns_info if col[1] not in ('ref', 'ID')]

            if not cols or 'ref' not in [col[1] for col in columns_info]:
                continue  # skip tables without 'ref'

            # Build GROUP_CONCAT for each column
            group_concat_cols = ",\n       ".join([f"GROUP_CONCAT([{col}], CHAR(10)) AS [{col}]" for col in cols])

            # Define CTE
            cte_name = f"{table}Agg"
            cte = f"{cte_name} AS (\n    SELECT ref,\n           {group_concat_cols}\n    FROM {table}\n    GROUP BY ref\n)"
            cte_list.append(cte)

            # Prepare LEFT JOIN
            join_list.append(f"LEFT JOIN {cte_name} {table[:4].lower()} ON {main_table}.ID = {table[:4].lower()}.ref")

        # Combine all CTEs
        cte_sql = ",\n".join(cte_list)

        # Select all columns from main_table + aggregated tables
        cursor.execute(f"PRAGMA table_info({main_table})")
        main_cols = [f"{main_table}.[{col[1]}]" for col in cursor.fetchall()]

        select_cols = main_cols.copy()
        for table in tables:
            cursor.execute(f"PRAGMA table_info({table})")
            for col in cursor.fetchall():
                if col[1] not in ('ref',"ID"):
                    select_cols.append(f"{table[:4].lower()}.[{col[1]}]")

        select_sql = ",\n    ".join(select_cols)
        join_sql = "\n".join(join_list)

        final_query = f"WITH {cte_sql}\nSELECT\n    {select_sql}\nFROM {main_table}\n{join_sql};"

        df = pd.read_sql_query(final_query, connection)

        # Make sure everything is readable for Excel
        def sanitize_excel_cells(df):
            def fix_value(val):
                if isinstance(val, str):
                    # If it starts with '=', prefix with a single quote to force Excel to treat it as text
                    if val.strip().startswith("="):
                        return "'" + val
                    # Remove control characters that break XML
                    return ''.join(ch for ch in val if ch.isprintable())
                return val

            for col in df.columns:
                df[col] = df[col].map(fix_value)
            return df


        # Clean both headers and cell values
        # df.columns = clean_excel_headers(df.columns)  # from earlier
        df = sanitize_excel_cells(df)

        # Write away the database as Excel
        with pd.ExcelWriter(DB_excel_saving_path + '/Database/C2Cdatabase ' + today + '.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="C2C DATABASE", index=False)
        st.write(f"Database exported to Excel as C2Cdatabase{today}.xlsx")


    except sqlite3.Error as e:
        # Catches SQLite-specific errors
        st.write("SQLite error:", e)
        traceback.print_exc()  # prints the full traceback

    except Exception as e:
        # Catches other Python errors
        st.write("General error:", e)
        traceback.print_exc()

    finally:
        # Always close connection if it was created
        try:
            connection.close()
            st.write("Connection closed.")
        except NameError:
            pass
def extraction_info_excels(database, template_path, CAS, folder, image_dir):
    '''Function saves CAS from DB to excel'''
    def db_to_excel_multiple_below(maindb, main_ref, linked_db, link_ref, column_to_get, lookup_column, lookup_value,
                                   label_excel):

        # Query the database for all matching values
        try:
            query = f"""
                 SELECT a.[{column_to_get}]
                 FROM {linked_db} a
                 JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                 WHERE c.{lookup_column} = ?
             """
            cursor.execute(query, (lookup_value,))
            results = cursor.fetchall()
            if not results:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            # Find the label in the worksheet
            for row in ws_template.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                        start_row = cell.row + 1
                        col = cell.column
                        # print(f"First test on{start_row, col}")

                        # Place each value in the first empty cell below the starting row
                        for result in results:
                            # If result is not none:

                            if result[0] != None:
                                # Start searching from start_row downward
                                target_row = start_row

                                # Keep moving down until we find an empty cell in the target column
                                while ws_template.cell(row=target_row, column=col).value not in (None, ''):
                                    target_row += 1

                                # print(f"target row{target_row}")

                                # Write the value in the first empty cell found
                                ws_template.cell(row=target_row, column=col).value = result[0]

                                print(
                                    f"Inserted '{result[0]}' into cell {ws_template.cell(row=target_row, column=col).coordinate}")

                        return

            print(f"Label '{label_excel}' not found in worksheet.")
        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_excel_source_right(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,    # base name, e.g. "source"
        lookup_column,
        lookup_value,
        label_excel,
        offset,
        max_suffix=5,    # how far to look for -1, -2 for the additional data
        include_resource=True # does it look for resources
    ):
        # helper: sanitize label like in add_info_CPS_right_until_empty
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]  # row[1] is column name

            # 2) Collect base + suffix columns for the main data (source, source-1, source-2, ...)
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                candidate = f"{column_to_get}-{i}"
                if candidate in all_cols:
                    matching_cols.append(candidate)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column name in SQL: resource-<sanitized label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(label_excel)
                candidate_resource_col = f"resource-{safe_label}"
                if candidate_resource_col in all_cols:
                    resource_sql_col = candidate_resource_col
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")

            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            # Split row into value columns and (optional) resource
            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # 5) Find the label in the worksheet
            for excel_row in ws_template.iter_rows():
                for cell in excel_row:
                    if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                        start_row = cell.row
                        start_col = cell.column + offset  # first column to write values

                        # 6) Write the main values horizontally to the right
                        current_col = start_col
                        for val in value_cols:
                            if val is not None and val != "":
                                ws_template.cell(row=start_row, column=current_col).value = val
                                print(
                                    f"Inserted '{val}' into cell "
                                    f"{ws_template.cell(row=start_row, column=current_col).coordinate}"
                                )
                            current_col += 1

                        # 7) If resource is present in SQL, put it in Excel "Resource" column
                        if include_resource and resource_val not in (None, ""):
                            resource_col_idx = None
                            # Find the "Resource" header column in the template
                            for hdr_row in ws_template.iter_rows():
                                for hdr_cell in hdr_row:
                                    if (
                                        isinstance(hdr_cell.value, str)
                                        and hdr_cell.value.strip().lower() == "resource"
                                    ):
                                        resource_col_idx = hdr_cell.column
                                        break
                                if resource_col_idx is not None:
                                    break

                            if resource_col_idx is not None:
                                ws_template.cell(row=start_row, column=resource_col_idx).value = resource_val
                                print(
                                    f"Inserted resource '{resource_val}' into cell "
                                    f"{ws_template.cell(row=start_row, column=resource_col_idx).coordinate}"
                                )
                            else:
                                print("Could not find 'Resource' column in Excel to write resource value.")

                        return  # done after first matching label

            print(f"Label '{label_excel}' not found in worksheet.")

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_column_names_unique(maindb, main_ref,linked_db, link_ref,lookup_column, lookup_value):
        """
        Returns a string with column names unique for each CAS, as a string
        """

        try:
            # Query EVERYTHING (*) from linked_db
            query = f"""
                SELECT a.*
                FROM {linked_db} a
                JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
            """

            cursor.execute(query, (lookup_value,))
            rows = cursor.fetchall()

            # If nothing found → return empty DataFrame (still safe)
            if not rows:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return pd.DataFrame()

            # Extract column names automatically from cursor.description
            colnames = [desc[0] for desc in cursor.description]

            dataframe = pd.DataFrame(rows, columns=colnames)

            # cutting columns with NULL values
            dataframe_cut = dataframe.dropna(axis=1, how='all')
            # dropping columns with ID and ref (not needed here)
            dataframe_cut = dataframe_cut.drop(columns=["ID", 'ref'])
            column_name = list(dataframe_cut.columns)
            # takes away from the string the resources names
            result = [c_name for c_name in column_name if "resource" not in c_name.lower() ]
            return result

        except sqlite3.Error as e:
            print("SQLite error:", e)
            return pd.DataFrame()

    def remove_text_from_string(string, target_name):
        '''removes text from string, used for Muta tests and SCL'''
        result = []
        for s in string:
            name = s.replace(target_name, "").strip()
            result.append(name)
        return result

    def write_list_right_of_label(ws_template: Worksheet, label_excel: str, offset: int, values_list: list):
        """
        Find the first cell whose value EXACTLY matches 'label_excel'
        (after stripping whitespace, case-insensitive),
        then write values from values_list to the right, moving downwards
        as long as the label cell below also exactly matches the label.
        """

        if not values_list:
            print("Value list is empty — nothing to write.")
            return

        # Normalize the target label once
        normalized_label = label_excel.strip().lower()

        # 1) Find the first exact match
        label_cell = None
        for row in ws_template.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == normalized_label:
                    label_cell = cell
                    break
            if label_cell:
                break

        if not label_cell:
            print(f"Exact label '{label_excel}' not found.")
            return

        start_row = label_cell.row
        label_col = label_cell.column
        target_col = label_col + offset
        max_row = ws_template.max_row

        current_row = start_row
        value_index = 0

        # 2) Write downward while exact match continues
        while value_index < len(values_list) and current_row <= max_row:
            current_cell_value = ws_template.cell(row=current_row, column=label_col).value

            # Check if current row still has EXACT match
            if not isinstance(current_cell_value, str) or \
               current_cell_value.strip().lower() != normalized_label:
                break

            ws_template.cell(row=current_row, column=target_col).value = values_list[value_index]

            print(
                f"Inserted '{values_list[value_index]}' into "
                f"{ws_template.cell(row=current_row, column=target_col).coordinate}"
            )

            value_index += 1
            current_row += 1

        if value_index < len(values_list):
            print(f"Warning: {len(values_list) - value_index} values not written — no more exact matching label rows.")

    def refdb_to_excel_source_after_two_targets(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,      # base name, e.g. "source"
        lookup_column,
        lookup_value,
        first_label_excel,  # label of first target cell
        second_label_excel, # label of second target cell (same row as first)
        max_suffix=5,       # how far to look for -1, -2, ... columns
        include_resource=True
    ):
        """
        Finds a row where both first_label_excel and second_label_excel are exact matches.
        Uses the rightmost of those as the "second target" and writes values starting at
        the first unmerged empty cell to its right.
        SQL resource column is: resource-<sanitized second_label_excel>.
        In Excel, resource always goes to the 'Resource' column.
        """
        #pre-step sanitizing labels
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]

            # 2) Collect base + suffix columns for the main data
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                cand = f"{column_to_get}-{i}"
                if cand in all_cols:
                    matching_cols.append(cand)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column in SQL: resource-<sanitized second_label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(second_label_excel)
                cand_res = f"resource-{safe_label}"
                if cand_res in all_cols:
                    resource_sql_col = cand_res
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")
            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # 5) Find a row where BOTH labels exist as exact cell matches
            row_found = False
            for excel_row in ws_template.iter_rows():
                first_col_idx = None
                second_col_idx = None

                for cell in excel_row:
                    if isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == str(first_label_excel).strip():
                            first_col_idx = cell.column
                        elif cell_val == str(second_label_excel).strip():
                            second_col_idx = cell.column

                if first_col_idx is not None and second_col_idx is not None:
                    row_found = True
                    target_row = excel_row[0].row
                    second_target_col = max(first_col_idx, second_col_idx)

                    # 6) Find the first *unmerged* empty cell to the right
                    current_col = second_target_col + 1
                    while True:
                        cell = ws_template.cell(row=target_row, column=current_col)

                        # Skip merged cells (non top-left)
                        if isinstance(cell, MergedCell):
                            current_col += 1
                            continue

                        if cell.value in (None, ""):
                            break

                        current_col += 1

                    # 7) Write values, skipping merged/occupied cells
                    for val in value_cols:
                        if val in (None, ""):
                            continue

                        while True:
                            cell = ws_template.cell(row=target_row, column=current_col)

                            if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                                current_col += 1
                                continue

                            cell.value = val
                            print(
                                f"Inserted '{val}' into cell {cell.coordinate}"
                            )
                            current_col += 1
                            break

                    # 8) Put resource into Excel "Resource" column (header cell == 'Resource')
                    if include_resource and resource_val not in (None, ""):
                        resource_col_idx = None
                        for hdr_row in ws_template.iter_rows():
                            for hdr_cell in hdr_row:
                                if (
                                    isinstance(hdr_cell.value, str)
                                    and hdr_cell.value.strip().lower() == "resource"
                                ):
                                    resource_col_idx = hdr_cell.column
                                    break
                            if resource_col_idx is not None:
                                break

                        if resource_col_idx is not None:
                            ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                            print(
                                f"Inserted resource '{resource_val}' into cell "
                                f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                            )
                        else:
                            print("Could not find 'Resource' column in Excel to write resource value.")
                    break

            if not row_found:
                print(
                    f"No row found where both '{first_label_excel}' "
                    f"and '{second_label_excel}' are present as exact matches."
                )

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_excel_source_after_two_targets_OECD(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,      # base name, e.g. "source"
        lookup_column,
        lookup_value,
        first_label_excel,  # label of first target cell
        second_label_excel, # label of second target cell (same row as first)
        max_suffix=5,       # how far to look for -1, -2, ... columns
        include_resource=True
    ):
        """
        Finds a row where both first_label_excel and second_label_excel are exact matches.
        Uses the rightmost of those as the "second target" and writes values starting at
        the first unmerged empty cell to its right.

        To work with the function getting info from SQL database, the No data could not be left as an empty cell as was before
        so it is written as "no data", however then the program to get the same looking template just has to skip it, not to
        write "do data" twice
        If second_label_excel == "No data" (case-insensitive):
          - Skip writing the SQL values for column_to_get and its suffixes.
          - Still write the SQL resource value (if available) into the Excel "Resource" column.
        """
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]

            # 2) Collect base + suffix columns for the main data
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                cand = f"{column_to_get}-{i}"
                if cand in all_cols:
                    matching_cols.append(cand)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column in SQL: resource-<sanitized second_label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(second_label_excel)
                cand_res = f"resource-{safe_label}"
                if cand_res in all_cols:
                    resource_sql_col = cand_res
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")
            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # Skipping no data: decide whether to skip writing the SQL values
            skip_value_write = str(second_label_excel).strip().lower() == "no data"

            # 5) Find a row where BOTH labels exist as exact cell matches
            row_found = False
            for excel_row in ws_template.iter_rows():
                first_col_idx = None
                second_col_idx = None

                for cell in excel_row:
                    if isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == str(first_label_excel).strip():
                            first_col_idx = cell.column
                        elif cell_val == str(second_label_excel).strip():
                            second_col_idx = cell.column

                if first_col_idx is not None and second_col_idx is not None:
                    row_found = True
                    target_row = excel_row[0].row
                    second_target_col = max(first_col_idx, second_col_idx)

                    # 6) Find the first *unmerged* empty cell to the right
                    current_col = second_target_col + 1
                    while True:
                        cell = ws_template.cell(row=target_row, column=current_col)

                        # Skip merged cells (non top-left)
                        if isinstance(cell, MergedCell):
                            current_col += 1
                            continue

                        if cell.value in (None, ""):
                            break

                        current_col += 1

                    # 7) Write values, skipping merged/occupied cells
                    # Only write values if NOT "No data"
                    if not skip_value_write:
                        for val in value_cols:
                            if val in (None, ""):
                                continue

                            while True:
                                cell = ws_template.cell(row=target_row, column=current_col)

                                if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                                    current_col += 1
                                    continue

                                cell.value = val
                                print(
                                    f"Inserted '{val}' into cell {cell.coordinate}"
                                )
                                current_col += 1
                                break
                    else:
                        print("Second label is 'No data' → skipping SQL values, but still handling resource if present.")

                    # 8) Put resource into Excel "Resource" column (header cell == 'Resource')
                    if include_resource and resource_val not in (None, ""):
                        resource_col_idx = None
                        for hdr_row in ws_template.iter_rows():
                            for hdr_cell in hdr_row:
                                if (
                                    isinstance(hdr_cell.value, str)
                                    and hdr_cell.value.strip().lower() == "resource"
                                ):
                                    resource_col_idx = hdr_cell.column
                                    break
                            if resource_col_idx is not None:
                                break

                        if resource_col_idx is not None:
                            ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                            print(
                                f"Inserted resource '{resource_val}' into cell "
                                f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                            )
                        else:
                            print("Could not find 'Resource' column in Excel to write resource value.")
                    break

            if not row_found:
                print(
                    f"No row found where both '{first_label_excel}' "
                    f"and '{second_label_excel}' are present as exact matches."
                )

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def insert_image_under_label(ws_template: Worksheet, label_excel: str, image_name: str,image_dir: str):
        """
        Find the first cell whose value exactly matches 'label_excel'
        then insert the image from image_dir/image_name into the cell
        directly below that label cell.
        """

        # Build image path
        image_path = Path(image_dir) / image_name

        if not image_path.exists():
            print(f"Image not found: {image_path}")
            return

        # Normalize the target label once
        normalized_label = label_excel.strip().lower()

        # 1) Find the first exact match for the label
        label_cell = None
        for row in ws_template.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == normalized_label:
                    label_cell = cell
                    break
            if label_cell:
                break

        if not label_cell:
            print(f"Exact label '{label_excel}' not found.")
            return

        # Cell directly below the label cell
        target_row = label_cell.row + 1
        target_col = label_cell.column
        anchor_coord = ws_template.cell(row=target_row, column=target_col).coordinate

        # 2) Create and place the image
        img = XLImage(str(image_path))
        img.anchor = anchor_coord
        ws_template.add_image(img)

        print(
            f"Inserted image '{image_name}' at {anchor_coord} "
            f"under label '{label_excel}' in sheet '{ws_template.title}'."
        )

    def put_template_into_CPS(filepath, template_path):
        '''Puts the template into the CPS excel that exists or creates a new one'''

        def open_or_create_xlsm(filepath):
            """
            Opens an existing .xlsm file with macros preserved.
            If it doesn't exist, creates a new one.
            Returns the workbook object.
            """
            if not filepath.lower().endswith(".xlsm"):
                print(f"ERROR: Expected an .xlsm file, but got: {filepath}")
                return None
            if os.path.exists(filepath):
                # Load workbook and preserve macros
                wb = load_workbook(filepath, keep_vba=True)
                print(f"Opened existing file: {filepath}")
            else:
                # Create new workbook and save as xlsm
                wb = Workbook()
                wb.save(filepath)
                wb2 = load_workbook(filepath, keep_vba=True)
                wb2.save(filepath)
                print(f"Created new xlsm file: {filepath}")
            return wb

        def add_new_sheet(filepath, new_sheet_name):
            """
            Opens (or creates) an xlsm file and adds a new sheet.
            Saves the file afterwards.
            """
            wb = open_or_create_xlsm(filepath)

            # If sheet already exists, create a unique name
            if new_sheet_name in wb.sheetnames:
                base = new_sheet_name
                i = 1
                while f"{base}_{i}" in wb.sheetnames:
                    i += 1
                new_sheet_name = f"{base}_{i}"

            # Create the sheet as first sheet
            ws = wb.create_sheet(new_sheet_name, 0)

            wb.save(filepath)
            print(f"Added sheet '{new_sheet_name}' to {filepath}")

            return ws

        def rename_with_date_and_move_to_back(filepath, sheet_name, date_format="%Y_%m_%d"):
            """
            Renames 'sheet_name' to 'sheet_name_YYYY_MM_DD'
            and moves it to the back of the workbook.
            """
            # Load workbook safely
            if filepath.lower().endswith(".xlsm"):
                wb = load_workbook(filepath, keep_vba=True)
            else:
                wb = load_workbook(filepath)

            # Ensure sheet exists
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' does not exist in workbook.")

            ws = wb[sheet_name]

            # Create date suffix
            date_str = datetime.now().strftime(date_format)

            # Build new name
            new_name = f"{sheet_name}_{date_str}"

            # Apply the new title
            ws.title = new_name

            # Move to the back
            wb._sheets.remove(ws)
            wb._sheets.append(ws)

            # Save file
            wb.save(filepath)

            print(f"Renamed '{sheet_name}' → '{new_name}' and moved to back.")

        def load_wb_any(path):
            """Load xlsx/xlsm, preserving VBA if present."""
            if path.lower().endswith(".xlsm"):
                return load_workbook(path, keep_vba=True)
            return load_workbook(path)

        def copy_sheet_to_other_workbook(
                src_path,
                src_sheet_name,
                dest_path,
                dest_sheet_name
        ):
            """
            Copy sheet `src_sheet_name` from src_path into dest_path
            as a new sheet called `dest_sheet_name`.

            Copies:
              - cell values
              - styles
              - merged cells
              - column widths
              - row heights
              - existing data validation (drop-down menus)
              - existing conditional formatting that applies to column I
            """

            # --- Load source workbook ---
            if not os.path.exists(src_path):
                raise FileNotFoundError(f"Source file not found: {src_path}")
            src_wb = load_wb_any(src_path)

            if src_sheet_name not in src_wb.sheetnames:
                raise ValueError(f"Sheet '{src_sheet_name}' not found in source workbook.")

            src_ws = src_wb[src_sheet_name]

            # --- Load or create destination workbook ---
            if os.path.exists(dest_path):
                dest_wb = load_wb_any(dest_path)
            else:
                from openpyxl import Workbook
                dest_wb = Workbook()
                # Clear the default sheet
                default_sheet = dest_wb.active
                dest_wb.remove(default_sheet)

            # If sheet with that name already exists in dest, remove it or rename first
            if dest_sheet_name in dest_wb.sheetnames:
                dest_wb.remove(dest_wb[dest_sheet_name])

            # Create destination sheet (as first sheet)
            dest_ws = dest_wb.create_sheet(title=dest_sheet_name, index=0)

            # --- Copy column dimensions (width, hidden, etc.) ---
            for col_letter, col_dim in src_ws.column_dimensions.items():
                new_dim = dest_ws.column_dimensions[col_letter]
                new_dim.width = col_dim.width
                new_dim.hidden = col_dim.hidden
                new_dim.outlineLevel = col_dim.outlineLevel
                new_dim.bestFit = col_dim.bestFit

            # --- Copy row dimensions (height, hidden, etc.) ---
            for row_idx, row_dim in src_ws.row_dimensions.items():
                new_row_dim = dest_ws.row_dimensions[row_idx]
                new_row_dim.height = row_dim.height
                new_row_dim.hidden = row_dim.hidden
                new_row_dim.outlineLevel = row_dim.outlineLevel

            # --- Copy cell values and styles ---
            for row in src_ws.iter_rows():
                for cell in row:
                    dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                    if cell.has_style:
                        dest_cell._style = copy(cell._style)

                    dest_cell.data_type = cell.data_type
                    dest_cell.number_format = cell.number_format
                    dest_cell.protection = copy(cell.protection)
                    dest_cell.alignment = copy(cell.alignment)

            # --- Copy merged cells ---
            if src_ws.merged_cells.ranges:
                for merged_range in src_ws.merged_cells.ranges:
                    dest_ws.merge_cells(str(merged_range.coord))

            # --- Copy existing data validation (drop-down menus) ---
            if src_ws.data_validations is not None:
                for dv in src_ws.data_validations.dataValidation:
                    new_dv = copy(dv)
                    dest_ws.add_data_validation(new_dv)

            # --- Copy conditional formatting only for column I ---
            col_I_idx = 9  # column I

            # Iterate internal CF structures similar to the SO snippet
            for cf in src_ws.conditional_formatting._cf_rules:
                for rng in cf.cells.ranges:
                    cr = CellRange(rng.coord)

                    # If the CF range covers column I
                    if cr.min_col <= col_I_idx <= cr.max_col:
                        # Intersect with column I: keep row range, force column I
                        dst_range = f"I{cr.min_row}:I{cr.max_row}"

                        for rule in cf.cfRule:
                            dest_ws.conditional_formatting.add(dst_range, copy(rule))
            # --- Copy sheet protection settings (so locked cells stay locked) ---
            dest_ws.protection = copy(src_ws.protection)

            # --- Save destination workbook ---
            dest_wb.save(dest_path)
            print(
                f"Copied sheet '{src_sheet_name}' from '{src_path}' "
                f"to '{dest_path}' as '{dest_sheet_name}' "
                f"(with data validation + CF for column I)."
            )

        # move the old C2C to the back and add a new C2C sheet in the front
        rename_with_date_and_move_to_back(filepath, "C2Coverview")
        add_new_sheet(filepath, "C2Coverview")
        # copy the template from the template
        copy_sheet_to_other_workbook(
            src_path=template_path,
            src_sheet_name="C2Coverview",
            dest_path=filepath,
            dest_sheet_name="C2Coverview"
        )

    ### Start with extracting
    # make a new sheet in the CAS specific folder
    st.write(f"Working on {CAS}")
    filepath = f"{folder}/CPS_CAS {CAS}.xlsm"

    if not os.path.exists(filepath):
        print(f"Creating new CAS excel file: {filepath}")
        template_wb = load_workbook(template_path, read_only=False, keep_vba=True)
        ws_template = template_wb["C2Coverview"]
        template_wb.save(filepath)

    else:
        print(f"CAS excel file: {filepath} already exists")

    put_template_into_CPS(filepath, template_path)

    template_wb = load_workbook(filepath, read_only=False, keep_vba=True)
    ws_template = template_wb["C2Coverview"]
    try:
        ### SQL SET-UP
        connection = sqlite3.connect(database)
        cursor = connection.cursor()

        print("Connected to SQLite database at:", db_path)

        # GENERAL INFO
        #Add general info
        namesDBcol_gen = ["Chemical name", "Common name", "CAS number", "EC number", "Linked CAS Read across",
                       "Linked CAS Monomers", "Linked CAS Degradation Products"]
        namesExcel_gen = ["Chemical name", "Common name", "CAS number", "EC number", "Linked CAS Read across",
                       "Linked CAS Monomers", "Linked CAS Degradation Products"]
        for namesDBcol_gen, namesExcel_gen in zip(namesDBcol_gen,namesExcel_gen):
            db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="GENERALINFO", link_ref="ref",
                                   column_to_get=namesDBcol_gen, lookup_column="ID",lookup_value =CAS, label_excel=namesExcel_gen)
        # ADD ASSESSORS
        db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="ASSESSORS", link_ref="ref",
                                   column_to_get="Name assessor", lookup_column="ID", lookup_value=CAS,
                                   label_excel="Name assessor")
        db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="ASSESSORS", link_ref="ref",
                                   column_to_get="Date assessed", lookup_column="ID", lookup_value=CAS,
                                   label_excel="Date created/updated")
        ## Add various info CHEMICAL CLASS
        namesDBcol_CC = ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]
        namesExcel_CC = ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]
        for names_DB, name_EX in zip(namesDBcol_CC, namesExcel_CC):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHEMICALCLASS", link_ref="ref",
                                        column_to_get=names_DB, lookup_column="ID", lookup_value=CAS,
                                        label_excel=name_EX, offset=2)

        # Adding other info
        namesDBcol_OTHER = ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]
        namesExcel_OTHER = ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]
        for names_DB_OTHER, name_EX_OTHER in zip(namesDBcol_OTHER, namesExcel_OTHER):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="OTHERINFO", link_ref="ref",
                                        column_to_get=names_DB_OTHER, lookup_column="ID", lookup_value=CAS,
                                        label_excel=name_EX_OTHER, offset=2)
        #  OTHER CRITERIA
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="OCRIT", link_ref="ref",
                                    column_to_get="Other comments", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Other comments", offset=1)
        # CARCINOGENICITY
        namesDBcols = ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK", "Carcinogenicity Classified IARC",
                        "Carcinogenicity Classified TLV","Carcinogenicity experimental evidence","Carcinogenicity Comments"]
        namesExcel = ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK", "Carcinogenicity Classified IARC",
                        "Carcinogenicity Classified TLV","Carcinogenicity experimental evidence","Carcinogenicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcols,namesExcel):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CARCINOGENICITY", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1,max_suffix=5,include_resource=True)

        # ED
        namesDBcols = ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]
        namesExcel = ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]
        for namesDBcol, nameExcel in zip(namesDBcols,namesExcel):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ENDOCRINE", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # MUTAGENICITY
        namesDBcol_MUT = ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]
        namesExcel_MUT = ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_MUT,namesExcel_MUT):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="MUTAGENICITY", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # MUTAGENICITY OECD TESTS
        #Point mutations
        point_mut_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT",
                                                       link_ref="ref",
                                                       lookup_column="ID", lookup_value=CAS)
        point_mut_names_cleared = remove_text_from_string(point_mut_names, "Point mutations:")
        #print(point_mut_names)

        write_list_right_of_label(ws_template, "Point mutations:", 1, point_mut_names_cleared)

        for namesDB, nameExcel in zip(point_mut_names, point_mut_names_cleared):
            refdb_to_excel_source_after_two_targets_OECD(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel="Point mutations:", second_label_excel=nameExcel, max_suffix=5, include_resource=True )

        # Chromosome damaging
        ch_dam_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM",
                                                    link_ref="ref",
                                                    lookup_column="ID", lookup_value=CAS)
        ch_dam_names_cleared = remove_text_from_string(ch_dam_names, "Chromosome damaging:")
        #print(ch_dam_names)
        write_list_right_of_label(ws_template, "Chromosome damaging:", 1, ch_dam_names_cleared)

        for namesDB, nameExcel in zip(ch_dam_names, ch_dam_names_cleared):
            refdb_to_excel_source_after_two_targets_OECD(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel="Chromosome damaging:", second_label_excel=nameExcel, max_suffix=5, include_resource=True )


        # REPROTOX
        namesDBcol_REP = ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]
        namesExcel_REP = ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_REP,namesExcel_REP):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="REPROTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # DEVELOPMENTAL TOX
        namesDBcol_DEV = ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]
        namesExcel_DEV = ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_DEV,namesExcel_DEV):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="DEVELOPTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # NEUROTOX
        namesDBcol_NETOX = ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]
        namesExcel_NETOX = ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_NETOX,namesExcel_NETOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="NEUROTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # ORAL TOX
        namesDBcol_ORTOX = ["Oral toxicity Acute Tox classified","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]
        namesExcel_ORTOX = ["Oral toxicity Acute Tox classified:","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_ORTOX,namesExcel_ORTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ORALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # INHALE TOX

        namesDBcol_INHTOX = ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]
        namesExcel_INHTOX = ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_INHTOX,namesExcel_INHTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="INHALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # DERMAL TOX
        namesDBcol_DERMTOX = ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]
        namesExcel_DERMTOX = ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_DERMTOX,namesExcel_DERMTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="DERMALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SKIN/EYE IRRIT/COR
        namesDBcol_IRR = ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]
        namesExcel_IRR = ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_IRR,namesExcel_IRR):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="IRRITCOR", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SENSITISATION
        namesDBcol_SENS = ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]
        namesExcel_SENS = ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_SENS,namesExcel_SENS):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="SENSITISATION", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SPECIFIC CONCENTRATION LIMITS
        SCL_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                 link_ref="ref",
                                                 lookup_column="ID", lookup_value=CAS)
        # cleaning the names so there is only distinct SCL names
        SCL_names_clean = remove_text_from_string(SCL_names, " - Lower Limit: (%)")
        SCL_names_clean = remove_text_from_string(SCL_names_clean, " - Upper Limit: (%)")
        SCL_names_dist = list(dict.fromkeys(SCL_names_clean))
        #print(SCL_names_dist)
        write_list_right_of_label(ws_template, "Hazard classification:", 1, SCL_names_dist)
        # # Lower limit
        # choosing SQL with only Lower limit data
        SCL_DB_names_lower = [s for s in SCL_names if "Lower Limit:" in s]
        print(SCL_DB_names_lower)
        SCL_EX_names_lower = remove_text_from_string(SCL_DB_names_lower, " - Lower Limit: (%)")
        print(SCL_EX_names_lower)
        # extracting Lower limit data
        for namesDB, nameExcel in zip(SCL_DB_names_lower, SCL_EX_names_lower):
            refdb_to_excel_source_after_two_targets(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel=nameExcel, second_label_excel="Lower Limit: (%)", max_suffix=5, include_resource=True )

        # # Upper limit
        # choosing SQL with only Lower limit data
        SCL_DB_names_upper = [s for s in SCL_names if "Upper Limit:" in s]
        SCL_EX_names_upper = remove_text_from_string(SCL_DB_names_upper, " - Upper Limit: (%)")
        # extracting Lower limit data
        for namesDB, nameExcel in zip(SCL_DB_names_upper, SCL_EX_names_upper):
            refdb_to_excel_source_after_two_targets(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                    link_ref="ref",
                                                    column_to_get=namesDB, lookup_column="ID", lookup_value=CAS,
                                                    first_label_excel=nameExcel, second_label_excel="Upper Limit: (%)",
                                                    max_suffix=5, include_resource=True)

        # AQUATIC TOXICITY

        namesDBcol_AQTOX = ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]
        namesExcel_AQTOX = ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]
        for namesDBcol, nameExcel in zip(namesDBcol_AQTOX,namesExcel_AQTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="AQUATOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # VERTEBRATE FISH
        namesDBcol_FISHTOX = ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]
        namesExcel_FISHTOX = ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_FISHTOX,namesExcel_FISHTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="FISHTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # INVERTEBRATE TOX
        namesDBcol_INVTOX = ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]
        namesExcel_INVTOX = ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_INVTOX,namesExcel_INVTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="INVTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # ALGAE TOX
        namesDBcol_ALGTOX = ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]
        namesExcel_ALGTOX = ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]
        for namesDBcol, nameExcel in zip(namesDBcol_ALGTOX, namesExcel_ALGTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ALGAETOX", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)

        # TERRESTRIAL TOX
        namesDBcol_TERTOX =  ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]
        namesExcel_TERTOX =  ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_TERTOX, namesExcel_TERTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="TERTOX", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)

        # OTHER SPECIES TOX
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="SPECTOX", link_ref="ref",
                                    column_to_get="Any other CLP species classification", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Any other CLP species classification", offset=1)
        # PERSISTENCE
        namesDBcol_PERS =  ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]
        namesExcel_PERS =  ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_PERS, namesExcel_PERS):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="PERSISTENCE", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        # BIOACCUMULATION
        namesDBcol_BIOAC =  ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]
        namesExcel_BIOAC =  ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_BIOAC, namesExcel_BIOAC):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="BIOACCUMULATION", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        # CLIMATIC RELEVANCE
        namesDBcol_CLIMREL =  ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]
        namesExcel_CLIMREL =  ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_CLIMREL, namesExcel_CLIMREL):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CLIMATICRELEVANCE", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        #  ADDITIONAL SOURCES
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ADDSOURCE", link_ref="ref",
                                    column_to_get="Additional sources", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Additional sources", offset=1)

        ### Image
        ex_label = "Molecular Formula or chemical picture"
        image_name = f"CPS_CAS {CAS}.png"

        insert_image_under_label(ws_template, ex_label, image_name, image_dir)

        ### SAVE THE FILLED IN CPS EXCEL ####
        name = f"Test CPS_CAS {CAS}.xlsm"
        saving_path = os.path.join(folder,name)
        template_wb.save(saving_path)

    except sqlite3.Error as e:
        st.write("SQLite error:", e)
def make_cas_report_excel(folder, *, base_name="Export_CAS", CASall=None, found=None, not_found=None, CAS_not_in_DB_but_in_excel=None,  CAS_not_in_DB_and_not_in_excel=None, CAS_older_than_3_years=None, CAS_needing_update=None, cas_with_up_to_date_info=None, cas_hazards=None, cas_with_no_json=None):
    """
    Creates an Excel report from the programme.
    Saves it automatically as Export_CAS_{date}.xlsx (or _v2, _v3... if needed).
    Returns the saved file path.
    """

    def _as_cell_text(value):
        """
        Converts function outputs to something Excel-friendly.
        - None -> ""
        - list/tuple/set -> newline-separated string
        - dict -> pretty key: value lines
        - everything else -> str(value)
        """
        if value is None:
            return ""
        if isinstance(value, dict):
            return "\n".join([f"{k}: {v}" for k, v in value.items()])
        if isinstance(value, (list, tuple, set)):
            return "\n".join([str(x) for x in value])
        return str(value)

    def get_unique_export_filename(folder, base_name="Export_CAS", ext=".xlsx"):
        date_str = datetime.now().strftime("%Y-%m-%d")

        # Initial filename
        filename = f"{base_name}_{date_str}{ext}"
        full_path = os.path.join(folder, filename)

        # If name exists, add _v2, _v3, ...
        version = 2
        while os.path.exists(full_path):
            filename = f"{base_name}_{date_str}_v{version}{ext}"
            full_path = os.path.join(folder, filename)
            version += 1

        return full_path

    os.makedirs(folder, exist_ok=True)
    out_path = get_unique_export_filename(folder, base_name=base_name, ext=".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "CAS Report"

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 70

    ws["A1"] = "Report of the CAS"
    ws["A1"].font = bold

    ws["A3"] = "Information"
    ws["A3"].font = bold
    ws["B3"] = "CAS"
    ws["B3"].font = bold

    ws["A4"] = "CAS analysed:"
    ws["B4"] = _as_cell_text(CASall)

    ws["A5"] = "CAS found in database:"
    ws["B5"] = _as_cell_text(found)

    ws["A6"] = "CAS not in database:"
    ws["B6"] = _as_cell_text(not_found)

    ws["A7"] = "New excel CPS added to DB for:"
    ws["B7"] = _as_cell_text(CAS_not_in_DB_but_in_excel)

    ws["A8"] = "CAS missing, need to be made:"
    ws["A8"].font = Font(bold=True, color="FF0000")
    ws["B8"] = _as_cell_text(CAS_not_in_DB_and_not_in_excel)
    ws["B8"].font = Font(bold=True, color="FF0000")

    ws["A9"] = "CPS older than 3 years for:"
    ws["A9"].font = Font(bold=True, color="FF0000")
    ws["B9"] = _as_cell_text(CAS_older_than_3_years)
    ws["B9"].font = Font(bold=True, color="FF0000")

    ws["A10"] = "CPS updated with new info from excel:"
    ws["B10"] = _as_cell_text(CAS_needing_update)

    ws["A11"] = "CnL info check"
    ws["A11"].font = bold

    ws["A12"] = "CnL info that is up to date:"
    ws["B12"] = _as_cell_text(cas_with_up_to_date_info)

    ws["A13"] = "CnL info is dfferent than in CPS for:"
    ws["A13"].font = Font(bold=True, color="FF0000")
    ws["B13"] = _as_cell_text(cas_hazards)
    ws["B13"].font = Font(bold=True, color="FF0000")

    ws["A14"] = "CnL info was not found for:"
    ws["B14"] = _as_cell_text(cas_with_no_json)

    for row in range(1, 16):
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                cell.alignment = wrap

    wb.save(out_path)
    return out_path


#### Create/update C2C database with CAS numbers from Excel files ####
# if connceted to database you can press run
if os.path.isfile(db_path):

    # find all other directories
    project_root = os.path.dirname(os.path.dirname(db_path))  # goes to folder new_DB_tests

    # save api key
    API_key = os.path.join(project_root, "Streamlit info", "NextSDS API key.txt")
    folder_excels = os.path.join(project_root, "CPS")
    save_json_dirr = os.path.join(project_root, "JSON")
    folder = folder_excels
    image_dir = os.path.join(project_root, "Chem_image")
    template_path = os.path.join(project_root, "Template", "CPS_CAS TEMPLATE V2.xlsm")
    folder_for_saving = os.path.join(project_root, "Downloads from Streamlit")
    folder_for_saving_excel_exports = os.path.join(project_root, "Downloads from Streamlit", "Report exports")
    folder_for_saving_CPS = os.path.join(project_root, "Downloads from Streamlit", "CPS downloads")
    db_backup_for_saving = os.path.join(project_root, "Database", "Backups")

    if st.button(":green[Run the code to screen selected CAS]"):
        ### Beginning: before starting download all available json files and make a backup of the DB
        st.success(":blue[Creating a DB back up.]")
        make_a_backup(db_path, db_backup_for_saving)
        st.success(":green[Back up complete.]")
        # json files download (first step to get new CnL info):
        st.success(":blue[Searching CnL website.]")
        CnL_json = check_json(CASall, API_key, save_json_dirr)
        if CnL_json is None:
            st.success(f":green[CnL extracted for: {', '.join(CASall)}.]")

        ### if the CAS exists and make a list with existing cas and the ones that have to be created
        # found => CAS exists in DB
        # not_found => CAS not in DB
        found, not_found = checking_if_CAS_exists(CASall, db_path)
        st.success(":blue[Searching if CAS is in the database.]")
        if found != []:
            st.success(f"CAS found in database: {', '.join(found)}")
        else:
            st.success(f"None of the CAS found in database.")
        if not_found != []:
            st.success(f"CAS not in database: {', '.join(not_found)}")
        else:
            st.success(f"All CAS are in DB")

        ### for CAS numbers that are NOT in the database (not_found)
        if not_found != []:
            st.success(f":blue[For CAS not found in database {', '.join(not_found)}. Searching if CAS is in the Excel file.]")
            # check if there is an Excel file with the given CAS
            CAS_not_in_DB_but_in_excel, CAS_not_in_DB_and_not_in_excel = check_if_excel_is_in_folder(folder_excels, not_found)

            # for CAS found as an Excel file
            if CAS_not_in_DB_but_in_excel != []:
                st.write(f"CAS found as an Excel: {', '.join(CAS_not_in_DB_but_in_excel)}. They need to be added to DB")
                st.write(f"Updating DB to add {', '.join(CAS_not_in_DB_but_in_excel)}")
                extract_info_form_excel_to_DB(db_path, folder_excels, CAS_not_in_DB_but_in_excel)
                st.success(f"DB updated successfully with {', '.join(CAS_not_in_DB_but_in_excel)}")
            else: st.success("There are no new Excel files for CAS not found in database.")
            # for CAS not in DB and not in Excel
            if CAS_not_in_DB_and_not_in_excel != []:
                st.success(f":red[CAS missing from DB and not in the Excel files: {', '.join(CAS_not_in_DB_and_not_in_excel)}. They need to be added to DB.]")
                ### LATER OPTION TO CREATE CPS FILES FROM CNL
                # st.write(f"Creating {', '.join(CAS_not_in_DB_and_not_in_excel)} as Excel files")
                # for CAS in CAS_not_in_DB_and_not_in_excel:
                #     extraction_info_excels(db_path, template_path, CAS, folder_for_saving_CPS, image_dir)
                #     st.toast(f"Excel file for {CAS} saved in your folder. Check it!")
                #     st.write(f"Excel file saved {CAS} in your folder.")
            else:
                st.success("There are no CAS missing from DB and not in the Excel files.")

        ### for CAS numbers that are in the database (found)
        if found != []:
            st.success(f":blue[For CAS found in database {', '.join(found)}. Searching if CAS is in the Excel file and if the info is up to date.]")
            # check if there is an Excel file with the given CAS
            CAS_in_folder, CAS_not_in_folder = check_if_excel_is_in_folder(folder_excels, found)
            if CAS_in_folder != []:
                st.write(f"Excel with CAS found for: {', '.join(CAS_in_folder)}")
                CAS_needing_update, CAS_older_than_3_years = is_DB_data_up_to_date_with_excel(db_path, folder_excels, CAS_in_folder)
                if CAS_older_than_3_years != []:
                    st.success(f":red[CAS older than 3 years: {', '.join(CAS_older_than_3_years)}]")
                if CAS_needing_update != []:
                    CAS_not_needing_update = [cas for cas in CAS_in_folder if cas not in CAS_needing_update]
                    st.success(f"CAS needing update: {', '.join(CAS_needing_update)}. CAS up to date: {', '.join(CAS_not_needing_update)} with info from Excel")
                    st.write(f"CAS needing update: {', '.join(CAS_needing_update)} will be updated now.")
                    extract_info_form_excel_to_DB(db_path, folder_excels, CAS_needing_update)
                    st.success(f"DB updated successfully with: {', '.join(CAS_needing_update)}")
                else:
                    if CAS_older_than_3_years == []:
                        st.success("All CAS are up to date with info from Excel.")
                    else:
                        st.success(f":red[Make sure to check {', '.join(CAS_older_than_3_years)} - old]")

            if CAS_not_in_folder != []:
                st.success(f":red[Those CAS are in DB but not as Excel files: {', '.join(CAS_not_in_folder)}. Making an Excel file in the folder]")
                # creating an Excel with info that is in the DB
                st.write("Generating an Excel file")
                for CAS in CAS_not_in_folder:
                    extraction_info_excels(db_path, template_path, CAS, folder_for_saving_CPS, image_dir)
                    st.toast("Excel file saved in your folder. Check it!")
                    st.write("Excel file saved in your folder.")

        ### Checking info from ECHA CnL
        # add together all the CAS for which CnL will be checked
        all_CAS_to_check_CnL = found + CAS_not_in_DB_but_in_excel
        st.success(f":blue[CnL info will be checked]")
        st.success(f"CnL info will be checked for: {', '.join(all_CAS_to_check_CnL)}")
        cas_hazards_list, cas_with_no_json = insert_json_info_to_DB(CnL_json, db_path, all_CAS_to_check_CnL)
        cas_hazards_updated = list(cas_hazards_list.keys())
        cas_with_up_to_date_info = [cas for cas in all_CAS_to_check_CnL if cas not in cas_hazards_updated]
        cas_with_up_to_date_info = [cas for cas in cas_with_up_to_date_info if cas not in cas_with_no_json]
        if cas_with_no_json != []:
            st.success(f":red[Data not checked for : {', '.join(cas_with_no_json)} (no info CnL file)]")
        if cas_with_up_to_date_info != []:
            st.success(f"CAS that have CnL info up to date: {', '.join(cas_with_up_to_date_info)}")
        if cas_hazards_updated != []:
            st.success(f":red[CnL info changed for: {', '.join(cas_hazards_updated)}]")
            for k, v in cas_hazards_list.items():
                st.write(f":red[{k},{v}]")

        out_file = make_cas_report_excel(
            folder= folder_for_saving_excel_exports,
            CASall=CASall,
            found=found,
            not_found=not_found,
            CAS_not_in_DB_but_in_excel=CAS_not_in_DB_but_in_excel,
            CAS_not_in_DB_and_not_in_excel=CAS_not_in_DB_and_not_in_excel,
            CAS_older_than_3_years=CAS_older_than_3_years,
            CAS_needing_update=CAS_needing_update,
            cas_with_up_to_date_info =cas_with_up_to_date_info,
            cas_hazards=cas_hazards_list,
            cas_with_no_json=cas_with_no_json,
        )
        st.success(f"Excel file with the documentation saved: {out_file}")

# saving the DB as excel file if needed
if os.path.isfile(db_path):

    # find all other directories
    project_root = os.path.dirname(os.path.dirname(db_path))  # goes to folder new_DB_tests

    # save api key
    API_key = os.path.join(project_root, "Streamlit info", "NextSDS API key.txt")
    folder_excels = os.path.join(project_root, "CPS")
    save_json_dirr = os.path.join(project_root, "JSON")
    folder = folder_excels
    image_dir = os.path.join(project_root, "Chem_image")
    template_path = os.path.join(project_root, "Template", "CPS_CAS TEMPLATE V2.xlsm")
    folder_for_saving = os.path.join(project_root, "Downloads from Streamlit")
    DB_excel_saving_path = folder_for_saving

    if st.button(":green[Do you want to make a current DB info as an Excel file?]"):
        st.write(f"DB info will be checked for: {db_path}")
        save_DB_to_excel(db_path, DB_excel_saving_path)
        st.toast("Excel file saved in your folder. Check it!")






