import streamlit as st
import pandas as pd
import sqlite3
import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
import traceback
import zipfile
import re
import requests
from bs4 import BeautifulSoup
from io import BytesIO, StringIO
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import logging
import zipfile
import random
import time
from datetime import datetime
import json
import copy

from streamlit import empty

st.title("ARCHE CAS database")
# text
st.markdown(
    """ 
    Checks and updates the CAS database.
    But before anything:
    """)
# green color:)
st.markdown("""
    <style>
    .stApp {
        background-color: #90EE90;
    }
    </style>
""", unsafe_allow_html=True)
# useless button
if st.button(":green[Press this button if you need some cheering up]"):
    st.balloons()
    st.toast("You can do this!", icon="🎉")
st.markdown("""
<style>
div.stButton > button {
    background-color: #0BDA51;     
    color: white;                 
    border-radius: 8px;
    border: none;
    padding: 0.6em 1.2em;
}
/* HOVER */
div.stButton > button:hover {
    background-color: #45a049;
    color: white;
}
""", unsafe_allow_html=True)
# CAS file upload
st.header(
    """ 
    Upload excel file with CAS numbers
    """)
uploaded_file = st.file_uploader("Upload Excel file with CAS: A column with name CAS containing all CAS/EC numbers to screen in individual rows below. This should be on the first sheet.", type=["xlsx", "xlsm"])
database_location = st.file_uploader("Upload a text file with database location in .txt", type=["txt"])
#st.selectbox
# save api key
API_key = '/Users/juliakulpa/Desktop/Test_echa/NextSDS API key.txt'
# folder with excels
folder_excels = "/Users/juliakulpa/Desktop/test/CPS"
# directory to save JSON
save_json_dirr = "/Users/juliakulpa/Desktop/test/JSON"


# Uploading the excel with CAS numbers
if uploaded_file is not None:
    # write if the file was uploaded
    CASallpd = pd.read_excel(uploaded_file)
    if "CAS" in CASallpd.columns:
        CASall = [cas.strip() for cas in CASallpd['CAS'].dropna().tolist()]  # Also remove white spaces per CAS
    else:
        st.success("The 'CAS' column was not found in the Excel file.")

    # CASall =["37872-24-5", "8028-89-5"]
    formatted_cas = [{"casNumber": cas, "ecNumber": ""} for cas in CASall]
    formatted_ec = [{"casNumber": "", "ecNumber": ec} for ec in CASall]

    st.success(f"Uploaded excel file with {len(CASall)} CAS numbers: {', '.join(CASall)}")
#starting with no path and then uploading it
db_path = 0
# Uploading the database location
if database_location is not None:
    db_path = database_location.read().decode("utf-8").strip()
    if os.path.isfile(db_path):
        st.success(f"Uploaded database directory: {db_path}")
    else:
        st.success(f":red[Database directory does not exist!!! Run button will only appear if you put the correct database directory]")

### FUNCTIONS
def make_a_backup(db_path):
    try:
        connection = sqlite3.connect(db_path)
        st.write("Connected to SQLite database:", db_path)

        # Create date-stamped backup filename
        today = datetime.now().strftime("%Y-%m-%d")
        base, ext = os.path.splitext(db_path)
        backup_path = f"{base}_backup_from_{today}{ext}"

        # Create backup connection
        backup_conn = sqlite3.connect(backup_path)

        # Perform the backup
        with backup_conn:
            connection.backup(backup_conn)

        st.write("Backup made:", backup_path)

    finally:
        if 'backup_conn' in locals():
            backup_conn.close()
        if connection:
            connection.close()
            st.write("Connection closed.")
def check_json(CASall, API_key, save_json_dirr):
    st.write('Checking API')

    # Load the API key from file: It's on the dropbox under Science/Data searches/ED screener/input databases/NextSDS API key.txt
    with open(API_key) as creds:
        api_key = creds.readlines()[0]  # API key is on the first line

    # Split up list in smaller parts (chunks)
    def chunk_list(lst, chunk_size):
        for i in range(0, len(lst), chunk_size):
            yield lst[i:i + chunk_size]

    start_url = "https://api.nextsds.com/jobs/start"
    status_url = "https://api.nextsds.com/jobs/retrieve"
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    # Step 1: Submit all jobs
    jobs = []
    for idx, cas_chunk in enumerate(chunk_list(CASall, 250)):
        data = {
            "taskId": "echa-api",
            "payload": cas_chunk
        }
        try:
            response = requests.post(start_url, headers=headers, json=data)
            if response.status_code == 200:
                job_id = response.json().get("id")
                jobs.append({"id": job_id, "index": idx + 1, "done": False, "output": None})
                st.write(f"Chunk {idx + 1}: Job submitted successfully: {job_id}")
            else:
                st.write(f"Chunk {idx + 1}: Failed to submit job")
        except Exception as e:
            st.write(f"Chunk {idx + 1}: Exception during job submission: {str(e)}")

    # Step 2: Monitor all jobs in one loop
    while not all(job["done"] for job in jobs):
        time.sleep(10)
        for job in jobs:
            if job["done"]:
                continue
            try:
                status_response = requests.get(f"{status_url}/{job['id']}", headers=headers)
                if status_response.status_code == 200:
                    status_data = status_response.json()
                    job_status = status_data.get("status")
                    st.write(f"Chunk {job['index']}: Job status: {job_status}")
                    if job_status not in ["STARTED", "EXECUTING"]:
                        job["done"] = True
                        job["output"] = status_data.get("output", [])
                elif status_response.status_code in [400, 404]:
                    st.write(f"Chunk {job['index']}: Job error ({status_response.status_code})")
                    job["done"] = True
            except Exception as e:
                st.write(f"Chunk {job['index']}: Exception during status check: {str(e)}")

    # Step 3: Combine all outputs
    CnL_json = []
    for job in jobs:
        if job["output"]:
            CnL_json.extend(job["output"])

    # Save to a JSON file
    exportpath = os.path.join(save_json_dirr,"output")
    st.write("Save to:",exportpath)
    if not os.path.exists(exportpath):
        os.makedirs(exportpath)
    formatted_time = datetime.now().strftime("%Y-%m-%d %H-%M")  # Customize format as needed
    exportjson = os.path.join(exportpath, "CnLscreener exportJSON " + formatted_time +".json")
    with open(exportjson, "w") as json_file:
        json.dump(CnL_json, json_file, indent=2)
    return CnL_json
def checking_if_CAS_exists(CASall, db_path):
    found = []
    not_found = []

    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()

        for cas in CASall:
            cursor.execute("SELECT 1 FROM C2C_DATABASE WHERE ID = ?", (cas,))
            row = cursor.fetchone()

            if row:
                found.append(cas)
            else:
                not_found.append(cas)

    except sqlite3.Error as e:
        st.write("SQLite error:", e)

    finally:
        if 'connection' in locals():
            connection.close()

    return found, not_found
def check_if_excel_is_in_folder(folder_excels, CAS_list):
    CAS_in_folder = []
    CAS_not_in_folder = []

    file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
    cas_pattern = re.compile(r'CAS (\d{2,7}[-‐-–—]\d{2,3}[-‐-–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
    ec_pattern = re.compile(r'EC (\d{2,7}[-‐-–—]\d{3}[-‐-–—]\d{1})')

    # collect all inventory numbers (CAS) found in the folder
    inv_in_folder = set()

    for filename in os.listdir(folder_excels):
        full_path = os.path.join(folder_excels, filename)
        if os.path.isfile(full_path):
            match = file_pattern.search(filename)
            if match:
                # Extract CAS inventory number (inv_number)
                match_inv = cas_pattern.search(filename)
                if match_inv:
                    inv_number = match_inv.group(1)  # this is the CAS
                    inv_in_folder.add(inv_number)
                else:
                    # if no CAS, then check for EC (kept from your code)
                    match_inv = ec_pattern.search(filename)
                    if match_inv:
                        inv_number = match_inv.group(1)
                    else:
                        st.write(f"Issue with: {filename}")

    # compare input CAS_list against what was found in folder
    for cas in CAS_list:
        if cas in inv_in_folder:
            CAS_in_folder.append(cas)
        else:
            CAS_not_in_folder.append(cas)

    return CAS_in_folder, CAS_not_in_folder
def insert_json_info_to_DB(CnL_json, db_path, target_cas_list):
    data = CnL_json
    cas_hazards = {} # used later to create a list of things to update
    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
     #id INTEGER PRIMARY KEY AUTOINCREMENT,
        # Ensure ECHACHEM_CL table exists
        cursor.execute("PRAGMA foreign_keys = ON;")
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS ECHACHEM_CL (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL,
            on_cl TEXT,
            cas TEXT,
            ec TEXT,
            name_echachem TEXT,
            type_classification TEXT,
            hazards TEXT,
            date_checked TEXT,
            FOREIGN KEY (code) REFERENCES C2C_DATABASE(ID)
        )
        """)


        st.write("Connected to SQLite database at:", db_path)

        today = datetime.now().date()

        for target_cas in target_cas_list:

            # Find the entry for the CAS you want
            entry = next((e for e in data if e.get("casNumber") == target_cas), None)

            if entry is None:
                st.write(f"CAS {target_cas} not found in JSON.")
            else:
                # Set up dictionary to collect all relevant info
                sqlinfo = {
                    "code": entry.get("casNumber"),
                    "on_cl": "-",
                    "cas": "-",
                    "ec": "-",
                    "name_echachem": "-",
                    "type_classification": "-",
                    "hazards": "-"
                }

                st.write(f"Testing for: {entry.get('casNumber')}")

                #### ECHA-CHEM C&L from NEXTSDS-API ####
                if entry.get("found") is False:  # If the chemical was NOT found on C&L
                    sqlinfo["on_cl"] = "No"
                else:  # If the chemical was found on C&L
                    sqlinfo["on_cl"] = "Yes"
                    sqlinfo["cas"] = entry.get("cas")
                    sqlinfo["ec"] = entry.get("ecNumber")
                    sqlinfo["name_echachem"] = entry.get("name")

                    if entry.get("isHarmonized") is True:
                        sqlinfo["type_classification"] = "Harmonized"
                    else:
                        sqlinfo["type_classification"] = "Self-classification"

                    # Safe hazards extraction (prevents crashes if hazards is missing/not a dict)
                    hazards = entry.get("hazards", {})
                    if isinstance(hazards, dict):
                        sqlinfo["hazards"] = hazards.get("hazardClasses", "-")
                #st.write(sqlinfo)

        #     with open(CnL_json, "r", encoding="utf-8") as f:
        #         data = json.load(f)
        #
        #         # If JSON is a list of entries
        #         for entry in data:
        #             print(entry)
        #         # Set up dictionary to collect all relevant info
        #         sqlinfo = {"code": entry.get("casNumber"), "on_cl": "-", "cas": "-", "ec": "-", "name_echachem": "-",
        #                    "type_classification": "-", "hazards": "-"}
        #         print(f"Adding chemical: {entry.get("casNumber")}")
        #
        #         #### ECHA-CHEM C&L from NEXTSDS-API ####
        #         if entry.get("found") == False:  # If the chemical was NOT found on C&L
        #             sqlinfo["on_cl"] = "No"
        #         else:  # If the chemical was found on C&L (then there is no "found" entry)
        #             sqlinfo["on_cl"] = "Yes"
        #             sqlinfo["cas"] = entry.get("cas")
        #             sqlinfo["ec"] = entry.get("ecNumber")
        #             sqlinfo["name_echachem"] = entry.get("name")
        #             if entry.get("isHarmonized") == True:
        #                 sqlinfo["type_classification"] = "Harmonized"
        #             else:
        #                 sqlinfo["type_classification"] = "Self-classification"
        #             sqlinfo["hazards"] = entry.get("hazards")["hazardClasses"]
        #
                # Check if CAS already exists
                cursor.execute("SELECT 1 FROM ECHACHEM_CL WHERE cas = ?", (sqlinfo["cas"],))
                exists = cursor.fetchone()
                if exists:
                    st.write(f"CAS {sqlinfo['cas']} already in database")
                    cursor.execute("SELECT 1 FROM ECHACHEM_CL WHERE cas = ? AND hazards = ?", (sqlinfo["cas"],sqlinfo["hazards"]))
                    same_hazard = cursor.fetchone()
                    if same_hazard:
                        st.write(f"Hazards for {sqlinfo['cas']} are the same as for the last update. NO ACTION NEEDED.")
                    else:
                        st.write(f"Inserting CAS {sqlinfo['cas']}...")
                        cursor.execute("""
                            UPDATE ECHACHEM_CL
                            SET hazards = ?, date_checked = ?
                            WHERE cas = ?
                        """, (
                            sqlinfo["hazards"],
                            today,
                            sqlinfo["cas"]
                        ))
                        connection.commit()
                        st.write(f"Hazards for {sqlinfo['cas']} are DIFFERENT as for the last update. INFO IN TABLE CnL UPDATED. ACTION REQUIRED.")

                        ### Needed for the next step to gather info and update
                        cursor.execute("SELECT hazards FROM ECHACHEM_CL WHERE code = ?",
                                       (sqlinfo["code"],))
                        # needed if info was added
                        row = cursor.fetchone()
                        hazards_list = row[0].split(",") if row and row[0] else []
                        cas_hazards[target_cas] = hazards_list
                        st.write(hazards_list)


                else:
                    st.write(f"CAS not in CnL database: {sqlinfo['cas']}")
                    cursor.execute(
                        "SELECT 1 FROM C2C_DATABASE WHERE ID = ?",
                        (sqlinfo["code"],)
                    )
                    exists = cursor.fetchone()
                    if not exists:
                        cursor.execute(
                            "INSERT INTO C2C_DATABASE (ID) VALUES (?)",
                            (sqlinfo["code"],)
                        )
                        st.write(f"CAS was not in the main C2C database. CAS added to the main C2C database: {sqlinfo['cas']}")
                    else:
                        st.write(f"CAS already exists in the main C2C database: {sqlinfo['cas']}")

                    cursor.execute("""
                        INSERT INTO ECHACHEM_CL (code, on_cl, cas, ec, name_echachem, type_classification, hazards, date_checked)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        sqlinfo["code"],
                        sqlinfo["on_cl"],
                        sqlinfo["cas"],
                        sqlinfo["ec"],
                        sqlinfo["name_echachem"],
                        sqlinfo["type_classification"],
                        sqlinfo["hazards"],
                        today
                    ))
                    connection.commit()
                    st.write(f"Information inserted to CnL database: {sqlinfo['cas']}")
                    ### Needed for the next step to gather info and update
                    cursor.execute("SELECT hazards FROM ECHACHEM_CL WHERE code = ?",
                                                 (sqlinfo["code"],))
                    #needed if info was added
                    row = cursor.fetchone()
                    hazards_list = row[0].split(",") if row and row[0] else []
                    cas_hazards[target_cas] = hazards_list
                    st.write(hazards_list)


        st.write(f"SQL updated, in case hazards were changed, they are here: {cas_hazards}")
        return cas_hazards
    #
    finally:
        if connection:
            connection.commit()
            connection.close()
            st.write("Connection closed.")
def is_DB_data_up_to_date_with_excel(db_path, folder_excels, CAS_list):
    excel_files_that_need_updating = []
    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
        st.write("Connected to SQLite database at", db_path)

        cursor.execute('''
                CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                    ID TEXT PRIMARY KEY,
                    LastUpdate TEXT,
                    FileName TEXT,
                    Comments TEXT
                )
            ''')
        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        for filename in os.listdir(folder_excels):
            full_path = os.path.join(folder_excels, filename)
            if not os.path.isfile(full_path):
                continue
            if not file_pattern.search(filename):
                continue
            # File date
            mod_time = os.path.getmtime(full_path)
            last_update = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d")

            # Extract CAS or EC
            match_inv = cas_pattern.search(filename)
            if match_inv is None:
                st.write("There should be something here. Please check.")

            if match_inv:
                inv_number = match_inv.group(1)
                comments = "CAS"
                if match_inv.group(2):
                    comments = "CAS, " + match_inv.group(2)
                #check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_list:
                    continue
            else:
                match_inv = ec_pattern.search(filename)
                if not match_inv:
                    st.write(f"Cannot extract CAS/EC from: {filename}")
                    continue
                inv_number = match_inv.group(1)
                comments = "EC"
                # check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_list:
                    continue

            cursor.execute("""
                    UPDATE C2C_DATABASE
                    SET LastUpdate = NULL
                    WHERE LastUpdate NOT LIKE '____-__-__'
                """)

            # Check what DB currently has
            cursor.execute("SELECT LastUpdate FROM C2C_DATABASE WHERE ID = ?", (inv_number,))
            row = cursor.fetchone()

            if row is None:
                # ID not in DB -> insert
                cursor.execute(
                    "INSERT INTO C2C_DATABASE (ID, LastUpdate, FileName, Comments) VALUES (?,?,?,?)",
                    (inv_number, last_update, filename, comments)
                )
                st.write(f"CHANGED (was not there before): inserted {inv_number}: set LastUpdate={last_update}")
                excel_files_that_need_updating.append(inv_number)

            else:
                db_last_update = row[0]  # ISO YYYY-MM-DD

                if db_last_update is None or db_last_update < last_update:
                    # File is newer -> update
                    cursor.execute(
                        "UPDATE C2C_DATABASE SET LastUpdate = ?, FileName = ?, Comments = ? WHERE ID = ?",
                        (last_update, filename, comments, inv_number)
                    )
                    st.write(f"CHANGED: inserted {inv_number}: {db_last_update} -> {last_update}")
                    excel_files_that_need_updating.append(inv_number)

                else:
                    # DB is newer or same -> skip
                    st.write(f"NO ACTION NEEDED. CAS that are up to date {inv_number}: DB {db_last_update} >= file {last_update}.")
        return excel_files_that_need_updating

    finally:
        if connection:
            connection.commit()
            connection.close()
        st.write("Connection closed.")
def extract_info_form_excel_to_DB(db_path, folder_excels, CAS_needing_DB_update):
    #### CUSTOM FUNCTIONS ####
    def add_info_CPS_below(sheet, search_strings, maindatabase, newdatabase, mainID):
        """
        Finds each search string somewhere in the sheet (not necessarily in the first row),
        then reads the values directly below it column-wise.
        Collects one "row" of data across all search strings,
        and inserts into the SQLite database.

        Stops when *all* searched columns are empty in the same row.

        search_strings:
            - list: Excel labels are used as SQL column names.
            - dict: {Excel label: SQL column name}
        """

        # Handle both list and dict
        if isinstance(search_strings, str):
            search_strings = [search_strings]
        if isinstance(search_strings, list):
            mapping = {s: s for s in search_strings}  # Excel label -> same SQL name
        elif isinstance(search_strings, dict):
            mapping = search_strings  # Excel label -> custom SQL name
        else:
            raise TypeError("search_strings must be a str, list, or dict")

        # Locate the target cells (positions of the labels in Excel)
        col_positions = {}
        for excel_label in mapping.keys():
            found = None
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and excel_label.lower() in str(cell.value).lower():
                        found = cell
                        break
                if found:
                    break
            if found:
                col_positions[excel_label] = (found.column, found.row + 1)  # start below label

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'

        # Check if the table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                       (newdatabase,))
        table_exists = cursor.fetchone()

        if not table_exists:
            # Create table with auto-increment ID and extracted columns
            cols_def = ", ".join([f"{q(sql_col)} TEXT" for sql_col in mapping.values()])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            # Add missing columns if needed
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            for sql_col in mapping.values():
                if sql_col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(sql_col)} TEXT")

        # Iterate row by row until all searched columns are empty
        row_offset = 0
        while True:
            row_data = {}
            all_empty = True
            for excel_label, (col, start_row) in col_positions.items():
                cell = sheet.cell(row=start_row + row_offset, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    sql_col = mapping[excel_label]
                    row_data[sql_col] = cell.value
                    all_empty = False
            if all_empty:
                break  # stop when all searched columns are empty in the same row

            if newdatabase == maindatabase:
                # --- Case 1: Append into the main database table ---
                # Check if the row for this ID exists
                cursor.execute(
                    f"SELECT 1 FROM {q(maindatabase)} WHERE ID = ?",
                    (mainID,)
                )
                exists = cursor.fetchone()

                if exists:
                    # Update only the new columns (append data)
                    update_clause = ", ".join([f"{q(col)} = ?" for col in row_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(maindatabase)} SET {update_clause} WHERE ID = ?",
                        list(row_data.values()) + [mainID]
                    )
                else:
                    # Insert new row with ID and these values
                    all_cols = ["ID"] + list(row_data.keys())
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(maindatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        [mainID] + list(row_data.values())
                    )

            else:
                # --- Case 2: Insert into child table ---
                row_data["ref"] = mainID

                # Build WHERE clause dynamically based on all row_data keys
                where_clause = " AND ".join([f"{q(col)} = ?" for col in row_data.keys()])
                params = list(row_data.values())

                cursor.execute(
                    f"SELECT 1 FROM {q(newdatabase)} WHERE {where_clause}",
                    params
                )
                exists = cursor.fetchone()

                if not exists:
                    # Insert new row
                    all_cols = list(row_data.keys())
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        list(row_data.values())
                    )

            row_offset += 1
    def add_info_CPS_one_cell_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
        """
        Finds rows containing `rowlabel`, extracts specified columns to the right,
        and inserts or updates the data in the SQLite database.

        Parameters:
            sheet: openpyxl worksheet
            rowlabel: string to search for in any row
            column_offsets: list of integers (e.g., [2, 3]) for columns to the right
            column_names: list of strings for custom SQL column names
            maindatabase: name of the main database (for foreign key reference)
            newdatabase: name of the table to update
            mainID: unique identifier for the row
        """

        if len(column_offsets) != len(column_names):
            raise ValueError("column_offsets and column_names must have the same length")

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'


        # Step 1: Find the row containing the rowlabel
        extracted_data = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                    row_idx = cell.row
                    col_idx = cell.column
                    for offset, col_name in zip(column_offsets, column_names):
                        target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                        extracted_data[col_name] = target_cell.value
                    break
            if extracted_data:
                break

        if not extracted_data:
            return  # nothing to insert

        # Step 2: Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        if not table_exists:
            # Create table with ID, ref, and extracted columns
            cols_def = ", ".join([f"{q(col)} TEXT" for col in column_names])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            # Add missing columns
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in column_names:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # Step 3: Insert or update
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                    [extracted_data[col] for col in column_names] + [mainID]
                )
            else:
                all_cols = ['ref'] + column_names
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in column_names]
                )

        else:  # when newdatabase == maindatabase
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                    [extracted_data[col] for col in column_names] + [mainID]
                )
            else:
                all_cols = ['ID'] + column_names
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in column_names]
                )
    def add_info_CPS_right_until_empty(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID, include_resource=True):
        """
        Like add_info_CPS_one_cell_right, but starting from column_offsets[0] to the right,
        keeps reading consecutive cells until it finds the first empty cell.
        Column naming:
          - first value uses column_names[0] (base)
          - next values use column_names[1:], if present
          - beyond that, auto-name as base-1, base-2, ...

        Optional behavior (when True):
          - If include_resource=True, captures the sheet's 'Resource' column value (if present and not empty)
            into SQL column 'resource-<sanitized rowlabel>'.
          - If no 'Resource' column exists or the cell is empty, skips creating that column.
        """

        if len(column_offsets) != len(column_names):
            raise ValueError("column_offsets and column_names must have the same length")
        if not column_offsets:
            return

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'

        # Sanitize rowlabel for safe SQL column naming
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        safe_rowlabel = sanitize_label(rowlabel)
        resource_colname = f"resource-{safe_rowlabel}"

        # --- locate the cell containing rowlabel ---
        match_row_idx = None
        match_col_idx = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                    match_row_idx = cell.row
                    match_col_idx = getattr(cell, "col_idx", cell.column)
                    break
            if match_row_idx is not None:
                break

        if match_row_idx is None:
            return  # nothing to insert

        # Optionally find the column index for "Resource"
        resource_col = None
        if include_resource:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == "resource":
                        resource_col = getattr(cell, "col_idx", cell.column)
                        break
                if resource_col:
                    break

        # Determine start offset and base name
        start_offset = column_offsets[0]
        base_name = column_names[0]

        # --- read to the right until the first empty cell ---
        extracted_data = {}
        k = 0
        max_col = sheet.max_column
        while (match_col_idx + start_offset + k) <= max_col:
            target = sheet.cell(row=match_row_idx, column=match_col_idx + start_offset + k)
            tv = target.value
            # stop at first empty/blank
            if tv is None or (isinstance(tv, str) and tv.strip() == ""):
                break

            # choose column name
            if k < len(column_names):
                col_name = column_names[k]
            else:
                col_name = f"{base_name}-{k - (len(column_names) - 1)}" if len(column_names) > 0 else f"col-{k}"

            extracted_data[col_name] = tv
            k += 1

        # Optionally add the Resource value, but only if it's not empty
        if include_resource and resource_col:
            resource_value = sheet.cell(row=match_row_idx, column=resource_col).value
            if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
                extracted_data[resource_colname] = resource_value

        if not extracted_data:
            return  # nothing to insert

        # --- ensure table exists and has needed columns ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- upsert logic (same keying as your working function) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                    [extracted_data[col] for col in needed_columns] + [mainID]
                )
            else:
                all_cols = ['ref'] + needed_columns
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in needed_columns]
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                    [extracted_data[col] for col in needed_columns] + [mainID]
                )
            else:
                all_cols = ['ID'] + needed_columns
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in needed_columns]
                )
    def loop_over_to_collect_right_values(sheet, rowlabel: str):
        """
        Finds the first cell whose value contains `rowlabel`, then collects the value
        in the cell to the right for that row and all consecutive rows below
        where the same label also appears in the same column.
        Returns a single string joining all collected right-hand values separated with commas.
        """

        # Case-insensitive: looking for upper or lower case
        def matches(value):
            if value is None:
                return False
            return str(value).strip().lower() == rowlabel.strip().lower()

        # Step 1: Find the first occurrence of the label
        start_row = None
        start_col = None
        for row in sheet.iter_rows():
            for cell in row:
                if matches(cell.value):
                    start_row = cell.row
                    start_col = getattr(cell, "col_idx", cell.column)
                    break
            if start_row is not None:
                break

        if start_row is None:
            return "No match with row label"

        # Step 2: Walk down while the label repeats and collect right-hand values
        collected = []
        r = start_row
        while r <= sheet.max_row:
            left_val = sheet.cell(row=r, column=start_col).value
            if not matches(left_val):
                break

            right_val = sheet.cell(row=r, column=start_col + 1).value
            if right_val is not None and str(right_val).strip() != "":
                collected.append(str(right_val).strip())

            r += 1

        # Step 2: Return as a single string
        return collected
    def add_info_CPS_from_row_with_two_markers(sheet, label1: str, label2: str, label3: str, label4: str, maindatabase, newdatabase, mainID):
        """
        label 1 - first row to match (e.g. Hazard classification)
        label 2 - second row to match (e.g. Eye Irrit. 2)
        label 3 & 4 looking in the row for them and taking values to the right
        1) Find a row where two adjacent cells match (label1, label2) left→right.
        2) In that row, find label3 and label4 (anywhere in the row), and for each:
           - take the value from the cell immediately to the right.
        3) Write to SQL columns named:
             - "{label3}-{label2}"  (for value next to label3)
             - "{label4}-{label2}"  (for value next to label4)
        Matching is case-insensitive 'contains'.
        Requires a DB cursor in outer scope named `cursor`.
        """

        # --- helpers ---
        def q(name: str) -> str:
            return f'"{name}"'

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        max_row = sheet.max_row
        max_col = sheet.max_column

        # --- 1) Find target row via adjacent (label1, label2) ---
        target_row = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = sheet.cell(row=r, column=c).value
                v2 = sheet.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    break
            if target_row is not None:
                break

        if target_row is None:
            return  # no matching row → nothing to insert

        # --- 2) Scan the row to find label3 and label4 targets; capture right-hand values ---
        extracted_data = {}

        # label3 → col name "{label3}-{label2}"
        col_name_3 = f"{label2} - {label3}"
        # label4 → col name "{label2}-{label4}"
        col_name_4 = f"{label2} - {label4}"

        # We search the entire row for each label, reading the cell to the right when found
        def capture_right_of_label(row: int, label: str):
            for c in range(1, max_col):  # up to max_col-1 to read c+1
                if matches(sheet.cell(row=row, column=c).value, label):
                    right_val = sheet.cell(row=row, column=c + 1).value
                    if right_val is None:
                        return None
                    if isinstance(right_val, str):
                        rv = right_val.strip()
                        return rv if rv != "" else None
                    return right_val
            return None

        val3 = capture_right_of_label(target_row, label3)
        val4 = capture_right_of_label(target_row, label4)

        if val3 is not None:
            extracted_data[col_name_3] = val3
        if val4 is not None:
            extracted_data[col_name_4] = val4

        if not extracted_data:
            return  # neither target produced a value → nothing to insert

        # --- 3) Ensure table/columns exist ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- 4) Upsert (same rules as your working pattern) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ref"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ID"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
    def add_info_right_two_markers_OECD(sheet, label1: str, label2: str, maindatabase, newdatabase, mainID, include_resource: bool = True):
        """
        label 1 - first row to match
        label 2 - second row to match
        1) Find a row where two adjacent cells match (label1, label2) left→right.
        2) Capture first non-empty cell to the right of label2.
        3) Write to SQL columns named: {label1}{label2}

        Optional behavior (when include_resource=True):
          - Captures the sheet's 'Resource' column value (if present and not empty)
            into SQL column 'resource-<sanitized label2>' for the same row.
          - If no 'Resource' column exists or the cell is empty, skips creating that column.
        """

        # --- helpers ---
        def q(name: str) -> str:
            return f'"{name}"'

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        # Sanitize label for safe SQL column naming (for resource column)
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        # --- NEW: if label2 is "no data", skip Excel and just write "no data" to SQL ---
        if label2.strip().lower() == "no data":
            safe_label2 = sanitize_label(label2)
            col_name = f"{label1}{label2}"   # same pattern as normal case
            extracted_data = {col_name: "no data"}

            # Ensure table/columns exist
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
            table_exists = cursor.fetchone()

            needed_columns = list(extracted_data.keys())

            if not table_exists:
                cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
                fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
                cursor.execute(f'''
                    CREATE TABLE {q(newdatabase)} (
                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                        ref TEXT
                        {"," if cols_def else ""} {cols_def}
                        {fk_clause}
                    )
                ''')
            else:
                cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
                existing_cols = [col[1] for col in cursor.fetchall()]
                if "ref" not in existing_cols and newdatabase != maindatabase:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
                for col in needed_columns:
                    if col not in existing_cols:
                        cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

            # Upsert
            if newdatabase != maindatabase:
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
                exists = cursor.fetchone()
                if exists:
                    set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                        list(extracted_data.values()) + [mainID]
                    )
                else:
                    cols = ["ref"] + list(extracted_data.keys())
                    placeholders = ", ".join(["?"] * len(cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                        [mainID] + list(extracted_data.values())
                    )
            else:
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
                exists = cursor.fetchone()
                if exists:
                    set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                        list(extracted_data.values()) + [mainID]
                    )
                else:
                    cols = ["ID"] + list(extracted_data.keys())
                    placeholders = ", ".join(["?"] * len(cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                        [mainID] + list(extracted_data.values())
                    )

            return  # done, no Excel lookup

        # --- Normal behavior below (when label2 is NOT "no data") ---

        max_row = sheet.max_row
        max_col = sheet.max_column

        safe_label2 = sanitize_label(label2)
        resource_colname = f"resource-{safe_label2}"

        # --- 1) Find target row via adjacent (label1, label2) ---
        target_row = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = sheet.cell(row=r, column=c).value
                v2 = sheet.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    break
            if target_row is not None:
                break

        if target_row is None:
            print("Target row not found")
            return  # no matching row → nothing to insert

        # --- Optionally find the column index for "Resource" ---
        resource_col = None
        if include_resource:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == "resource":
                        resource_col = getattr(cell, "col_idx", cell.column)
                        break
                if resource_col:
                    break

        # --- 2) Scan the row to find targets; capture right-hand values ---
        extracted_data = {}

        # label → col name
        col_name = f"{label1}{label2}"

        # move to look for the first value to the right
        def capture_right_of_label(row: int, label: str):
            # Find the column containing the label
            for c in range(1, max_col):
                cell_value = sheet.cell(row=row, column=c).value
                if matches(cell_value, label):

                    # Start searching to the right of this column
                    for cc in range(c + 1, max_col + 1):
                        right_val = sheet.cell(row=row, column=cc).value

                        # Skip empty or whitespace-only
                        if right_val is None:
                            continue

                        if isinstance(right_val, str):
                            rv = right_val.strip()
                            if rv == "":
                                continue
                            return rv  # return first non-empty string

                        # Non-string, non-None → return immediately
                        return right_val

                    # If no value was found to the right
                    return None

            # Label not found at all
            return None

        val = capture_right_of_label(target_row, label2)

        if val is not None:
            extracted_data[col_name] = val

        # Optionally add the Resource value, but only if it's not empty
        if include_resource and resource_col:
            resource_value = sheet.cell(row=target_row, column=resource_col).value
            if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
                extracted_data[resource_colname] = resource_value

        if not extracted_data:
            print("Target extracted not found")
            return  # nothing to insert

        # --- 3) Ensure table/columns exist ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- 4) Upsert (same rules as your working pattern) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ref"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ID"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )

    try:
        ### SQL SET-UP
        connection = sqlite3.connect(db_path)
        st.write("Connected to SQLite database at", db_path)

        # Create main C2C_DATABASE table if not existing
        cursor = connection.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                ID TEXT PRIMARY KEY,
                LastUpdate TEXT,
                FileName TEXT,
                Comments TEXT
            )
        ''')

        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        # Loop through Excel files with CAS number and add their info from the template
        for filename in os.listdir(folder_excels):
            full_path = os.path.join(folder_excels, filename)
            if not os.path.isfile(full_path):
                continue
            if not file_pattern.search(filename):
                continue
            # File date
            mod_time = os.path.getmtime(full_path)
            last_update = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d")

            # Extract CAS number or EC number if applicable
            match_inv = cas_pattern.search(filename)    # Check for CAS number
            comments = "There should be something here. Please check."
            if match_inv:
                inv_number = match_inv.group(1)
                comments = "CAS"
                if match_inv.group(2):
                    comments = "CAS, " + match_inv.group(2)
                #check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_needing_DB_update:
                    continue
            else:
                match_inv = ec_pattern.search(filename)
                if not match_inv:
                    st.write(f"Cannot extract CAS/EC from: {filename}")
                    continue
                inv_number = match_inv.group(1)
                comments = "EC"
                # check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_needing_DB_update:
                    continue
            # Check which inv you work with
            st.write(f"Updating CAS: {inv_number}")

            # Open the Excel file
            CPS_wb_obj = openpyxl.load_workbook(full_path)
            CPSsheet = CPS_wb_obj.active

            # Add general info
            for g_info in ["Chemical name","Common name","CAS number", "EC number", "Linked CAS Read across", "Linked CAS Monomers", "Linked CAS Degradation Products"]:
                add_info_CPS_below(CPSsheet, g_info,"C2C_DATABASE","GENERALINFO",inv_number )

            # ASSESSOR
            add_info_CPS_below(CPSsheet, {"Name assessor":"Name assessor","Date created/updated" : "Date assessed"},"C2C_DATABASE","ASSESSORS",inv_number)

            # Add various info
            for info in ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]:
                   add_info_CPS_one_cell_right(CPSsheet,info,[2],[info],
                       "C2C_DATABASE","CHEMICALCLASS",inv_number)

            # Adding other info
            for o_info in ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]:
                   add_info_CPS_right_until_empty(CPSsheet,o_info,[2],[o_info],
                       "C2C_DATABASE","OTHERINFO",inv_number, include_resource=False)

            # CARCINOGENICITY
            for carc_type in ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK","Carcinogenicity Classified IARC",
                "Carcinogenicity Classified TLV", "Carcinogenicity experimental evidence", "Carcinogenicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,carc_type,[1],[carc_type],
                    "C2C_DATABASE","CARCINOGENICITY",inv_number)

            # ED
            for ED_type in ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,ED_type,[1],[ED_type],
                    "C2C_DATABASE","ENDOCRINE",inv_number)

            # MUTAGENICITY/GENOTOXICITY
            # General information
            for muta_type in ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,muta_type,[1],[muta_type],
                    "C2C_DATABASE","MUTAGENICITY",inv_number)
            # Point mutations
            P_mut = loop_over_to_collect_right_values(CPSsheet,"Point mutations:")  # make a string with all point mut
            #print("Point mut",P_mut) #print to see if it makes a string with the point mut
            for p_mut in P_mut:
                 add_info_right_two_markers_OECD(CPSsheet,"Point mutations:",p_mut, "C2C_DATABASE","POINTMUT",inv_number)


            # Chromosomal mutations
            Ch_mut = loop_over_to_collect_right_values(CPSsheet,
                                                      "Chromosome damaging:")  # make a string with all point mut
            #print("Ch mut",Ch_mut)  # print to see if it makes a string with the point mut
            for ch_mut in Ch_mut:
                add_info_right_two_markers_OECD(CPSsheet,"Chromosome damaging:",ch_mut,"C2C_DATABASE","CHROMDAM",inv_number)

            # REPRODUCTIVE TOXICITY
            for repro_type in ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,repro_type,[1],[repro_type],
                    "C2C_DATABASE","REPROTOX",inv_number)

            # DEVELOPMENTAL TOXICITY
            for devo_type in ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,devo_type,[1],[devo_type],
                    "C2C_DATABASE","DEVELOPTOX",inv_number)

            # NEUROTOXICITY
            for neuro_type in ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,neuro_type,[1],[neuro_type],
                    "C2C_DATABASE","NEUROTOX",inv_number)

            # ORAL TOXICITY
            for oral_type in ["Oral toxicity Acute Tox classified","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,oral_type,[1],[oral_type],
                    "C2C_DATABASE","ORALTOX",inv_number)

            # INHALATIVE TOXICITY
            for inhal_type in ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,inhal_type,[1],[inhal_type],
                    "C2C_DATABASE","INHALTOX",inv_number)

            # DERMAL TOXICITY
            for dermal_type in ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,dermal_type,[1],[dermal_type],
                    "C2C_DATABASE","DERMALTOX",inv_number)

            # SKIN/EYE IRRITATION/CORROSION
            for irrit_type in ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]:
                add_info_CPS_right_until_empty(CPSsheet,irrit_type,[1],[irrit_type],
                    "C2C_DATABASE","IRRITCOR",inv_number)

            # SENSITISATION
            for sens_type in ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]:
                add_info_CPS_right_until_empty(CPSsheet,sens_type,[1],[sens_type],
                    "C2C_DATABASE","SENSITISATION",inv_number)

            # ADD Specific concentration limits

            SCL = loop_over_to_collect_right_values(CPSsheet, "Hazard classification:") # make a string with all SCL for each CAS
            #print(SCL) #print to see if it makes a string with the specific conc limits
            for spec_conc_lim in SCL:
                add_info_CPS_from_row_with_two_markers(CPSsheet,"Hazard classification:", spec_conc_lim, "Lower Limit: (%)", "Upper Limit: (%)" , "C2C_DATABASE","SCONCLIM",inv_number)

            #  OTHER CRITERIA
            for other_criteria in ["Other comments"]:
                add_info_CPS_right_until_empty(CPSsheet, other_criteria, [1], [other_criteria],
                                   "C2C_DATABASE", "OCRIT", inv_number)

            # AQUATIC TOXICITY
            for aqtox_type in ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]:
                add_info_CPS_right_until_empty(CPSsheet, aqtox_type, [1], [aqtox_type],
                                    "C2C_DATABASE", "AQUATOX", inv_number)
                # VERTEBRATE FISH
            for fish_type in ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet,fish_type,[1],[fish_type],
                    "C2C_DATABASE","FISHTOX",inv_number)
                # INVERTEBRATE
            for inv_type in ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet, inv_type, [1], [inv_type],
                                   "C2C_DATABASE", "INVTOX", inv_number)
                # ALGAE
            for algae_type in ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]:
                add_info_CPS_right_until_empty(CPSsheet, algae_type, [1], [algae_type],
                                   "C2C_DATABASE", "ALGAETOX", inv_number)

            # TERRESTRIAL TOXICITY
            for tertox_type in ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet, tertox_type, [1], [tertox_type],
                                   "C2C_DATABASE", "TERTOX", inv_number)
            # Other species toxicity
            for spec_tox_type in ["Any other CLP species classification"]:
                add_info_CPS_right_until_empty(CPSsheet, spec_tox_type, [1], [spec_tox_type],
                                   "C2C_DATABASE", "SPECTOX", inv_number)

            # PERSISTENCE
            for pers_type in ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]:
                add_info_CPS_right_until_empty(CPSsheet,pers_type,[1],[pers_type],
                    "C2C_DATABASE","PERSISTENCE",inv_number)

            # BIOACCUMULATION
            for bioac_type in ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]:
                add_info_CPS_right_until_empty(CPSsheet,bioac_type,[1],[bioac_type],
                    "C2C_DATABASE","BIOACCUMULATION",inv_number)

            # CLIMATIC RELEVANCE
            for clima_type in ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]:
                add_info_CPS_right_until_empty(CPSsheet,clima_type,[1],[clima_type],
                    "C2C_DATABASE","CLIMATICRELEVANCE",inv_number)

            #  ADDITIONAL SOURCES
            for add_sources in ["Additional sources"]:
                add_info_CPS_right_until_empty(CPSsheet, add_sources, [1], [add_sources],
                                   "C2C_DATABASE", "ADDSOURCE", inv_number)

        connection.commit()
        st.write("SQL updated")
    except sqlite3.Error as e:
        st.write("SQLite error", e, inv_number)

    finally:
        if connection:
            connection.commit()
            connection.close()
        st.write("Connection closed.")



#### Create/update C2C database with CAS numbers from Excel files ####
#if connceted to database you can press run
if os.path.isfile(db_path):
    if st.button(":green[Run]"):
        ### Beginning: before starting download all available json files and make a backup of the DB
        # json files download (first step to get new CnL info):
        #CnL_json = check_json(CASall, API_key, save_json_dirr)
        #st.success(":green[CnL data extracted from the website.]")
        make_a_backup(db_path)
        st.success(":green[Backup complete.]")

        ### if the CAS exists and make a list with existing cas and the ones that have to be created
        # found => CAS exists in DB
        # not_found => CAS not in DB
        found, not_found = checking_if_CAS_exists(CASall, db_path)
        st.write(f"CAS found in database: {', '.join(found)}")
        st.write(f"CAS not in database: {', '.join(not_found)}")

        ### for CAS numbers that are NOT in the database (not_found)
        # check if there is an Excel file with the given CAS
        CAS_in_folder_1, CAS_not_in_folder_1 = check_if_excel_is_in_folder(folder_excels, not_found)
        st.write(f"CAS found as an Excel: {', '.join(CAS_in_folder_1)}")
        st.write(f"CAS not found as an Excel: {', '.join(CAS_not_in_folder_1)}")



        ### for CAS numbers that are in the database (found)
        # check if there is an Excel file with the given CAS
        CAS_in_folder, CAS_not_in_folder = check_if_excel_is_in_folder(folder_excels, found)
        st.write(f"CAS found as an Excel: {', '.join(CAS_in_folder)}")
        if CAS_not_in_folder != []:
            st.write(f":red[Those CAS are in DB but not as Excel files: {', '.join(CAS_not_in_folder)}]")

        #list = insert_json_info_to_DB(CnL_json, db_path, CASall)
        #st.write(list)
        ### Recent data?
        excels_needing_an_update = is_DB_data_up_to_date_with_excel(db_path, folder_excels, CAS_list=["10-00-0", "1592-23-0"])
        st.write(f"Files need updating: {', '.join(excels_needing_an_update)}")

        extract_info_form_excel_to_DB(db_path, folder_excels, CAS_needing_DB_update=["10-00-0", "1592-23-0"])







