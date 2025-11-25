### This script goes to the CPS directory, checks all CAS files present and adds the info to the SQLite database.

#### SET UP ####
import sqlite3
import os
import re
import pandas as pd
from datetime import date
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from datetime import date
from copy import copy
from openpyxl import load_workbook, Workbook

# Define the path to your SQLite database file
C2Cpath = "/Users/juliakulpa/Desktop/Test_excel_imports"
C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# C2Cfiles_path = os.path.join(C2Cpath)
db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")
# /Users/juliakulpa/Desktop/Test_excel_imports/Database/C2Cdatabase.db
print(db_path)
# # LOAD EXCEL CPS TEMPLATE
template_path = "/Users/juliakulpa/Desktop/Test_excel_imports/Template/CPS_CAS TEMPLATE V2.xlsm"
template_wb = load_workbook(template_path, read_only=False, keep_vba=True)
ws_template = template_wb["C2Coverview"]

database = '/Users/juliakulpa/Desktop/Test_excel_imports/Database /C2Cdatabase.db'
CAS = "10-00-0"
folder = '/Users/juliakulpa/Desktop/Test_excel_imports/Testing/'


### CUSTOM FUNCTIONS ###

# def db_to_excel_one_below(table_name, column_to_get,lookup_column,lookup_value,label_excel):
#     # Query the database
#     query = f"SELECT [{column_to_get}] FROM {table_name} WHERE {lookup_column} = ?"
#     cursor.execute(query, (lookup_value,))
#     result = cursor.fetchone()
#     if not result:
#         print(f"No result found for {lookup_column} = {lookup_value}")
#         return
#     value_to_insert = result[0]
#     # Find the label in the worksheet and insert the value below it
#     for row in ws_template.iter_rows():
#         for cell in row:
#             if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#                 ws_template.cell(row=cell.row + 1, column=cell.column).value = value_to_insert
#                 print(f"Inserted '{value_to_insert}' below '{label_excel}' in cell {cell.coordinate}")
#                 return
#     print(f"Label '{label_excel}' not found in worksheet.")
#
# def db_to_excel_x_right(table_name, columns_to_get, lookup_column, lookup_value, labels_excel,offset):
#     # Ensure both inputs are lists
#     if isinstance(columns_to_get, str):
#         columns_to_get = [columns_to_get]
#     if isinstance(labels_excel, str):
#         labels_excel = [labels_excel]
#
#     # Query the database for multiple columns
#     query = f"SELECT {', '.join(f'[{col}]' for col in columns_to_get)} FROM {table_name} WHERE {lookup_column} = ?"
#     cursor.execute(query, (lookup_value,))
#     result = cursor.fetchone()
#     if not result:
#         print(f"No result found for {lookup_column} = {lookup_value}")
#         return
#     # Map column -> value from database
#     col_value_map = dict(zip(columns_to_get, result))
#
#     # Loop over each label and insert corresponding column value
#     for col_name, label_excel in zip(columns_to_get, labels_excel):
#         value_to_insert = col_value_map[col_name]
#         inserted = False
#         for row in ws_template.iter_rows():
#             for cell in row:
#                 if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
#                     # Add the value in the column x (= offset) columns to the right
#                     ws_template.cell(row=cell.row , column=cell.column + offset).value = value_to_insert
#                     print(f"Inserted '{value_to_insert}' to the right of '{label_excel}' in cell {cell.coordinate}")
#                     inserted = True
#                     break
#             if inserted:
#                 break
#         if not inserted:
#             print(f"Label '{label_excel}' not found in worksheet.")
#


### Function

def extraction_info_excels(database, CAS, ws_template, folder):
    def db_to_excel_multiple_below(maindb, main_ref, linked_db, link_ref, column_to_get, lookup_column, lookup_value,
                                   label_excel):

        # Query the database for all matching values
        try:
            query = f"""
                 SELECT a.[{column_to_get}]
                 FROM {linked_db} a
                 JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                 WHERE c.{lookup_column} = ?
             """
            cursor.execute(query, (lookup_value,))
            results = cursor.fetchall()
            if not results:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            # Find the label in the worksheet
            for row in ws_template.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                        start_row = cell.row + 1
                        col = cell.column
                        # print(f"First test on{start_row, col}")

                        # Place each value in the first empty cell below the starting row
                        for result in results:
                            # If result is not none:

                            if result[0] != None:
                                # Start searching from start_row downward
                                target_row = start_row

                                # Keep moving down until we find an empty cell in the target column
                                while ws_template.cell(row=target_row, column=col).value not in (None, ''):
                                    target_row += 1

                                # print(f"target row{target_row}")

                                # Write the value in the first empty cell found
                                ws_template.cell(row=target_row, column=col).value = result[0]

                                print(
                                    f"Inserted '{result[0]}' into cell {ws_template.cell(row=target_row, column=col).coordinate}")

                        return

            print(f"Label '{label_excel}' not found in worksheet.")
        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_excel_source_right(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,    # base name, e.g. "source"
        lookup_column,
        lookup_value,
        label_excel,
        offset,
        max_suffix=5,    # how far to look for -1, -2 for the additional data
        include_resource=True # does it look for resources
    ):
        # helper: sanitize label like in add_info_CPS_right_until_empty
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]  # row[1] is column name

            # 2) Collect base + suffix columns for the main data (source, source-1, source-2, ...)
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                candidate = f"{column_to_get}-{i}"
                if candidate in all_cols:
                    matching_cols.append(candidate)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column name in SQL: resource-<sanitized label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(label_excel)
                candidate_resource_col = f"resource-{safe_label}"
                if candidate_resource_col in all_cols:
                    resource_sql_col = candidate_resource_col
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")

            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            # Split row into value columns and (optional) resource
            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # 5) Find the label in the worksheet
            for excel_row in ws_template.iter_rows():
                for cell in excel_row:
                    if cell.value and isinstance(cell.value, str) and label_excel.lower() in cell.value.lower():
                        start_row = cell.row
                        start_col = cell.column + offset  # first column to write values

                        # 6) Write the main values horizontally to the right
                        current_col = start_col
                        for val in value_cols:
                            if val is not None and val != "":
                                ws_template.cell(row=start_row, column=current_col).value = val
                                print(
                                    f"Inserted '{val}' into cell "
                                    f"{ws_template.cell(row=start_row, column=current_col).coordinate}"
                                )
                            current_col += 1

                        # 7) If resource is present in SQL, put it in Excel "Resource" column
                        if include_resource and resource_val not in (None, ""):
                            resource_col_idx = None
                            # Find the "Resource" header column in the template
                            for hdr_row in ws_template.iter_rows():
                                for hdr_cell in hdr_row:
                                    if (
                                        isinstance(hdr_cell.value, str)
                                        and hdr_cell.value.strip().lower() == "resource"
                                    ):
                                        resource_col_idx = hdr_cell.column
                                        break
                                if resource_col_idx is not None:
                                    break

                            if resource_col_idx is not None:
                                ws_template.cell(row=start_row, column=resource_col_idx).value = resource_val
                                print(
                                    f"Inserted resource '{resource_val}' into cell "
                                    f"{ws_template.cell(row=start_row, column=resource_col_idx).coordinate}"
                                )
                            else:
                                print("Could not find 'Resource' column in Excel to write resource value.")

                        return  # done after first matching label

            print(f"Label '{label_excel}' not found in worksheet.")

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_column_names_unique(maindb, main_ref,linked_db, link_ref,lookup_column, lookup_value):
        """
        Returns a string with column names unique for each CAS, as a string
        """

        try:
            # Query EVERYTHING (*) from linked_db
            query = f"""
                SELECT a.*
                FROM {linked_db} a
                JOIN {maindb} c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
            """

            cursor.execute(query, (lookup_value,))
            rows = cursor.fetchall()

            # If nothing found → return empty DataFrame (still safe)
            if not rows:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return pd.DataFrame()

            # Extract column names automatically from cursor.description
            colnames = [desc[0] for desc in cursor.description]

            dataframe = pd.DataFrame(rows, columns=colnames)

            # cutting columns with NULL values
            dataframe_cut = dataframe.dropna(axis=1, how='all')
            # dropping columns with ID and ref (not needed here)
            dataframe_cut = dataframe_cut.drop(columns=["ID", 'ref'])
            column_name = list(dataframe_cut.columns)
            # takes away from the string the resources names
            result = [c_name for c_name in column_name if "resource" not in c_name.lower() ]
            return result

        except sqlite3.Error as e:
            print("SQLite error:", e)
            return pd.DataFrame()

    def remove_text_from_string(string, target_name):
        '''removes text from string, used for Muta tests and SCL'''
        result = []
        for s in string:
            name = s.replace(target_name, "").strip()
            result.append(name)
        return result

    def write_list_right_of_label(ws_template: Worksheet, label_excel: str, offset: int, values_list: list):
        """
        Find the first cell whose value EXACTLY matches 'label_excel'
        (after stripping whitespace, case-insensitive),
        then write values from values_list to the right, moving downwards
        as long as the label cell below also exactly matches the label.
        """

        if not values_list:
            print("Value list is empty — nothing to write.")
            return

        # Normalize the target label once
        normalized_label = label_excel.strip().lower()

        # 1) Find the first exact match
        label_cell = None
        for row in ws_template.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == normalized_label:
                    label_cell = cell
                    break
            if label_cell:
                break

        if not label_cell:
            print(f"Exact label '{label_excel}' not found.")
            return

        start_row = label_cell.row
        label_col = label_cell.column
        target_col = label_col + offset
        max_row = ws_template.max_row

        current_row = start_row
        value_index = 0

        # 2) Write downward while exact match continues
        while value_index < len(values_list) and current_row <= max_row:
            current_cell_value = ws_template.cell(row=current_row, column=label_col).value

            # Check if current row still has EXACT match
            if not isinstance(current_cell_value, str) or \
               current_cell_value.strip().lower() != normalized_label:
                break

            ws_template.cell(row=current_row, column=target_col).value = values_list[value_index]

            print(
                f"Inserted '{values_list[value_index]}' into "
                f"{ws_template.cell(row=current_row, column=target_col).coordinate}"
            )

            value_index += 1
            current_row += 1

        if value_index < len(values_list):
            print(f"Warning: {len(values_list) - value_index} values not written — no more exact matching label rows.")

    def refdb_to_excel_source_after_two_targets(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,      # base name, e.g. "source"
        lookup_column,
        lookup_value,
        first_label_excel,  # label of first target cell
        second_label_excel, # label of second target cell (same row as first)
        max_suffix=5,       # how far to look for -1, -2, ... columns
        include_resource=True
    ):
        """
        Finds a row where both first_label_excel and second_label_excel are exact matches.
        Uses the rightmost of those as the "second target" and writes values starting at
        the first unmerged empty cell to its right.
        SQL resource column is: resource-<sanitized second_label_excel>.
        In Excel, resource always goes to the 'Resource' column.
        """
        #pre-step sanitizing labels
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]

            # 2) Collect base + suffix columns for the main data
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                cand = f"{column_to_get}-{i}"
                if cand in all_cols:
                    matching_cols.append(cand)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column in SQL: resource-<sanitized second_label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(second_label_excel)
                cand_res = f"resource-{safe_label}"
                if cand_res in all_cols:
                    resource_sql_col = cand_res
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")
            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # 5) Find a row where BOTH labels exist as exact cell matches
            row_found = False
            for excel_row in ws_template.iter_rows():
                first_col_idx = None
                second_col_idx = None

                for cell in excel_row:
                    if isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == str(first_label_excel).strip():
                            first_col_idx = cell.column
                        elif cell_val == str(second_label_excel).strip():
                            second_col_idx = cell.column

                if first_col_idx is not None and second_col_idx is not None:
                    row_found = True
                    target_row = excel_row[0].row
                    second_target_col = max(first_col_idx, second_col_idx)

                    # 6) Find the first *unmerged* empty cell to the right
                    current_col = second_target_col + 1
                    while True:
                        cell = ws_template.cell(row=target_row, column=current_col)

                        # Skip merged cells (non top-left)
                        if isinstance(cell, MergedCell):
                            current_col += 1
                            continue

                        if cell.value in (None, ""):
                            break

                        current_col += 1

                    # 7) Write values, skipping merged/occupied cells
                    for val in value_cols:
                        if val in (None, ""):
                            continue

                        while True:
                            cell = ws_template.cell(row=target_row, column=current_col)

                            if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                                current_col += 1
                                continue

                            cell.value = val
                            print(
                                f"Inserted '{val}' into cell {cell.coordinate}"
                            )
                            current_col += 1
                            break

                    # 8) Put resource into Excel "Resource" column (header cell == 'Resource')
                    if include_resource and resource_val not in (None, ""):
                        resource_col_idx = None
                        for hdr_row in ws_template.iter_rows():
                            for hdr_cell in hdr_row:
                                if (
                                    isinstance(hdr_cell.value, str)
                                    and hdr_cell.value.strip().lower() == "resource"
                                ):
                                    resource_col_idx = hdr_cell.column
                                    break
                            if resource_col_idx is not None:
                                break

                        if resource_col_idx is not None:
                            ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                            print(
                                f"Inserted resource '{resource_val}' into cell "
                                f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                            )
                        else:
                            print("Could not find 'Resource' column in Excel to write resource value.")
                    break

            if not row_found:
                print(
                    f"No row found where both '{first_label_excel}' "
                    f"and '{second_label_excel}' are present as exact matches."
                )

        except sqlite3.Error as e:
            print("SQLite error:", e)

    def refdb_to_excel_source_after_two_targets_OECD(
        maindb,
        main_ref,
        linked_db,
        link_ref,
        column_to_get,      # base name, e.g. "source"
        lookup_column,
        lookup_value,
        first_label_excel,  # label of first target cell
        second_label_excel, # label of second target cell (same row as first)
        max_suffix=5,       # how far to look for -1, -2, ... columns
        include_resource=True
    ):
        """
        Finds a row where both first_label_excel and second_label_excel are exact matches.
        Uses the rightmost of those as the "second target" and writes values starting at
        the first unmerged empty cell to its right.

        To work with the function getting info from SQL database, the No data could not be left as an empty cell as was before
        so it is written as "no data", however then the program to get the same looking template just has to skip it, not to
        write "do data" twice
        If second_label_excel == "No data" (case-insensitive):
          - Skip writing the SQL values for column_to_get and its suffixes.
          - Still write the SQL resource value (if available) into the Excel "Resource" column.
        """
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        try:
            # 1) Get list of all columns in linked_db
            pragma_sql = f"PRAGMA table_info([{linked_db}])"
            cursor.execute(pragma_sql)
            cols_info = cursor.fetchall()
            all_cols = [row[1] for row in cols_info]

            # 2) Collect base + suffix columns for the main data
            matching_cols = []
            if column_to_get in all_cols:
                matching_cols.append(column_to_get)

            for i in range(1, max_suffix + 1):
                cand = f"{column_to_get}-{i}"
                if cand in all_cols:
                    matching_cols.append(cand)

            if not matching_cols:
                print(f"No columns found for base '{column_to_get}' in table '{linked_db}'")
                return

            # 3) Resource column in SQL: resource-<sanitized second_label_excel>
            resource_sql_col = None
            select_resource = False
            if include_resource:
                safe_label = sanitize_label(second_label_excel)
                cand_res = f"resource-{safe_label}"
                if cand_res in all_cols:
                    resource_sql_col = cand_res
                    select_resource = True

            # 4) Build SELECT list
            select_parts = [f"a.[{col}]" for col in matching_cols]
            if select_resource:
                select_parts.append(f"a.[{resource_sql_col}]")
            select_list = ", ".join(select_parts)

            query = f"""
                SELECT {select_list}
                FROM {linked_db} a
                JOIN {maindb}  c ON a.{link_ref} = c.{main_ref}
                WHERE c.{lookup_column} = ?
                LIMIT 1
            """

            cursor.execute(query, (lookup_value,))
            row = cursor.fetchone()
            if not row:
                print(f"No results found for {lookup_column} = {lookup_value}")
                return

            num_val_cols = len(matching_cols)
            value_cols = row[:num_val_cols]
            resource_val = row[num_val_cols] if select_resource else None

            # Skipping no data: decide whether to skip writing the SQL values
            skip_value_write = str(second_label_excel).strip().lower() == "no data"

            # 5) Find a row where BOTH labels exist as exact cell matches
            row_found = False
            for excel_row in ws_template.iter_rows():
                first_col_idx = None
                second_col_idx = None

                for cell in excel_row:
                    if isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == str(first_label_excel).strip():
                            first_col_idx = cell.column
                        elif cell_val == str(second_label_excel).strip():
                            second_col_idx = cell.column

                if first_col_idx is not None and second_col_idx is not None:
                    row_found = True
                    target_row = excel_row[0].row
                    second_target_col = max(first_col_idx, second_col_idx)

                    # 6) Find the first *unmerged* empty cell to the right
                    current_col = second_target_col + 1
                    while True:
                        cell = ws_template.cell(row=target_row, column=current_col)

                        # Skip merged cells (non top-left)
                        if isinstance(cell, MergedCell):
                            current_col += 1
                            continue

                        if cell.value in (None, ""):
                            break

                        current_col += 1

                    # 7) Write values, skipping merged/occupied cells
                    # Only write values if NOT "No data"
                    if not skip_value_write:
                        for val in value_cols:
                            if val in (None, ""):
                                continue

                            while True:
                                cell = ws_template.cell(row=target_row, column=current_col)

                                if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                                    current_col += 1
                                    continue

                                cell.value = val
                                print(
                                    f"Inserted '{val}' into cell {cell.coordinate}"
                                )
                                current_col += 1
                                break
                    else:
                        print("Second label is 'No data' → skipping SQL values, but still handling resource if present.")

                    # 8) Put resource into Excel "Resource" column (header cell == 'Resource')
                    if include_resource and resource_val not in (None, ""):
                        resource_col_idx = None
                        for hdr_row in ws_template.iter_rows():
                            for hdr_cell in hdr_row:
                                if (
                                    isinstance(hdr_cell.value, str)
                                    and hdr_cell.value.strip().lower() == "resource"
                                ):
                                    resource_col_idx = hdr_cell.column
                                    break
                            if resource_col_idx is not None:
                                break

                        if resource_col_idx is not None:
                            ws_template.cell(row=target_row, column=resource_col_idx).value = resource_val
                            print(
                                f"Inserted resource '{resource_val}' into cell "
                                f"{ws_template.cell(row=target_row, column=resource_col_idx).coordinate}"
                            )
                        else:
                            print("Could not find 'Resource' column in Excel to write resource value.")
                    break

            if not row_found:
                print(
                    f"No row found where both '{first_label_excel}' "
                    f"and '{second_label_excel}' are present as exact matches."
                )

        except sqlite3.Error as e:
            print("SQLite error:", e)


    ### Start with extracting
    try:
        ### SQL SET-UP
        connection = sqlite3.connect(database)
        cursor = connection.cursor()

        print("Connected to SQLite database at:", db_path)

        # GENERAL INFO
        #Add general info
        namesDBcol_gen = ["Chemical name", "Common name", "CAS number", "EC number", "Linked CAS Read across",
                       "Linked CAS Monomers", "Linked CAS Degradation Products"]
        namesExcel_gen = ["Chemical name", "Common name", "CAS number", "EC number", "Linked CAS Read across",
                       "Linked CAS Monomers", "Linked CAS Degradation Products"]
        for namesDBcol_gen, namesExcel_gen in zip(namesDBcol_gen,namesExcel_gen):
            db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="GENERALINFO", link_ref="ref",
                                   column_to_get=namesDBcol_gen, lookup_column="ID",lookup_value =CAS, label_excel=namesExcel_gen)
        # ADD ASSESSORS
        db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="ASSESSORS", link_ref="ref",
                                   column_to_get="Name assessor", lookup_column="ID", lookup_value=CAS,
                                   label_excel="Name assessor")
        db_to_excel_multiple_below(maindb="C2C_DATABASE", main_ref="ID", linked_db="ASSESSORS", link_ref="ref",
                                   column_to_get="Date assessed", lookup_column="ID", lookup_value=CAS,
                                   label_excel="Date created/updated")
        ## Add various info CHEMICAL CLASS
        namesDBcol_CC = ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]
        namesExcel_CC = ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]
        for names_DB, name_EX in zip(namesDBcol_CC, namesExcel_CC):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHEMICALCLASS", link_ref="ref",
                                        column_to_get=names_DB, lookup_column="ID", lookup_value=CAS,
                                        label_excel=name_EX, offset=2)

        # Adding other info
        namesDBcol_OTHER = ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]
        namesExcel_OTHER = ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]
        for names_DB_OTHER, name_EX_OTHER in zip(namesDBcol_OTHER, namesExcel_OTHER):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="OTHERINFO", link_ref="ref",
                                        column_to_get=names_DB_OTHER, lookup_column="ID", lookup_value=CAS,
                                        label_excel=name_EX_OTHER, offset=2)
        #  OTHER CRITERIA
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="OCRIT", link_ref="ref",
                                    column_to_get="Other comments", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Other comments", offset=1)
        # CARCINOGENICITY
        namesDBcols = ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK", "Carcinogenicity Classified IARC",
                        "Carcinogenicity Classified TLV","Carcinogenicity experimental evidence","Carcinogenicity Comments"]
        namesExcel = ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK", "Carcinogenicity Classified IARC",
                        "Carcinogenicity Classified TLV","Carcinogenicity experimental evidence","Carcinogenicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcols,namesExcel):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CARCINOGENICITY", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1,max_suffix=5,include_resource=True)

        # ED
        namesDBcols = ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]
        namesExcel = ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]
        for namesDBcol, nameExcel in zip(namesDBcols,namesExcel):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ENDOCRINE", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # MUTAGENICITY
        namesDBcol_MUT = ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]
        namesExcel_MUT = ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_MUT,namesExcel_MUT):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="MUTAGENICITY", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # MUTAGENICITY OECD TESTS
        #Point mutations
        point_mut_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT",
                                                       link_ref="ref",
                                                       lookup_column="ID", lookup_value=CAS)
        point_mut_names_cleared = remove_text_from_string(point_mut_names, "Point mutations:")
        #print(point_mut_names)

        write_list_right_of_label(ws_template, "Point mutations:", 1, point_mut_names_cleared)

        for namesDB, nameExcel in zip(point_mut_names, point_mut_names_cleared):
            refdb_to_excel_source_after_two_targets_OECD(maindb="C2C_DATABASE", main_ref="ID", linked_db="POINTMUT", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel="Point mutations:", second_label_excel=nameExcel, max_suffix=5, include_resource=True )

        # Chromosome damaging
        ch_dam_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM",
                                                    link_ref="ref",
                                                    lookup_column="ID", lookup_value=CAS)
        ch_dam_names_cleared = remove_text_from_string(ch_dam_names, "Chromosome damaging:")
        #print(ch_dam_names)
        write_list_right_of_label(ws_template, "Chromosome damaging:", 1, ch_dam_names_cleared)

        for namesDB, nameExcel in zip(ch_dam_names, ch_dam_names_cleared):
            refdb_to_excel_source_after_two_targets_OECD(maindb="C2C_DATABASE", main_ref="ID", linked_db="CHROMDAM", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel="Chromosome damaging:", second_label_excel=nameExcel, max_suffix=5, include_resource=True )


        # REPROTOX
        namesDBcol_REP = ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]
        namesExcel_REP = ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_REP,namesExcel_REP):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="REPROTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # DEVELOPMENTAL TOX
        namesDBcol_DEV = ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]
        namesExcel_DEV = ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_DEV,namesExcel_DEV):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="DEVELOPTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # NEUROTOX
        namesDBcol_NETOX = ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]
        namesExcel_NETOX = ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_NETOX,namesExcel_NETOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="NEUROTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # ORAL TOX
        namesDBcol_ORTOX = ["Oral toxicity Acute Tox classified","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]
        namesExcel_ORTOX = ["Oral toxicity Acute Tox classified:","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_ORTOX,namesExcel_ORTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ORALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # INHALE TOX

        namesDBcol_INHTOX = ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]
        namesExcel_INHTOX = ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_INHTOX,namesExcel_INHTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="INHALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # DERMAL TOX
        namesDBcol_DERMTOX = ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]
        namesExcel_DERMTOX = ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_DERMTOX,namesExcel_DERMTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="DERMALTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SKIN/EYE IRRIT/COR
        namesDBcol_IRR = ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]
        namesExcel_IRR = ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_IRR,namesExcel_IRR):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="IRRITCOR", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SENSITISATION
        namesDBcol_SENS = ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]
        namesExcel_SENS = ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_SENS,namesExcel_SENS):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="SENSITISATION", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # SPECIFIC CONCENTRATION LIMITS
        SCL_names = refdb_to_column_names_unique(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                 link_ref="ref",
                                                 lookup_column="ID", lookup_value=CAS)
        # cleaning the names so there is only distinct SCL names
        SCL_names_clean = remove_text_from_string(SCL_names, " - Lower Limit: (%)")
        SCL_names_clean = remove_text_from_string(SCL_names_clean, " - Upper Limit: (%)")
        SCL_names_dist = list(dict.fromkeys(SCL_names_clean))
        #print(SCL_names_dist)
        write_list_right_of_label(ws_template, "Hazard classification:", 1, SCL_names_dist)
        # # Lower limit
        # choosing SQL with only Lower limit data
        SCL_DB_names_lower = [s for s in SCL_names if "Lower Limit:" in s]
        print(SCL_DB_names_lower)
        SCL_EX_names_lower = remove_text_from_string(SCL_DB_names_lower, " - Lower Limit: (%)")
        print(SCL_EX_names_lower)
        # extracting Lower limit data
        for namesDB, nameExcel in zip(SCL_DB_names_lower, SCL_EX_names_lower):
            refdb_to_excel_source_after_two_targets(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM", link_ref="ref",
                                       column_to_get=namesDB, lookup_column="ID",lookup_value =CAS, first_label_excel=nameExcel, second_label_excel="Lower Limit: (%)", max_suffix=5, include_resource=True )

        # # Upper limit
        # choosing SQL with only Lower limit data
        SCL_DB_names_upper = [s for s in SCL_names if "Upper Limit:" in s]
        SCL_EX_names_upper = remove_text_from_string(SCL_DB_names_upper, " - Upper Limit: (%)")
        # extracting Lower limit data
        for namesDB, nameExcel in zip(SCL_DB_names_upper, SCL_EX_names_upper):
            refdb_to_excel_source_after_two_targets(maindb="C2C_DATABASE", main_ref="ID", linked_db="SCONCLIM",
                                                    link_ref="ref",
                                                    column_to_get=namesDB, lookup_column="ID", lookup_value=CAS,
                                                    first_label_excel=nameExcel, second_label_excel="Upper Limit: (%)",
                                                    max_suffix=5, include_resource=True)

        # AQUATIC TOXICITY

        namesDBcol_AQTOX = ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]
        namesExcel_AQTOX = ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]
        for namesDBcol, nameExcel in zip(namesDBcol_AQTOX,namesExcel_AQTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="AQUATOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)
        # VERTEBRATE FISH
        namesDBcol_FISHTOX = ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]
        namesExcel_FISHTOX = ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_FISHTOX,namesExcel_FISHTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="FISHTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # INVERTEBRATE TOX
        namesDBcol_INVTOX = ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]
        namesExcel_INVTOX = ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_INVTOX,namesExcel_INVTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="INVTOX", link_ref="ref",
                                   column_to_get=namesDBcol, lookup_column="ID",lookup_value =CAS, label_excel=nameExcel,offset=1)

        # ALGAE TOX
        namesDBcol_ALGTOX = ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]
        namesExcel_ALGTOX = ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]
        for namesDBcol, nameExcel in zip(namesDBcol_ALGTOX, namesExcel_ALGTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ALGAETOX", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)

        # TERRESTRIAL TOX
        namesDBcol_TERTOX =  ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]
        namesExcel_TERTOX =  ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_TERTOX, namesExcel_TERTOX):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="TERTOX", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)

        # OTHER SPECIES TOX
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="SPECTOX", link_ref="ref",
                                    column_to_get="Any other CLP species classification", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Any other CLP species classification", offset=1)
        # PERSISTENCE
        namesDBcol_PERS =  ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]
        namesExcel_PERS =  ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_PERS, namesExcel_PERS):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="PERSISTENCE", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        # BIOACCUMULATION
        namesDBcol_BIOAC =  ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]
        namesExcel_BIOAC =  ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_BIOAC, namesExcel_BIOAC):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="BIOACCUMULATION", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        # CLIMATIC RELEVANCE
        namesDBcol_CLIMREL =  ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]
        namesExcel_CLIMREL =  ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]
        for namesDBcol, nameExcel in zip(namesDBcol_CLIMREL, namesExcel_CLIMREL):
            refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="CLIMATICRELEVANCE", link_ref="ref",
                                        column_to_get=namesDBcol, lookup_column="ID", lookup_value=CAS,
                                        label_excel=nameExcel, offset=1)
        #  ADDITIONAL SOURCES
        refdb_to_excel_source_right(maindb="C2C_DATABASE", main_ref="ID", linked_db="ADDSOURCE", link_ref="ref",
                                    column_to_get="Additional sources", lookup_column="ID", lookup_value=CAS,
                                    label_excel="Additional sources", offset=1)
        #### SAVE THE FILLED IN CPS EXCEL ####
        name = f"Test CPS_CAS {CAS}.xlsm"
        saving_path = os.path.join(folder,name)
        template_wb.save(saving_path)

        # def copy_sheet_with_formatting(src_ws, dest_ws):
        #     """Copy values + formatting from src_ws to dest_ws."""
        #     # Cells: values + styles
        #     for row in src_ws.iter_rows():
        #         for cell in row:
        #             new_cell = dest_ws[cell.coordinate]
        #             new_cell.value = cell.value
        #
        #             if cell.has_style:
        #                 new_cell.font = copy(cell.font)
        #                 new_cell.border = copy(cell.border)
        #                 new_cell.fill = copy(cell.fill)
        #                 new_cell.number_format = cell.number_format
        #                 new_cell.protection = copy(cell.protection)
        #                 new_cell.alignment = copy(cell.alignment)
        #
        #     # Column widths
        #     for col_letter, col_dim in src_ws.column_dimensions.items():
        #         dest_ws.column_dimensions[col_letter].width = col_dim.width
        #
        #     # Row heights
        #     for row_idx, row_dim in src_ws.row_dimensions.items():
        #         dest_ws.row_dimensions[row_idx].height = row_dim.height
        #
        # # def copy_sheet_with_formatting(src_ws, dest_ws):
        # #     """Copy values, styles, merged cells, CF and data validation from src_ws to dest_ws."""
        # #
        # #     # --- 1) Copy cell values + styles ---
        # #     for row in src_ws.iter_rows():
        # #         for cell in row:
        # #             new_cell = dest_ws[cell.coordinate]
        # #             new_cell.value = cell.value
        # #
        # #             if cell.has_style:
        # #                 new_cell.font = copy(cell.font)
        # #                 new_cell.border = copy(cell.border)
        # #                 new_cell.fill = copy(cell.fill)
        # #                 new_cell.number_format = cell.number_format
        # #                 new_cell.protection = copy(cell.protection)
        # #                 new_cell.alignment = copy(cell.alignment)
        # #
        # #     # --- 2) Copy column widths ---
        # #     for col_letter, col_dim in src_ws.column_dimensions.items():
        # #         dest_ws.column_dimensions[col_letter].width = col_dim.width
        # #
        # #     # --- 3) Copy row heights ---
        # #     for row_idx, row_dim in src_ws.row_dimensions.items():
        # #         dest_ws.row_dimensions[row_idx].height = row_dim.height
        # #
        # #     # --- 4) Copy merged cells ---
        # #     # (new sheet should be empty but we clear just in case)
        # #     if dest_ws.merged_cells.ranges:
        # #         for r in list(dest_ws.merged_cells.ranges):
        # #             dest_ws.unmerge_cells(str(r))
        # #
        # #     for merge_range in src_ws.merged_cells.ranges:
        # #         dest_ws.merge_cells(str(merge_range))
        # #
        # #     # --- 5) Copy conditional formatting ---
        # #     # _cf_rules is an OrderedDict: {ConditionalFormattingObj -> [Rule, Rule, ...]}
        # #     cf = src_ws.conditional_formatting
        # #     for cf_obj, rules in getattr(cf, "_cf_rules", {}).items():
        # #         for rule in rules:
        # #             dest_ws.conditional_formatting.add(cf_obj.sqref, copy(rule))
        # #
        # #     # --- 6) Copy data validation (the bit that crashed before) ---
        # #     # We must copy *each* DataValidation, not the DataValidationList itself
        # #     dv_list = src_ws.data_validations
        # #     if dv_list is not None and getattr(dv_list, "dataValidation", None):
        # #         for dv in dv_list.dataValidation:
        # #             dv_copy = copy(dv)
        # #             dest_ws.add_data_validation(dv_copy)
        #
        # def save_template_as_first_sheet(template_wb, template_sheet_name, output_folder, base_filename):
        #     """
        #     - template_wb: workbook with your filled template sheet
        #     - template_sheet_name: name of the sheet you filled (e.g. "C2Coverview")
        #     - output_folder: folder where the file should be saved
        #     - base_filename: filename *without* extension (e.g. "MyExport")
        #
        #     Behaviour:
        #       - If <base>.xlsm exists → use it (keep_vba=True), add new first sheet, save as <base>.xlsm
        #       - Else if <base>.xlsx exists → use it, add new first sheet, save as <base>.xlsm
        #       - Else → create new workbook, add first sheet, save as <base>.xlsm
        #
        #     Result: final file is always .xlsm
        #     """
        #     src_ws = template_wb[template_sheet_name]
        #
        #     today_str = date.today().strftime("%Y-%m-%d")
        #     base_sheet_name = f"{template_sheet_name}_{today_str}"
        #
        #     os.makedirs(output_folder, exist_ok=True)
        #
        #     xlsm_path = os.path.join(output_folder, f"{base_filename}.xlsm")
        #     xlsx_path = os.path.join(output_folder, f"{base_filename}.xlsx")
        #
        #     # Decide which workbook to load / create
        #     if os.path.exists(xlsm_path):
        #         # Prefer existing .xlsm and keep macros if there are any
        #         wb = load_workbook(xlsm_path, keep_vba=True)
        #         full_path = xlsm_path
        #     elif os.path.exists(xlsx_path):
        #         # Use existing .xlsx as base, but we will save to .xlsm
        #         wb = load_workbook(xlsx_path)
        #         full_path = xlsm_path
        #     else:
        #         # No file exists → create new workbook
        #         wb = Workbook()
        #         # remove default empty sheet
        #         wb.remove(wb.active)
        #         full_path = xlsm_path
        #
        #     # Ensure sheet name is unique
        #     new_sheet_name = base_sheet_name
        #     i = 1
        #     while new_sheet_name in wb.sheetnames:
        #         new_sheet_name = f"{base_sheet_name}_{i}"
        #         i += 1
        #
        #     # Create as FIRST sheet
        #     new_ws = wb.create_sheet(title=new_sheet_name, index=0)
        #
        #     # Copy template (values + formatting)
        #     copy_sheet_with_formatting(src_ws, new_ws)
        #
        #     # Save as .xlsm
        #     wb.save(full_path)
        #
        # template_sheet_name = "C2Coverview"
        # EX_name = f"CPS_CAS {CAS}"
        # save_template_as_first_sheet(template_wb, template_sheet_name, folder, EX_name)

    except sqlite3.Error as e:
        print("SQLite error:", e)



extraction_info_excels(database, CAS, ws_template, folder)