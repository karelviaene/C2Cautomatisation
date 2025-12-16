### This script goes to the CPS directory, checks all CAS files present and adds the info to the SQLite database.
import requests
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import os
import json
import sqlite3
import time
import re
import time
import os
import re
from datetime import datetime
def checking_if_CAS_exists(CASall, db_path):
    found = []
    not_found = []

    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()

        for cas in CASall:
            cursor.execute("SELECT 1 FROM C2C_DATABASE WHERE ID = ?", (cas,))
            row = cursor.fetchone()

            if row:
                found.append(cas)
            else:
                not_found.append(cas)

        print(f"CAS found in database: {found}")
        print(f"CAS not in database: {not_found}")

    except sqlite3.Error as e:
        print("SQLite error:", e)

    finally:
        if 'connection' in locals():
            connection.close()

    return found, not_found

db_path = '/Users/juliakulpa/Desktop/test/Database/C2Cdatabase.db'
CASall = ['50-00-0', '110-54-3', "50-32-8", "71-43-2", '64-17-5']
CnL_json = '/Users/juliakulpa/Desktop/test/JSON/output/CnLscreener exportJSON 2025-12-16 10-49.json'
folder_excels = "/Users/juliakulpa/Desktop/test/CPS"
# found, not_found = checking_if_CAS_exists(CASall,db_path)

CnL_json = '/Users/juliakulpa/Desktop/test/JSON/output/CnLscreener exportJSON 2025-12-16 11-09.json'

#### Upload info to CnL database ####

# Define the path to your SQLite database file
# db_path = os.path.join("/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/Database/C2Cdatabase.db")
# # C2Cfiles_path = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling"
# C2Cpath = "/Users/juliakulpa/Desktop/Test_echa/Testing"
# db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")
# C2Cfiles_path = os.path.join(C2Cpath,"CPS")
C2Cpath = "/Users/juliakulpa/Desktop/test"
C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# C2Cfiles_path = os.path.join(C2Cpath)

import sqlite3
import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
import traceback
import zipfile

# Define the path to your SQLite database file
C2Cpath = "/Users/juliakulpa/Desktop/test"
C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# C2Cfiles_path = os.path.join(C2Cpath)
db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")

# Format today’s date
today = datetime.today().strftime("%Y%m%d")  # e.g., 20250903
# Specify whether you want to load-in the CPS files or whether you want to start from a preloaded database
READ_IN_CPS = True


# Format today’s date
today = datetime.today().strftime("%Y%m%d")  # e.g., 20250903


db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")
folder_excels = "/Users/juliakulpa/Desktop/test/CPS"

READ_IN_CPS = True

CAS_needing_DB_update = ["110-54-3","50-00-0"]

def extract_info_form_excel_to_DB(db_path, folder_excels, CAS_needing_DB_update):
    #### CUSTOM FUNCTIONS ####
    def add_info_CPS_below(sheet, search_strings, maindatabase, newdatabase, mainID):
        """
        Finds each search string somewhere in the sheet (not necessarily in the first row),
        then reads the values directly below it column-wise.
        Collects one "row" of data across all search strings,
        and inserts into the SQLite database.

        Stops when *all* searched columns are empty in the same row.

        search_strings:
            - list: Excel labels are used as SQL column names.
            - dict: {Excel label: SQL column name}
        """

        # Handle both list and dict
        if isinstance(search_strings, str):
            search_strings = [search_strings]
        if isinstance(search_strings, list):
            mapping = {s: s for s in search_strings}  # Excel label -> same SQL name
        elif isinstance(search_strings, dict):
            mapping = search_strings  # Excel label -> custom SQL name
        else:
            raise TypeError("search_strings must be a str, list, or dict")

        # Locate the target cells (positions of the labels in Excel)
        col_positions = {}
        for excel_label in mapping.keys():
            found = None
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and excel_label.lower() in str(cell.value).lower():
                        found = cell
                        break
                if found:
                    break
            if found:
                col_positions[excel_label] = (found.column, found.row + 1)  # start below label

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'

        # Check if the table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                       (newdatabase,))
        table_exists = cursor.fetchone()

        if not table_exists:
            # Create table with auto-increment ID and extracted columns
            cols_def = ", ".join([f"{q(sql_col)} TEXT" for sql_col in mapping.values()])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            # Add missing columns if needed
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            for sql_col in mapping.values():
                if sql_col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(sql_col)} TEXT")

        # Iterate row by row until all searched columns are empty
        row_offset = 0
        while True:
            row_data = {}
            all_empty = True
            for excel_label, (col, start_row) in col_positions.items():
                cell = sheet.cell(row=start_row + row_offset, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    sql_col = mapping[excel_label]
                    row_data[sql_col] = cell.value
                    all_empty = False
            if all_empty:
                break  # stop when all searched columns are empty in the same row

            if newdatabase == maindatabase:
                # --- Case 1: Append into the main database table ---
                # Check if the row for this ID exists
                cursor.execute(
                    f"SELECT 1 FROM {q(maindatabase)} WHERE ID = ?",
                    (mainID,)
                )
                exists = cursor.fetchone()

                if exists:
                    # Update only the new columns (append data)
                    update_clause = ", ".join([f"{q(col)} = ?" for col in row_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(maindatabase)} SET {update_clause} WHERE ID = ?",
                        list(row_data.values()) + [mainID]
                    )
                else:
                    # Insert new row with ID and these values
                    all_cols = ["ID"] + list(row_data.keys())
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(maindatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        [mainID] + list(row_data.values())
                    )

            else:
                # --- Case 2: Insert into child table ---
                row_data["ref"] = mainID

                # Build WHERE clause dynamically based on all row_data keys
                where_clause = " AND ".join([f"{q(col)} = ?" for col in row_data.keys()])
                params = list(row_data.values())

                cursor.execute(
                    f"SELECT 1 FROM {q(newdatabase)} WHERE {where_clause}",
                    params
                )
                exists = cursor.fetchone()

                if not exists:
                    # Insert new row
                    all_cols = list(row_data.keys())
                    placeholders = ", ".join(["?"] * len(all_cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                        list(row_data.values())
                    )

            row_offset += 1
    def add_info_CPS_one_cell_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
        """
        Finds rows containing `rowlabel`, extracts specified columns to the right,
        and inserts or updates the data in the SQLite database.

        Parameters:
            sheet: openpyxl worksheet
            rowlabel: string to search for in any row
            column_offsets: list of integers (e.g., [2, 3]) for columns to the right
            column_names: list of strings for custom SQL column names
            maindatabase: name of the main database (for foreign key reference)
            newdatabase: name of the table to update
            mainID: unique identifier for the row
        """

        if len(column_offsets) != len(column_names):
            raise ValueError("column_offsets and column_names must have the same length")

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'


        # Step 1: Find the row containing the rowlabel
        extracted_data = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                    row_idx = cell.row
                    col_idx = cell.column
                    for offset, col_name in zip(column_offsets, column_names):
                        target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                        extracted_data[col_name] = target_cell.value
                    break
            if extracted_data:
                break

        if not extracted_data:
            return  # nothing to insert

        # Step 2: Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        if not table_exists:
            # Create table with ID, ref, and extracted columns
            cols_def = ", ".join([f"{q(col)} TEXT" for col in column_names])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            # Add missing columns
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in column_names:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # Step 3: Insert or update
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                    [extracted_data[col] for col in column_names] + [mainID]
                )
            else:
                all_cols = ['ref'] + column_names
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in column_names]
                )

        else:  # when newdatabase == maindatabase
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                    [extracted_data[col] for col in column_names] + [mainID]
                )
            else:
                all_cols = ['ID'] + column_names
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in column_names]
                )
    def add_info_CPS_right_until_empty(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID, include_resource=True):
        """
        Like add_info_CPS_one_cell_right, but starting from column_offsets[0] to the right,
        keeps reading consecutive cells until it finds the first empty cell.
        Column naming:
          - first value uses column_names[0] (base)
          - next values use column_names[1:], if present
          - beyond that, auto-name as base-1, base-2, ...

        Optional behavior (when True):
          - If include_resource=True, captures the sheet's 'Resource' column value (if present and not empty)
            into SQL column 'resource-<sanitized rowlabel>'.
          - If no 'Resource' column exists or the cell is empty, skips creating that column.
        """

        if len(column_offsets) != len(column_names):
            raise ValueError("column_offsets and column_names must have the same length")
        if not column_offsets:
            return

        # Quote identifiers for SQL safety
        def q(name: str) -> str:
            return f'"{name}"'

        # Sanitize rowlabel for safe SQL column naming
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        safe_rowlabel = sanitize_label(rowlabel)
        resource_colname = f"resource-{safe_rowlabel}"

        # --- locate the cell containing rowlabel ---
        match_row_idx = None
        match_col_idx = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
                    match_row_idx = cell.row
                    match_col_idx = getattr(cell, "col_idx", cell.column)
                    break
            if match_row_idx is not None:
                break

        if match_row_idx is None:
            return  # nothing to insert

        # Optionally find the column index for "Resource"
        resource_col = None
        if include_resource:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == "resource":
                        resource_col = getattr(cell, "col_idx", cell.column)
                        break
                if resource_col:
                    break

        # Determine start offset and base name
        start_offset = column_offsets[0]
        base_name = column_names[0]

        # --- read to the right until the first empty cell ---
        extracted_data = {}
        k = 0
        max_col = sheet.max_column
        while (match_col_idx + start_offset + k) <= max_col:
            target = sheet.cell(row=match_row_idx, column=match_col_idx + start_offset + k)
            tv = target.value
            # stop at first empty/blank
            if tv is None or (isinstance(tv, str) and tv.strip() == ""):
                break

            # choose column name
            if k < len(column_names):
                col_name = column_names[k]
            else:
                col_name = f"{base_name}-{k - (len(column_names) - 1)}" if len(column_names) > 0 else f"col-{k}"

            extracted_data[col_name] = tv
            k += 1

        # Optionally add the Resource value, but only if it's not empty
        if include_resource and resource_col:
            resource_value = sheet.cell(row=match_row_idx, column=resource_col).value
            if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
                extracted_data[resource_colname] = resource_value

        if not extracted_data:
            return  # nothing to insert

        # --- ensure table exists and has needed columns ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- upsert logic (same keying as your working function) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
                    [extracted_data[col] for col in needed_columns] + [mainID]
                )
            else:
                all_cols = ['ref'] + needed_columns
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in needed_columns]
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                update_clause = ", ".join([f"{q(col)} = ?" for col in needed_columns])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
                    [extracted_data[col] for col in needed_columns] + [mainID]
                )
            else:
                all_cols = ['ID'] + needed_columns
                placeholders = ", ".join(["?"] * len(all_cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
                    [mainID] + [extracted_data[col] for col in needed_columns]
                )
    def loop_over_to_collect_right_values(sheet, rowlabel: str):
        """
        Finds the first cell whose value contains `rowlabel`, then collects the value
        in the cell to the right for that row and all consecutive rows below
        where the same label also appears in the same column.
        Returns a single string joining all collected right-hand values separated with commas.
        """

        # Case-insensitive: looking for upper or lower case
        def matches(value):
            if value is None:
                return False
            return str(value).strip().lower() == rowlabel.strip().lower()

        # Step 1: Find the first occurrence of the label
        start_row = None
        start_col = None
        for row in sheet.iter_rows():
            for cell in row:
                if matches(cell.value):
                    start_row = cell.row
                    start_col = getattr(cell, "col_idx", cell.column)
                    break
            if start_row is not None:
                break

        if start_row is None:
            return "No match with row label"

        # Step 2: Walk down while the label repeats and collect right-hand values
        collected = []
        r = start_row
        while r <= sheet.max_row:
            left_val = sheet.cell(row=r, column=start_col).value
            if not matches(left_val):
                break

            right_val = sheet.cell(row=r, column=start_col + 1).value
            if right_val is not None and str(right_val).strip() != "":
                collected.append(str(right_val).strip())

            r += 1

        # Step 2: Return as a single string
        return collected
    def add_info_CPS_from_row_with_two_markers(sheet, label1: str, label2: str, label3: str, label4: str, maindatabase, newdatabase, mainID):
        """
        label 1 - first row to match (e.g. Hazard classification)
        label 2 - second row to match (e.g. Eye Irrit. 2)
        label 3 & 4 looking in the row for them and taking values to the right
        1) Find a row where two adjacent cells match (label1, label2) left→right.
        2) In that row, find label3 and label4 (anywhere in the row), and for each:
           - take the value from the cell immediately to the right.
        3) Write to SQL columns named:
             - "{label3}-{label2}"  (for value next to label3)
             - "{label4}-{label2}"  (for value next to label4)
        Matching is case-insensitive 'contains'.
        Requires a DB cursor in outer scope named `cursor`.
        """

        # --- helpers ---
        def q(name: str) -> str:
            return f'"{name}"'

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        max_row = sheet.max_row
        max_col = sheet.max_column

        # --- 1) Find target row via adjacent (label1, label2) ---
        target_row = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = sheet.cell(row=r, column=c).value
                v2 = sheet.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    break
            if target_row is not None:
                break

        if target_row is None:
            return  # no matching row → nothing to insert

        # --- 2) Scan the row to find label3 and label4 targets; capture right-hand values ---
        extracted_data = {}

        # label3 → col name "{label3}-{label2}"
        col_name_3 = f"{label2} - {label3}"
        # label4 → col name "{label2}-{label4}"
        col_name_4 = f"{label2} - {label4}"

        # We search the entire row for each label, reading the cell to the right when found
        def capture_right_of_label(row: int, label: str):
            for c in range(1, max_col):  # up to max_col-1 to read c+1
                if matches(sheet.cell(row=row, column=c).value, label):
                    right_val = sheet.cell(row=row, column=c + 1).value
                    if right_val is None:
                        return None
                    if isinstance(right_val, str):
                        rv = right_val.strip()
                        return rv if rv != "" else None
                    return right_val
            return None

        val3 = capture_right_of_label(target_row, label3)
        val4 = capture_right_of_label(target_row, label4)

        if val3 is not None:
            extracted_data[col_name_3] = val3
        if val4 is not None:
            extracted_data[col_name_4] = val4

        if not extracted_data:
            return  # neither target produced a value → nothing to insert

        # --- 3) Ensure table/columns exist ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- 4) Upsert (same rules as your working pattern) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ref"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ID"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
    def add_info_right_two_markers_OECD(sheet, label1: str, label2: str, maindatabase, newdatabase, mainID, include_resource: bool = True):
        """
        label 1 - first row to match
        label 2 - second row to match
        1) Find a row where two adjacent cells match (label1, label2) left→right.
        2) Capture first non-empty cell to the right of label2.
        3) Write to SQL columns named: {label1}{label2}

        Optional behavior (when include_resource=True):
          - Captures the sheet's 'Resource' column value (if present and not empty)
            into SQL column 'resource-<sanitized label2>' for the same row.
          - If no 'Resource' column exists or the cell is empty, skips creating that column.
        """

        # --- helpers ---
        def q(name: str) -> str:
            return f'"{name}"'

        def matches(val, needle: str) -> bool:
            if val is None:
                return False
            return needle.lower() in str(val).lower()

        # Sanitize label for safe SQL column naming (for resource column)
        def sanitize_label(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace(" ", "-")
            s = re.sub(r"[^a-z0-9_\-]", "", s)
            s = re.sub(r"-{2,}", "-", s)
            return s or "unnamed"

        # --- NEW: if label2 is "no data", skip Excel and just write "no data" to SQL ---
        if label2.strip().lower() == "no data":
            safe_label2 = sanitize_label(label2)
            col_name = f"{label1}{label2}"   # same pattern as normal case
            extracted_data = {col_name: "no data"}

            # Ensure table/columns exist
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
            table_exists = cursor.fetchone()

            needed_columns = list(extracted_data.keys())

            if not table_exists:
                cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
                fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
                cursor.execute(f'''
                    CREATE TABLE {q(newdatabase)} (
                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                        ref TEXT
                        {"," if cols_def else ""} {cols_def}
                        {fk_clause}
                    )
                ''')
            else:
                cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
                existing_cols = [col[1] for col in cursor.fetchall()]
                if "ref" not in existing_cols and newdatabase != maindatabase:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
                for col in needed_columns:
                    if col not in existing_cols:
                        cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

            # Upsert
            if newdatabase != maindatabase:
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
                exists = cursor.fetchone()
                if exists:
                    set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                        list(extracted_data.values()) + [mainID]
                    )
                else:
                    cols = ["ref"] + list(extracted_data.keys())
                    placeholders = ", ".join(["?"] * len(cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                        [mainID] + list(extracted_data.values())
                    )
            else:
                cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
                exists = cursor.fetchone()
                if exists:
                    set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                    cursor.execute(
                        f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                        list(extracted_data.values()) + [mainID]
                    )
                else:
                    cols = ["ID"] + list(extracted_data.keys())
                    placeholders = ", ".join(["?"] * len(cols))
                    cursor.execute(
                        f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                        [mainID] + list(extracted_data.values())
                    )

            return  # done, no Excel lookup

        # --- Normal behavior below (when label2 is NOT "no data") ---

        max_row = sheet.max_row
        max_col = sheet.max_column

        safe_label2 = sanitize_label(label2)
        resource_colname = f"resource-{safe_label2}"

        # --- 1) Find target row via adjacent (label1, label2) ---
        target_row = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col):  # up to max_col-1 because we check c and c+1
                v1 = sheet.cell(row=r, column=c).value
                v2 = sheet.cell(row=r, column=c + 1).value
                if matches(v1, label1) and matches(v2, label2):
                    target_row = r
                    break
            if target_row is not None:
                break

        if target_row is None:
            print("Target row not found")
            return  # no matching row → nothing to insert

        # --- Optionally find the column index for "Resource" ---
        resource_col = None
        if include_resource:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == "resource":
                        resource_col = getattr(cell, "col_idx", cell.column)
                        break
                if resource_col:
                    break

        # --- 2) Scan the row to find targets; capture right-hand values ---
        extracted_data = {}

        # label → col name
        col_name = f"{label1}{label2}"

        # move to look for the first value to the right
        def capture_right_of_label(row: int, label: str):
            # Find the column containing the label
            for c in range(1, max_col):
                cell_value = sheet.cell(row=row, column=c).value
                if matches(cell_value, label):

                    # Start searching to the right of this column
                    for cc in range(c + 1, max_col + 1):
                        right_val = sheet.cell(row=row, column=cc).value

                        # Skip empty or whitespace-only
                        if right_val is None:
                            continue

                        if isinstance(right_val, str):
                            rv = right_val.strip()
                            if rv == "":
                                continue
                            return rv  # return first non-empty string

                        # Non-string, non-None → return immediately
                        return right_val

                    # If no value was found to the right
                    return None

            # Label not found at all
            return None

        val = capture_right_of_label(target_row, label2)

        if val is not None:
            extracted_data[col_name] = val

        # Optionally add the Resource value, but only if it's not empty
        if include_resource and resource_col:
            resource_value = sheet.cell(row=target_row, column=resource_col).value
            if resource_value is not None and (not isinstance(resource_value, str) or resource_value.strip() != ""):
                extracted_data[resource_colname] = resource_value

        if not extracted_data:
            print("Target extracted not found")
            return  # nothing to insert

        # --- 3) Ensure table/columns exist ---
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
        table_exists = cursor.fetchone()

        needed_columns = list(extracted_data.keys())

        if not table_exists:
            cols_def = ", ".join([f"{q(col)} TEXT" for col in needed_columns])
            fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
            cursor.execute(f'''
                CREATE TABLE {q(newdatabase)} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    ref TEXT
                    {"," if cols_def else ""} {cols_def}
                    {fk_clause}
                )
            ''')
        else:
            cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
            existing_cols = [col[1] for col in cursor.fetchall()]
            if "ref" not in existing_cols and newdatabase != maindatabase:
                cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
            for col in needed_columns:
                if col not in existing_cols:
                    cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")

        # --- 4) Upsert (same rules as your working pattern) ---
        if newdatabase != maindatabase:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ref = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ref"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )
        else:
            cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
            exists = cursor.fetchone()
            if exists:
                set_clause = ", ".join([f"{q(k)} = ?" for k in extracted_data.keys()])
                cursor.execute(
                    f"UPDATE {q(newdatabase)} SET {set_clause} WHERE ID = ?",
                    list(extracted_data.values()) + [mainID]
                )
            else:
                cols = ["ID"] + list(extracted_data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                cursor.execute(
                    f"INSERT INTO {q(newdatabase)} ({', '.join(q(c) for c in cols)}) VALUES ({placeholders})",
                    [mainID] + list(extracted_data.values())
                )

    try:
        ### SQL SET-UP
        connection = sqlite3.connect(db_path)
        print("Connected to SQLite database at", db_path)

        # Create main C2C_DATABASE table if not existing
        cursor = connection.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                ID TEXT PRIMARY KEY,
                LastUpdate TEXT,
                FileName TEXT,
                Comments TEXT
            )
        ''')

        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        # Loop through Excel files with CAS number and add their info from the template
        for filename in os.listdir(folder_excels):
            full_path = os.path.join(folder_excels, filename)
            if not os.path.isfile(full_path):
                continue
            if not file_pattern.search(filename):
                continue
            # File date
            mod_time = os.path.getmtime(full_path)
            last_update = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d")

            # Extract CAS number or EC number if applicable
            match_inv = cas_pattern.search(filename)    # Check for CAS number
            comments = "There should be something here. Please check."
            if match_inv:
                inv_number = match_inv.group(1)
                comments = "CAS"
                if match_inv.group(2):
                    comments = "CAS, " + match_inv.group(2)
                #check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_needing_DB_update:
                    continue
            else:
                match_inv = ec_pattern.search(filename)
                if not match_inv:
                    print(f"Cannot extract CAS/EC from: {filename}")
                    continue
                inv_number = match_inv.group(1)
                comments = "EC"
                # check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_needing_DB_update:
                    continue
            # Check which inv you work with
            print(f"Updating CAS: {inv_number}")

            # Open the Excel file
            CPS_wb_obj = openpyxl.load_workbook(full_path)
            CPSsheet = CPS_wb_obj.active

            # Add general info
            for g_info in ["Chemical name","Common name","CAS number", "EC number", "Linked CAS Read across", "Linked CAS Monomers", "Linked CAS Degradation Products"]:
                add_info_CPS_below(CPSsheet, g_info,"C2C_DATABASE","GENERALINFO",inv_number )

            # ASSESSOR
            add_info_CPS_below(CPSsheet, {"Name assessor":"Name assessor","Date created/updated" : "Date assessed"},"C2C_DATABASE","ASSESSORS",inv_number)

            # Add various info
            for info in ["Organohalogen","Toxic metal", "Colourant", "Geological", "Biological", "Polymer", "SVHC", "VOC"]:
                   add_info_CPS_one_cell_right(CPSsheet,info,[2],[info],
                       "C2C_DATABASE","CHEMICALCLASS",inv_number)

            # Adding other info
            for o_info in ["Molecular weight","Boiling point", "Log kow (octanol-water partition coefficient)", "Vapor pressure", "Water solubility", "pH", "SMILES"]:
                   add_info_CPS_right_until_empty(CPSsheet,o_info,[2],[o_info],
                       "C2C_DATABASE","OTHERINFO",inv_number, include_resource=False)

            # CARCINOGENICITY
            for carc_type in ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK","Carcinogenicity Classified IARC",
                "Carcinogenicity Classified TLV", "Carcinogenicity experimental evidence", "Carcinogenicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,carc_type,[1],[carc_type],
                    "C2C_DATABASE","CARCINOGENICITY",inv_number)

            # ED
            for ED_type in ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,ED_type,[1],[ED_type],
                    "C2C_DATABASE","ENDOCRINE",inv_number)

            # MUTAGENICITY/GENOTOXICITY
            # General information
            for muta_type in ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,muta_type,[1],[muta_type],
                    "C2C_DATABASE","MUTAGENICITY",inv_number)
            # Point mutations
            P_mut = loop_over_to_collect_right_values(CPSsheet,"Point mutations:")  # make a string with all point mut
            #print("Point mut",P_mut) #print to see if it makes a string with the point mut
            for p_mut in P_mut:
                 add_info_right_two_markers_OECD(CPSsheet,"Point mutations:",p_mut, "C2C_DATABASE","POINTMUT",inv_number)


            # Chromosomal mutations
            Ch_mut = loop_over_to_collect_right_values(CPSsheet,
                                                      "Chromosome damaging:")  # make a string with all point mut
            #print("Ch mut",Ch_mut)  # print to see if it makes a string with the point mut
            for ch_mut in Ch_mut:
                add_info_right_two_markers_OECD(CPSsheet,"Chromosome damaging:",ch_mut,"C2C_DATABASE","CHROMDAM",inv_number)

            # REPRODUCTIVE TOXICITY
            for repro_type in ["Reprotox Classified CLP", "Reprotox Classified MAK", "Reprotox Oral NOAEL =",
                               "Reprotox Inhalation NOAEL =", "Reproductive Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,repro_type,[1],[repro_type],
                    "C2C_DATABASE","REPROTOX",inv_number)

            # DEVELOPMENTAL TOXICITY
            for devo_type in ["Developmental Classified CLP", "Developmental Classified MAK", "Developmental Oral NOAEL =",
                               "Developmental Inhalation NOAEL =", "Developmental Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,devo_type,[1],[devo_type],
                    "C2C_DATABASE","DEVELOPTOX",inv_number)

            # NEUROTOXICITY
            for neuro_type in ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
                "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,neuro_type,[1],[neuro_type],
                    "C2C_DATABASE","NEUROTOX",inv_number)

            # ORAL TOXICITY
            for oral_type in ["Oral toxicity Acute Tox classified","Oral toxicity Asp Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
                "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,oral_type,[1],[oral_type],
                    "C2C_DATABASE","ORALTOX",inv_number)

            # INHALATIVE TOXICITY
            for inhal_type in ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
                "Inhalative toxicity Acute: LC50 (gas) =", "Inhalative toxicity Acute: LC50 (vapor) =", "Inhalative toxicity Acute: LC50 (dust/mist/aerosol) =", "Inhalative toxicity Chronic: LOAEL (gas) =",
                "Inhalative toxicity Chronic: LOAEL (vapor) =", "Inhalative toxicity Chronic: LOAEL (dust/mist/aerosol) =", "Inhalative Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,inhal_type,[1],[inhal_type],
                    "C2C_DATABASE","INHALTOX",inv_number)

            # DERMAL TOXICITY
            for dermal_type in ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
                                "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]:
                add_info_CPS_right_until_empty(CPSsheet,dermal_type,[1],[dermal_type],
                    "C2C_DATABASE","DERMALTOX",inv_number)

            # SKIN/EYE IRRITATION/CORROSION
            for irrit_type in ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
                "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion", "Corrosion/irritation comments"]:
                add_info_CPS_right_until_empty(CPSsheet,irrit_type,[1],[irrit_type],
                    "C2C_DATABASE","IRRITCOR",inv_number)

            # SENSITISATION
            for sens_type in ["Skin sensitization CLP classification", "Skin sensitization MAK classification",
                "Skin sensitization testing conclusion", "Respiratory sensitization CLP classification",
                "Respiratory sensitization MAK classification", "Respiratory sensitization testing conclusion", "Sensitization comments"]:
                add_info_CPS_right_until_empty(CPSsheet,sens_type,[1],[sens_type],
                    "C2C_DATABASE","SENSITISATION",inv_number)

            # ADD Specific concentration limits

            SCL = loop_over_to_collect_right_values(CPSsheet, "Hazard classification:") # make a string with all SCL for each CAS
            #print(SCL) #print to see if it makes a string with the specific conc limits
            for spec_conc_lim in SCL:
                add_info_CPS_from_row_with_two_markers(CPSsheet,"Hazard classification:", spec_conc_lim, "Lower Limit: (%)", "Upper Limit: (%)" , "C2C_DATABASE","SCONCLIM",inv_number)

            #  OTHER CRITERIA
            for other_criteria in ["Other comments"]:
                add_info_CPS_right_until_empty(CPSsheet, other_criteria, [1], [other_criteria],
                                   "C2C_DATABASE", "OCRIT", inv_number)

            # AQUATIC TOXICITY
            for aqtox_type in ["Aquatic toxicity Acute Tox classified", "Aquatic toxicity Chronic Tox classified","Water solubility", "M factor"]:
                add_info_CPS_right_until_empty(CPSsheet, aqtox_type, [1], [aqtox_type],
                                    "C2C_DATABASE", "AQUATOX", inv_number)
                # VERTEBRATE FISH
            for fish_type in ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC =", "Fish toxicity Acute QSAR: LC50 =", "Fish toxicity Chronic QSAR: NOEC =", "Fish toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet,fish_type,[1],[fish_type],
                    "C2C_DATABASE","FISHTOX",inv_number)
                # INVERTEBRATE
            for inv_type in ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC =", "Invertebrae toxicity Acute QSAR: LC50 =", "Invertebrae toxicity Chronic QSAR: NOEC =", "Invertebrate toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet, inv_type, [1], [inv_type],
                                   "C2C_DATABASE", "INVTOX", inv_number)
                # ALGAE
            for algae_type in ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC =", "Algae toxicity Acute QSAR: LC50 =", "Algae toxicity Chronic QSAR: NOEC =", "Algae toxicity comments:"]:
                add_info_CPS_right_until_empty(CPSsheet, algae_type, [1], [algae_type],
                                   "C2C_DATABASE", "ALGAETOX", inv_number)

            # TERRESTRIAL TOXICITY
            for tertox_type in ["Terrestial toxicity CLP classification", "Terrestial toxicity Acute (Chicken): LD50=", "Terrestial toxicity Acute (Duck): LD50=",
                                "Terrestial toxicity Acute (Worm): EC50=", "Terrestial toxicity Chronic (Chicken): NOEC=", "Terrestial toxicity Chronic (Duck): NOEC=",
                                "Terrestial toxicity Chronic (Worm): NOEC=", "Terrestial toxicity comments"]:
                add_info_CPS_right_until_empty(CPSsheet, tertox_type, [1], [tertox_type],
                                   "C2C_DATABASE", "TERTOX", inv_number)
            # Other species toxicity
            for spec_tox_type in ["Any other CLP species classification"]:
                add_info_CPS_right_until_empty(CPSsheet, spec_tox_type, [1], [spec_tox_type],
                                   "C2C_DATABASE", "SPECTOX", inv_number)

            # PERSISTENCE
            for pers_type in ["OECD 301: % DOC biodegradation after 28 days", "OECD 301: % ThOD biodegradation after 28 days",
                "OECD 302 or OECD 304A: % inherent biodegradation: ", "OECD 311","QSAR prediction", "Half-life (T1/2) Air", "Half-life (T1/2) Water", "Half-life (T1/2) Soil or sediment", "Persistence comments"]:
                add_info_CPS_right_until_empty(CPSsheet,pers_type,[1],[pers_type],
                    "C2C_DATABASE","PERSISTENCE",inv_number)

            # BIOACCUMULATION
            for bioac_type in ["BCF/BAF (experimental)", "BCF/BAF (QSAR)", "Bioaccumulation comments"]:
                add_info_CPS_right_until_empty(CPSsheet,bioac_type,[1],[bioac_type],
                    "C2C_DATABASE","BIOACCUMULATION",inv_number)

            # CLIMATIC RELEVANCE
            for clima_type in ["Climatic listed?", "100 year GWP", "ODP", "Climatic relevance comments"]:
                add_info_CPS_right_until_empty(CPSsheet,clima_type,[1],[clima_type],
                    "C2C_DATABASE","CLIMATICRELEVANCE",inv_number)

            #  ADDITIONAL SOURCES
            for add_sources in ["Additional sources"]:
                add_info_CPS_right_until_empty(CPSsheet, add_sources, [1], [add_sources],
                                   "C2C_DATABASE", "ADDSOURCE", inv_number)

        connection.commit()
        print("SQL updated")
    except sqlite3.Error as e:
        print("SQLite error", e, inv_number)

    finally:
        if connection:
            connection.commit()
            connection.close()
        print("Connection closed.")




extract_info_form_excel_to_DB(db_path, folder_excels, CAS_needing_DB_update)
 ###
def make_a_backup(db_path):
    try:
        connection = sqlite3.connect(db_path)
        print("Connected to SQLite database:", db_path)

        # Create date-stamped backup filename
        today = datetime.now().strftime("%Y-%m-%d")
        base, ext = os.path.splitext(db_path)
        backup_path = f"{base}_backup_from_{today}{ext}"

        # Create backup connection
        backup_conn = sqlite3.connect(backup_path)

        # Perform the backup
        with backup_conn:
            connection.backup(backup_conn)

        print("Backup made:", backup_path)

    finally:
        if 'backup_conn' in locals():
            backup_conn.close()
        if connection:
            connection.close()
            print("Connection closed.")


### Function if recent?
#
def is_DB_data_up_to_date_with_excel(db_path, folder_excels, CAS_list):
    excel_files_that_need_updating = []
    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
        print("Connected to SQLite database at", db_path)

        cursor.execute('''
                CREATE TABLE IF NOT EXISTS C2C_DATABASE (
                    ID TEXT PRIMARY KEY,
                    LastUpdate TEXT,
                    FileName TEXT,
                    Comments TEXT
                )
            ''')
        # Regex pattern to extract CAS number from filename
        file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
        cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
        cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
        ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        for filename in os.listdir(folder_excels):
            full_path = os.path.join(folder_excels, filename)
            if not os.path.isfile(full_path):
                continue
            if not file_pattern.search(filename):
                continue
            # File date
            mod_time = os.path.getmtime(full_path)
            last_update = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d")

            # Extract CAS or EC
            match_inv = cas_pattern.search(filename)
            if match_inv is None:
                print("There should be something here. Please check.")

            if match_inv:
                inv_number = match_inv.group(1)
                comments = "CAS"
                if match_inv.group(2):
                    comments = "CAS, " + match_inv.group(2)
                #check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_list:
                    continue
            else:
                match_inv = ec_pattern.search(filename)
                if not match_inv:
                    print(f"Cannot extract CAS/EC from: {filename}")
                    continue
                inv_number = match_inv.group(1)
                comments = "EC"
                # check if CAS is in the list if not then it is skipped
                if inv_number not in CAS_list:
                    continue

            cursor.execute("""
                    UPDATE C2C_DATABASE
                    SET LastUpdate = NULL
                    WHERE LastUpdate NOT LIKE '____-__-__'
                """)

            # Check what DB currently has
            cursor.execute("SELECT LastUpdate FROM C2C_DATABASE WHERE ID = ?", (inv_number,))
            row = cursor.fetchone()

            if row is None:
                # ID not in DB -> insert
                cursor.execute(
                    "INSERT INTO C2C_DATABASE (ID, LastUpdate, FileName, Comments) VALUES (?,?,?,?)",
                    (inv_number, last_update, filename, comments)
                )
                print(f"CHANGED (was not there before): inserted {inv_number}: set LastUpdate={last_update}")
                excel_files_that_need_updating.append(inv_number)

            else:
                db_last_update = row[0]  # ISO YYYY-MM-DD

                if db_last_update is None or db_last_update < last_update:
                    # File is newer -> update
                    cursor.execute(
                        "UPDATE C2C_DATABASE SET LastUpdate = ?, FileName = ?, Comments = ? WHERE ID = ?",
                        (last_update, filename, comments, inv_number)
                    )
                    print(f"CHANGED: inserted {inv_number}: {db_last_update} -> {last_update}")
                    excel_files_that_need_updating.append(inv_number)

                else:
                    # DB is newer or same -> skip
                    print(f"NO ACTION NEEDED. CAS that are up to date {inv_number}: DB {db_last_update} >= file {last_update}.")
        return excel_files_that_need_updating

    finally:
        if connection:
            connection.commit()
            connection.close()
        print("Connection closed.")
#
# a = is_DB_data_up_to_date_with_excel(db_path, folder_excels, CAS_list=["10-00-0", "AAA"])
# print(a)
# #### Create/update C2C database with CAS numbers from Excel files ####
# def insert_json_info_to_DB(CnL_json, db_path, target_cas_list):
#     cas_hazards = {} # used later to create a list of things to update
#     with open(CnL_json, "r", encoding="utf-8") as f:
#         data = json.load(f)
#     try:
#         connection = sqlite3.connect(db_path)
#         cursor = connection.cursor()
#      #id INTEGER PRIMARY KEY AUTOINCREMENT,
#         # Ensure ECHACHEM_CL table exists
#         cursor.execute("PRAGMA foreign_keys = ON;")
#         cursor.execute("""
#         CREATE TABLE IF NOT EXISTS ECHACHEM_CL (
#             id INTEGER PRIMARY KEY AUTOINCREMENT,
#             code TEXT NOT NULL,
#             on_cl TEXT,
#             cas TEXT,
#             ec TEXT,
#             name_echachem TEXT,
#             type_classification TEXT,
#             hazards TEXT,
#             date_checked TEXT,
#             FOREIGN KEY (code) REFERENCES C2C_DATABASE(ID)
#         )
#         """)
#
#
#         print("Connected to SQLite database at:", db_path)
#
#         today = datetime.now().date()
#
#         for target_cas in target_cas_list:
#
#             # Find the entry for the CAS you want
#             entry = next((e for e in data if e.get("casNumber") == target_cas), None)
#
#             if entry is None:
#                 print(f"CAS {target_cas} not found in JSON.")
#             else:
#                 # Set up dictionary to collect all relevant info
#                 sqlinfo = {
#                     "code": entry.get("casNumber"),
#                     "on_cl": "-",
#                     "cas": "-",
#                     "ec": "-",
#                     "name_echachem": "-",
#                     "type_classification": "-",
#                     "hazards": "-"
#                 }
#
#                 print(f"Testing for: {entry.get('casNumber')}")
#
#                 #### ECHA-CHEM C&L from NEXTSDS-API ####
#                 if entry.get("found") is False:  # If the chemical was NOT found on C&L
#                     sqlinfo["on_cl"] = "No"
#                 else:  # If the chemical was found on C&L
#                     sqlinfo["on_cl"] = "Yes"
#                     sqlinfo["cas"] = entry.get("cas")
#                     sqlinfo["ec"] = entry.get("ecNumber")
#                     sqlinfo["name_echachem"] = entry.get("name")
#
#                     if entry.get("isHarmonized") is True:
#                         sqlinfo["type_classification"] = "Harmonized"
#                     else:
#                         sqlinfo["type_classification"] = "Self-classification"
#
#                     # Safe hazards extraction (prevents crashes if hazards is missing/not a dict)
#                     hazards = entry.get("hazards", {})
#                     if isinstance(hazards, dict):
#                         sqlinfo["hazards"] = hazards.get("hazardClasses", "-")
#                 print(sqlinfo)
#
#         #     with open(CnL_json, "r", encoding="utf-8") as f:
#         #         data = json.load(f)
#         #
#         #         # If JSON is a list of entries
#         #         for entry in data:
#         #             print(entry)
#         #         # Set up dictionary to collect all relevant info
#         #         sqlinfo = {"code": entry.get("casNumber"), "on_cl": "-", "cas": "-", "ec": "-", "name_echachem": "-",
#         #                    "type_classification": "-", "hazards": "-"}
#         #         print(f"Adding chemical: {entry.get("casNumber")}")
#         #
#         #         #### ECHA-CHEM C&L from NEXTSDS-API ####
#         #         if entry.get("found") == False:  # If the chemical was NOT found on C&L
#         #             sqlinfo["on_cl"] = "No"
#         #         else:  # If the chemical was found on C&L (then there is no "found" entry)
#         #             sqlinfo["on_cl"] = "Yes"
#         #             sqlinfo["cas"] = entry.get("cas")
#         #             sqlinfo["ec"] = entry.get("ecNumber")
#         #             sqlinfo["name_echachem"] = entry.get("name")
#         #             if entry.get("isHarmonized") == True:
#         #                 sqlinfo["type_classification"] = "Harmonized"
#         #             else:
#         #                 sqlinfo["type_classification"] = "Self-classification"
#         #             sqlinfo["hazards"] = entry.get("hazards")["hazardClasses"]
#         #
#                 # Check if CAS already exists
#                 cursor.execute("SELECT 1 FROM ECHACHEM_CL WHERE cas = ?", (sqlinfo["cas"],))
#                 exists = cursor.fetchone()
#                 if exists:
#                     print(f"CAS {sqlinfo['cas']} already in database")
#                     cursor.execute("SELECT 1 FROM ECHACHEM_CL WHERE cas = ? AND hazards = ?", (sqlinfo["cas"],sqlinfo["hazards"]))
#                     same_hazard = cursor.fetchone()
#                     if same_hazard:
#                         print(f"Hazards for {sqlinfo['cas']} are the same as for the last update. NO ACTION NEEDED.")
#                     else:
#                         print(f"Inserting CAS {sqlinfo['cas']}...")
#                         cursor.execute("""
#                             UPDATE ECHACHEM_CL
#                             SET hazards = ?, date_checked = ?
#                             WHERE cas = ?
#                         """, (
#                             sqlinfo["hazards"],
#                             today,
#                             sqlinfo["cas"]
#                         ))
#                         connection.commit()
#                         print(f"Hazards for {sqlinfo['cas']} are DIFFERENT as for the last update. INFO IN TABLE CnL UPDATED. ACTION REQUIRED.")
#
#                         ### Needed for the next step to gather info and update
#                         cursor.execute("SELECT hazards FROM ECHACHEM_CL WHERE code = ?",
#                                        (sqlinfo["code"],))
#                         # needed if info was added
#                         row = cursor.fetchone()
#                         hazards_list = row[0].split(",") if row and row[0] else []
#                         cas_hazards[target_cas] = hazards_list
#                         print(hazards_list)
#
#
#                 else:
#                     print(f"CAS not in CnL database: {sqlinfo['cas']}")
#                     cursor.execute(
#                         "SELECT 1 FROM C2C_DATABASE WHERE ID = ?",
#                         (sqlinfo["code"],)
#                     )
#                     exists = cursor.fetchone()
#                     if not exists:
#                         cursor.execute(
#                             "INSERT INTO C2C_DATABASE (ID) VALUES (?)",
#                             (sqlinfo["code"],)
#                         )
#                         print(f"CAS was not in the main C2C database. CAS added to the main C2C database: {sqlinfo['cas']}")
#                     else:
#                         print(f"CAS already exists in the main C2C database: {sqlinfo['cas']}")
#
#                     cursor.execute("""
#                         INSERT INTO ECHACHEM_CL (code, on_cl, cas, ec, name_echachem, type_classification, hazards, date_checked)
#                         VALUES (?, ?, ?, ?, ?, ?, ?, ?)
#                     """, (
#                         sqlinfo["code"],
#                         sqlinfo["on_cl"],
#                         sqlinfo["cas"],
#                         sqlinfo["ec"],
#                         sqlinfo["name_echachem"],
#                         sqlinfo["type_classification"],
#                         sqlinfo["hazards"],
#                         today
#                     ))
#                     connection.commit()
#                     print(f"Information inserted to CnL database: {sqlinfo['cas']}")
#                     ### Needed for the next step to gather info and update
#                     cursor.execute("SELECT hazards FROM ECHACHEM_CL WHERE code = ?",
#                                                  (sqlinfo["code"],))
#                     #needed if info was added
#                     row = cursor.fetchone()
#                     hazards_list = row[0].split(",") if row and row[0] else []
#                     cas_hazards[target_cas] = hazards_list
#                     print(hazards_list)
#
#
#         print("SQL updated")
#         return cas_hazards
#     #
#     finally:
#         if connection:
#             connection.commit()
#             connection.close()
#             print("Connection closed.")
#
# list = insert_json_info_to_DB(CnL_json, db_path, CASall)
# print(list)

###
# # Loop through Excel files with CAS number
#
# def check_if_excel_is_in_folder(folder_excels, CAS_list):
#     CAS_in_folder = []
#     CAS_not_in_folder = []
#
#     file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
#     cas_pattern = re.compile(r'CAS (\d{2,7}[-‐-–—]\d{2,3}[-‐-–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
#     ec_pattern = re.compile(r'EC (\d{2,7}[-‐-–—]\d{3}[-‐-–—]\d{1})')
#
#     # collect all inventory numbers (CAS) found in the folder
#     inv_in_folder = set()
#
#     for filename in os.listdir(folder_excels):
#         full_path = os.path.join(folder_excels, filename)
#         if os.path.isfile(full_path):
#             match = file_pattern.search(filename)
#             if match:
#                 # Extract CAS inventory number (inv_number)
#                 match_inv = cas_pattern.search(filename)
#                 if match_inv:
#                     inv_number = match_inv.group(1)  # this is the CAS
#                     inv_in_folder.add(inv_number)
#                 else:
#                     # if no CAS, then check for EC (kept from your code)
#                     match_inv = ec_pattern.search(filename)
#                     if match_inv:
#                         inv_number = match_inv.group(1)
#                     else:
#                         print(f"Issue with: {filename}")
#
#     # compare input CAS_list against what was found in folder
#     for cas in CAS_list:
#         if cas in inv_in_folder:
#             CAS_in_folder.append(cas)
#         else:
#             CAS_not_in_folder.append(cas)
#
#     return CAS_in_folder, CAS_not_in_folder
#
#
# CAS_in_folder, CAS_not_in_folder = check_if_excel_is_in_folder(folder_excels, CASall)
# print("solution:",CAS_in_folder, CAS_not_in_folder)
# def check_if_excel_is_in_folder(folder_excels)
#     CAS_in_folder = []
#
#     file_pattern = re.compile(r'CAS (.*?)\.(xlsx|xlsm)$')
#     cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.(xlsx|xlsm)$', re.IGNORECASE)
#     cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
#     ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')
#
#     for filename in os.listdir(folder_excels):
#         full_path = os.path.join(folder_excels, filename)
#         if os.path.isfile(full_path):
#             match = file_pattern.search(filename)
#             if match:
#                 print(filename)
#                 #### Update database: general info #####
#                 # Get last modification time and format it as DD/MM/YYYY
#                 mod_time = os.path.getmtime(full_path)
#                 last_update = datetime.fromtimestamp(mod_time).strftime("%d/%m/%Y")
#
#                 # Extract CAS number or EC number if applicable
#                 match_inv = cas_pattern.search(filename)  # Check for CAS number
#                 comments = "There should be something here. Please check."
#                 if match_inv:  # If CAS
#                     comments = "CAS"
#                     if match_inv.group(2):  # If there is additional info after the CAS number, save it in comments
#                         comments = "CAS, " + match_inv.group(2)
#
#                 else:  # if no CAS, then check for EC
#                     match_inv = ec_pattern.search(filename)
#                     comments = "EC"
#                 if match_inv:  # If a standard format is found, save for use in the database
#                     inv_number = match_inv.group(1)
#                 else:  # Else print the file for which there is an issue
#                     print(f"Issue with: {filename}")
#
#


# def checing_if_CAS_exists(CASall, db_path):
#     try:
#         connection = sqlite3.connect(db_path)
#         #st.success(f"Connected to SQLite database at: {db_path}")
#         cursor = connection.cursor()
#         found = []  # CAS numbers that exist in the DB
#         not_found = []  # CAS numbers that do NOT exist
#         for cas in CASall:
#             cursor.execute("SELECT 1 FROM C2C_DATABASE WHERE ID = ?", (cas,))
#             row = cursor.fetchone()
#
#             if row:
#                 found.append(cas)
#             else:
#                 not_found.append(cas)
#
#         print("CAS found in database:", found)
#         print("CAS not in database:", not_found)
#     except sqlite3.Error as e:
#         print("SQLite error:", e)
#
# db_path = '/Users/juliakulpa/Desktop/test/Database/C2Cdatabase.db'
# CASall = ['50-00-0', '110-54-3', "00-00-1"]
#
# checing_if_CAS_exists(CASall, db_path)

#
# def check_jason(CASall, API_key):
#     print('Checking API')
#
#     # Load the API key from file: It's on the dropbox under Science/Data searches/ED screener/input databases/NextSDS API key.txt
#     with open(API_key) as creds:
#         api_key = creds.readlines()[0]  # API key is on the first line
#
#     # Split up list in smaller parts (chunks)
#     def chunk_list(lst, chunk_size):
#         for i in range(0, len(lst), chunk_size):
#             yield lst[i:i + chunk_size]
#
#     start_url = "https://api.nextsds.com/jobs/start"
#     status_url = "https://api.nextsds.com/jobs/retrieve"
#     headers = {
#         "accept": "application/json",
#         "Authorization": f"Bearer {api_key}",
#         "Content-Type": "application/json"
#     }
#     # Step 1: Submit all jobs
#     jobs = []
#     for idx, cas_chunk in enumerate(chunk_list(CASall, 250)):
#         data = {
#             "taskId": "echa-api",
#             "payload": cas_chunk
#         }
#         try:
#             response = requests.post(start_url, headers=headers, json=data)
#             if response.status_code == 200:
#                 job_id = response.json().get("id")
#                 jobs.append({"id": job_id, "index": idx + 1, "done": False, "output": None})
#                 print(f"Chunk {idx + 1}: Job submitted successfully: {job_id}")
#             else:
#                 print(f"Chunk {idx + 1}: Failed to submit job")
#         except Exception as e:
#             print(f"Chunk {idx + 1}: Exception during job submission: {str(e)}")
#
#     # Step 2: Monitor all jobs in one loop
#     while not all(job["done"] for job in jobs):
#         time.sleep(10)
#         for job in jobs:
#             if job["done"]:
#                 continue
#             try:
#                 status_response = requests.get(f"{status_url}/{job['id']}", headers=headers)
#                 if status_response.status_code == 200:
#                     status_data = status_response.json()
#                     job_status = status_data.get("status")
#                     print(f"Chunk {job['index']}: Job status: {job_status}")
#                     if job_status not in ["STARTED", "EXECUTING"]:
#                         job["done"] = True
#                         job["output"] = status_data.get("output", [])
#                 elif status_response.status_code in [400, 404]:
#                     print(f"Chunk {job['index']}: Job error ({status_response.status_code})")
#                     job["done"] = True
#             except Exception as e:
#                 print(f"Chunk {job['index']}: Exception during status check: {str(e)}")
#
#     # Step 3: Combine all outputs
#     CnL_json = []
#     for job in jobs:
#         if job["output"]:
#             CnL_json.extend(job["output"])
#
#     # Save to a JSON file
#     currentdir = os.getcwd()
#     exportpath = os.path.join(currentdir,"output")
#     if not os.path.exists(exportpath):
#         os.makedirs(exportpath)
#     formatted_time = datetime.now().strftime("%Y-%m-%d %H-%M")  # Customize format as needed
#     exportjson = os.path.join(exportpath, "CnLscreener exportJSON " + formatted_time +".json")
#     with open(exportjson, "w") as json_file:
#         json.dump(CnL_json, json_file, indent=2)
#     return CnL_json
#
#
# API_key = '/Users/juliakulpa/Desktop/Test_echa/NextSDS API key.txt'
# CASall = ['50-00-0', '110-54-3']
#
# check_jason(CASall, API_key)
#
#
#
#
#
#
#
# #### SET UP ####
# import sqlite3
# import os
# import re
# import pandas as pd
# import openpyxl
# from datetime import datetime
# import traceback

# # Define the path to your SQLite database file
# C2Cpath = "/Users/juliakulpa/Desktop/test"
# C2Cfiles_path = os.path.join(C2Cpath,"CPS")
# # C2Cpath = "/Users/arche/Arche Dropbox/C2C/08_Chemical Profiling/"
# # C2Cfiles_path = os.path.join(C2Cpath)
# db_path = os.path.join(C2Cpath,"Database/C2Cdatabase.db")
#
# # Format today’s date
# today = datetime.today().strftime("%Y%m%d")  # e.g., 20250903
# # Specify whether you want to load-in the CPS files or whether you want to start from a preloaded database
# READ_IN_CPS = True

#sheet = pd.read_excel("/Users/juliakulpa/Desktop/test/CPS/CPS_CAS 10-00-0.xlsx")
# def extract_notifiers_resources_wide(sheet, mainID):
#     """
#     Scan the sheet for 'Notifiers' and 'Resources' headers (case-insensitive).
#     For each header, read up to 250 rows below (skip blanks). For each row:
#       - name := value 6 columns to the left of the header
#       - section value := the cell under the header
#     Returns a list of dicts, one per name, e.g.:
#       {'ref': mainID, 'name': 'Alpha', 'Notifiers': 'Company A', 'Resources': 5}
#     """
#
#     TARGETS = {"notifiers": "Notifiers", "resources": "Resources"}
#
#     max_row = sheet.max_row
#     max_col = sheet.max_column
#
#     # Collect into a dict keyed by name so sections merge onto one row
#     rows_by_name = {}  # name -> dict
#
#     def coerce(v):
#         # Keep numbers as numbers; trim strings
#         if v is None:
#             return None
#         if isinstance(v, str):
#             s = v.strip()
#             # try numeric cast if the string looks like a number
#             try:
#                 if "." in s:
#                     return float(s)
#                 return int(s)
#             except ValueError:
#                 return s
#         return v
#
#     def process_column(header_row, header_col, section_label):
#         for r in range(header_row + 1, min(header_row + 251, max_row + 1)):
#             val = sheet.cell(row=r, column=header_col).value
#             if val is None or (isinstance(val, str) and val.strip() == ""):
#                 continue  # skip blanks
#
#             name_col = max(1, header_col - 6)
#             name_val = sheet.cell(row=r, column=name_col).value
#             name = "" if name_val is None else str(name_val).strip()
#             if name == "":
#                 # skip rows without a usable name
#                 continue
#
#             if name not in rows_by_name:
#                 rows_by_name[name] = {"ref": mainID, "name": name}
#
#             rows_by_name[name][section_label] = coerce(val)
#
#     # Find headers and process each
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value is None:
#                 continue
#             txt = str(cell.value).strip().lower()
#             if txt in TARGETS:
#                 section = TARGETS[txt]
#                 header_row = cell.row
#                 header_col = getattr(cell, "col_idx", cell.column)
#                 if 1 <= header_col <= max_col:
#                     process_column(header_row, header_col, section)
#
#     # Convert map to list (stable, sorted by name for convenience)
#     return [rows_by_name[k] for k in sorted(rows_by_name.keys())]

#extract_notifiers_resources_wide(sheet, None)

# #### CUSTOM FUNCTIONS ####
# # Custom function for extracting info from Excel: it takes all the info below a certain cell until an empty string is reached
# # It then adds the info to a new SQL database connected to the main database
# def add_info_CPS_below(sheet, search_strings, maindatabase, newdatabase, mainID):
#     """
#     Finds each search string somewhere in the sheet (not necessarily in the first row),
#     then reads the values directly below it column-wise.
#     Collects one "row" of data across all search strings,
#     and inserts into the SQLite database.
#
#     Stops when *all* searched columns are empty in the same row.
#
#     search_strings:
#         - list: Excel labels are used as SQL column names.
#         - dict: {Excel label: SQL column name}
#     """
#
#     # Handle both list and dict
#     if isinstance(search_strings, str):
#         search_strings = [search_strings]
#     if isinstance(search_strings, list):
#         mapping = {s: s for s in search_strings}  # Excel label -> same SQL name
#     elif isinstance(search_strings, dict):
#         mapping = search_strings  # Excel label -> custom SQL name
#     else:
#         raise TypeError("search_strings must be a str, list, or dict")
#
#     # Locate the target cells (positions of the labels in Excel)
#     col_positions = {}
#     for excel_label in mapping.keys():
#         found = None
#         for row in sheet.iter_rows():
#             for cell in row:
#                 if cell.value is not None and excel_label.lower() in str(cell.value).lower():
#                     found = cell
#                     break
#             if found:
#                 break
#         if found:
#             col_positions[excel_label] = (found.column, found.row + 1)  # start below label
#
#     # Quote identifiers for SQL safety
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#     # Check if the table exists
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?",
#                    (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     if not table_exists:
#         # Create table with auto-increment ID and extracted columns
#         cols_def = ", ".join([f"{q(sql_col)} TEXT" for sql_col in mapping.values()])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
#                 ref TEXT
#                 {"," if cols_def else ""} {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         # Add missing columns if needed
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         for sql_col in mapping.values():
#             if sql_col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(sql_col)} TEXT")
#
#     # Iterate row by row until all searched columns are empty
#     row_offset = 0
#     while True:
#         row_data = {}
#         all_empty = True
#         for excel_label, (col, start_row) in col_positions.items():
#             cell = sheet.cell(row=start_row + row_offset, column=col)
#             if cell.value is not None and str(cell.value).strip() != "":
#                 sql_col = mapping[excel_label]
#                 row_data[sql_col] = cell.value
#                 all_empty = False
#         if all_empty:
#             break  # stop when all searched columns are empty in the same row
#
#         if newdatabase == maindatabase:
#             # --- Case 1: Append into the main database table ---
#             # Check if the row for this ID exists
#             cursor.execute(
#                 f"SELECT 1 FROM {q(maindatabase)} WHERE ID = ?",
#                 (mainID,)
#             )
#             exists = cursor.fetchone()
#
#             if exists:
#                 # Update only the new columns (append data)
#                 update_clause = ", ".join([f"{q(col)} = ?" for col in row_data.keys()])
#                 cursor.execute(
#                     f"UPDATE {q(maindatabase)} SET {update_clause} WHERE ID = ?",
#                     list(row_data.values()) + [mainID]
#                 )
#             else:
#                 # Insert new row with ID and these values
#                 all_cols = ["ID"] + list(row_data.keys())
#                 placeholders = ", ".join(["?"] * len(all_cols))
#                 cursor.execute(
#                     f"INSERT INTO {q(maindatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                     [mainID] + list(row_data.values())
#                 )
#
#         else:
#             # --- Case 2: Insert into child table ---
#             row_data["ref"] = mainID
#
#             # Build WHERE clause dynamically based on all row_data keys
#             where_clause = " AND ".join([f"{q(col)} = ?" for col in row_data.keys()])
#             params = list(row_data.values())
#
#             cursor.execute(
#                 f"SELECT 1 FROM {q(newdatabase)} WHERE {where_clause}",
#                 params
#             )
#             exists = cursor.fetchone()
#
#             if not exists:
#                 # Insert new row
#                 all_cols = list(row_data.keys())
#                 placeholders = ", ".join(["?"] * len(all_cols))
#                 cursor.execute(
#                     f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                     list(row_data.values())
#                 )
#
#         row_offset += 1
#
# def add_info_CPS_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
#     """
#     Finds rows containing `rowlabel`, extracts specified columns to the right,
#     plus the value in the "Resource" column of the same row,
#     and inserts or updates the data in the SQLite database.
#
#     Parameters:
#         sheet: openpyxl worksheet
#         rowlabel: string to search for in any row
#         column_offsets: list of integers (e.g., [2, 3]) for columns to the right
#         column_names: list of strings for custom SQL column names
#         maindatabase: name of the main database (for foreign key reference)
#         newdatabase: name of the table to update
#         mainID: unique identifier for the row
#     """
#
#     if len(column_offsets) != len(column_names):
#         raise ValueError("column_offsets and column_names must have the same length")
#
#     # Quote identifiers for SQL safety
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#     # Step 1a: Find the column index for "Resource"
#     resource_col = None
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value and str(cell.value).strip().lower() == "resource":
#                 resource_col = cell.column
#                 break
#         if resource_col:
#             break
#
#     # Step 1b: Find the row containing the rowlabel
#     extracted_data = {}
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
#                 row_idx = cell.row
#                 col_idx = cell.column
#                 for offset, col_name in zip(column_offsets, column_names):
#                     target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
#                     extracted_data[col_name] = target_cell.value
#                 # Always try to grab Resource from the same row
#                 extracted_data["Resource"] = (
#                     sheet.cell(row=row_idx, column=resource_col).value if resource_col else None
#                 )
#                 break
#         if extracted_data:
#             break
#
#     if not extracted_data:
#         return  # nothing to insert
#
#     # Step 2: Check if table exists
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     if not table_exists:
#         # Create table with ID, ref, and extracted columns
#         cols_def = ", ".join([f"{q(col)} TEXT" for col in column_names + ["Resource"]])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
#                 ref TEXT
#                 {"," if cols_def else ""} {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         # Add missing columns
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         if "ref" not in existing_cols and newdatabase != maindatabase:
#             cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
#         for col in column_names + ["Resource"]:
#             if col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
#
#     # Step 3: Insert or update
#     if newdatabase != maindatabase:
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             update_clause = ", ".join([f"{q(col)} = ?" for col in column_names + ["Resource"]])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
#                 [extracted_data[col] for col in column_names + ["Resource"]] + [mainID]
#             )
#         else:
#             all_cols = ['ref'] + column_names + ["Resource"]
#             placeholders = ", ".join(["?"] * len(all_cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                 [mainID] + [extracted_data[col] for col in column_names + ["Resource"]]
#             )
#
#     else:  # when newdatabase == maindatabase
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             update_clause = ", ".join([f"{q(col)} = ?" for col in column_names + ["Resource"]])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
#                 [extracted_data[col] for col in column_names + ["Resource"]] + [mainID]
#             )
#         else:
#             all_cols = ['ID'] + column_names + ["Resource"]
#             placeholders = ", ".join(["?"] * len(all_cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                 [mainID] + [extracted_data[col] for col in column_names + ["Resource"]]
#             )
#
# def add_all_info_CPS_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
#     """
#     Scans all cells in the Excel sheet for `rowlabel`, extracts specified columns to the right,
#     plus the 'Resource' column (if present), and inserts or updates each match into the SQLite database.
#
#     Parameters:
#         sheet: openpyxl worksheet
#         rowlabel: string to search for in any cell
#         column_offsets: list of integers (e.g., [2, 3]) for columns to the right
#         column_names: list of strings for custom SQL column names
#         maindatabase: name of the main database (for foreign key reference)
#         newdatabase: name of the table to update
#         mainID: foreign key reference to maindatabase
#     """
#
#     if len(column_offsets) != len(column_names):
#         raise ValueError("column_offsets and column_names must have the same length")
#
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#     # Step 1: Find the column index for "Resource"
#     resource_col = None
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value and str(cell.value).strip().lower() == "resource":
#                 resource_col = cell.column
#                 break
#         if resource_col:
#             break
#
#     # Step 2: Prepare the table
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     all_columns = column_names + ["Resource"]
#
#     if not table_exists:
#         cols_def = ", ".join([f"{q(col)} TEXT" for col in all_columns])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID TEXT PRIMARY KEY,
#                 ref TEXT,
#                 {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         for col in all_columns:
#             if col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
#
#     # Step 3: Scan all cells and process matches
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
#                 row_idx = cell.row
#                 col_idx = cell.column
#
#                 # Extract custom columns
#                 extracted_data = {}
#                 for offset, col_name in zip(column_offsets, column_names):
#                     target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
#                     extracted_data[col_name] = target_cell.value
#
#                 # Extract Resource if available
#                 extracted_data["Resource"] = (
#                     sheet.cell(row=row_idx, column=resource_col).value if resource_col else None
#                 )
#
#                 # Build unique ID (row_idx_col_idx ensures uniqueness)
#                 newID = f"{row_idx}_{col_idx}"
#
#                 # Check if row already exists
#                 cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (newID,))
#                 exists = cursor.fetchone()
#
#                 if exists:
#                     update_clause = ", ".join([f"{q(col)} = ?" for col in all_columns])
#                     cursor.execute(
#                         f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
#                         [extracted_data[col] for col in all_columns] + [newID]
#                     )
#                 else:
#                     all_cols = ['ID', 'ref'] + all_columns
#                     placeholders = ", ".join(["?"] * len(all_cols))
#                     cursor.execute(
#                         f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                         [newID, mainID] + [extracted_data[col] for col in all_columns]
#                     )
#
# def add_info_CPS_one_cell_right(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
#     """
#     Finds rows containing `rowlabel`, extracts specified columns to the right,
#     and inserts or updates the data in the SQLite database.
#
#     Parameters:
#         sheet: openpyxl worksheet
#         rowlabel: string to search for in any row
#         column_offsets: list of integers (e.g., [2, 3]) for columns to the right
#         column_names: list of strings for custom SQL column names
#         maindatabase: name of the main database (for foreign key reference)
#         newdatabase: name of the table to update
#         mainID: unique identifier for the row
#     """
#
#     if len(column_offsets) != len(column_names):
#         raise ValueError("column_offsets and column_names must have the same length")
#
#     # Quote identifiers for SQL safety
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#
#     # Step 1b: Find the row containing the rowlabel
#     extracted_data = {}
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
#                 row_idx = cell.row
#                 col_idx = cell.column
#                 for offset, col_name in zip(column_offsets, column_names):
#                     target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
#                     extracted_data[col_name] = target_cell.value
#                 break
#         if extracted_data:
#             break
#
#     if not extracted_data:
#         return  # nothing to insert
#
#     # Step 2: Check if table exists
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     if not table_exists:
#         # Create table with ID, ref, and extracted columns
#         cols_def = ", ".join([f"{q(col)} TEXT" for col in column_names + ["Resource"]])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID INTEGER PRIMARY KEY AUTOINCREMENT,
#                 ref TEXT
#                 {"," if cols_def else ""} {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         # Add missing columns
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         if "ref" not in existing_cols and newdatabase != maindatabase:
#             cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN ref TEXT")
#         for col in column_names:
#             if col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
#
#     # Step 3: Insert or update
#     if newdatabase != maindatabase:
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ref = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ref = ?",
#                 [extracted_data[col] for col in column_names] + [mainID]
#             )
#         else:
#             all_cols = ['ref'] + column_names
#             placeholders = ", ".join(["?"] * len(all_cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                 [mainID] + [extracted_data[col] for col in column_names]
#             )
#
#     else:  # when newdatabase == maindatabase
#         cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (mainID,))
#         exists = cursor.fetchone()
#         if exists:
#             update_clause = ", ".join([f"{q(col)} = ?" for col in column_names])
#             cursor.execute(
#                 f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
#                 [extracted_data[col] for col in column_names] + [mainID]
#             )
#         else:
#             all_cols = ['ID'] + column_names
#             placeholders = ", ".join(["?"] * len(all_cols))
#             cursor.execute(
#                 f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                 [mainID] + [extracted_data[col] for col in column_names]
#             )
#
# def add_all_info_CPS_right_no_resource(sheet, rowlabel, column_offsets, column_names, maindatabase, newdatabase, mainID):
#     """
#     Scans all cells in the Excel sheet for `rowlabel`, extracts specified columns to the right.
#
#     Parameters:
#         sheet: openpyxl worksheet
#         rowlabel: string to search for in any cell
#         column_offsets: list of integers (e.g., [2, 3]) for columns to the right
#         column_names: list of strings for custom SQL column names
#         maindatabase: name of the main database (for foreign key reference)
#         newdatabase: name of the table to update
#         mainID: foreign key reference to maindatabase
#     """
#
#     if len(column_offsets) != len(column_names):
#         raise ValueError("column_offsets and column_names must have the same length")
#
#     def q(name: str) -> str:
#         return f'"{name}"'
#
#     # Step 1: Prepare the table
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (newdatabase,))
#     table_exists = cursor.fetchone()
#
#     all_columns = column_names
#
#     if not table_exists:
#         cols_def = ", ".join([f"{q(col)} TEXT" for col in all_columns])
#         fk_clause = f", FOREIGN KEY (ref) REFERENCES {q(maindatabase)}(ID)" if newdatabase != maindatabase else ""
#         cursor.execute(f'''
#             CREATE TABLE {q(newdatabase)} (
#                 ID TEXT PRIMARY KEY,
#                 ref TEXT,
#                 {cols_def}
#                 {fk_clause}
#             )
#         ''')
#     else:
#         cursor.execute(f"PRAGMA table_info({q(newdatabase)})")
#         existing_cols = [col[1] for col in cursor.fetchall()]
#         for col in all_columns:
#             if col not in existing_cols:
#                 cursor.execute(f"ALTER TABLE {q(newdatabase)} ADD COLUMN {q(col)} TEXT")
#
#     # Step 3: Scan all cells and process matches
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value is not None and rowlabel.lower() in str(cell.value).lower():
#                 row_idx = cell.row
#                 col_idx = cell.column
#
#                 # Extract custom columns
#                 extracted_data = {}
#                 for offset, col_name in zip(column_offsets, column_names):
#                     target_cell = sheet.cell(row=row_idx, column=col_idx + offset)
#                     extracted_data[col_name] = target_cell.value
#
#                 # Build unique ID (row_idx_col_idx ensures uniqueness)
#                 newID = f"{row_idx}_{col_idx}"
#
#                 # Check if row already exists
#                 cursor.execute(f"SELECT 1 FROM {q(newdatabase)} WHERE ID = ?", (newID,))
#                 exists = cursor.fetchone()
#
#                 if exists:
#                     update_clause = ", ".join([f"{q(col)} = ?" for col in all_columns])
#                     cursor.execute(
#                         f"UPDATE {q(newdatabase)} SET {update_clause} WHERE ID = ?",
#                         [extracted_data[col] for col in all_columns] + [newID]
#                     )
#                 else:
#                     all_cols = ['ID', 'ref'] + all_columns
#                     placeholders = ", ".join(["?"] * len(all_cols))
#                     cursor.execute(
#                         f"INSERT INTO {q(newdatabase)} ({', '.join(q(col) for col in all_cols)}) VALUES ({placeholders})",
#                         [newID, mainID] + [extracted_data[col] for col in all_columns]
#                     )
#
# def extract_notifiers_resources(sheet, mainID):
#     """
#     Finds 'Notifiers' and 'Resources' headers in the sheet.
#     Reads up to 250 cells below each, skipping blanks.
#     For each row:
#         - 'name' = cell 6 columns to the left
#         - 'value' = cell under the header
#     Returns a list of dicts with fields:
#         ref, section, name, value
#     """
#
#     TARGETS = {"notifiers": "Notifiers", "resources": "Resources"}
#     results = []
#
#     max_row = sheet.max_row
#     max_col = sheet.max_column
#
#     def process_column(header_row, header_col, section_label):
#         """Read up to 250 cells below a given header"""
#         for r in range(header_row + 1, min(header_row + 251, max_row + 1)):
#             cell_val = sheet.cell(row=r, column=header_col).value
#             if cell_val is None or (isinstance(cell_val, str) and cell_val.strip() == ""):
#                 continue  # skip empty cells, don't stop
#
#             name_col = max(1, header_col - 6)
#             name_val = sheet.cell(row=r, column=name_col).value
#
#             entry = {
#                 "ref": mainID,
#                 "section": section_label,
#                 "name": "" if name_val is None else str(name_val).strip(),
#                 "value": str(cell_val).strip(),
#             }
#             results.append(entry)
#
#     # Scan for 'Notifiers' and 'Resources' headers
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value is None:
#                 continue
#             text = str(cell.value).strip().lower()
#             if text in TARGETS:
#                 section = TARGETS[text]
#                 header_row = cell.row
#                 header_col = getattr(cell, "col_idx", cell.column)
#                 if 1 <= header_col <= max_col:
#                     process_column(header_row, header_col, section)
#
#     return results
#
# extract_notifiers_resources()

# def add_info_checkstring(table_name,id_column,mainID,search_column,search_string,update_column,update_string):
#     """
#     Checks if `search_string` exists in `search_column` for a given `mainID`.
#     If found, updates `update_column` with `update_string`.
#
#     Parameters:
#         db_path (str): Path to the SQLite database file
#         table_name (str): Name of the table to query
#         id_column (str): Name of the ID column (e.g., "ID")
#         mainID (str): The ID value to look for
#         search_column (str): Column to search for the string
#         search_string (str): String to search for
#         update_column (str): Column to update
#         update_string (str): String to insert if match is found
#     """
#
#     # Get existing columns
#     cursor.execute(f"PRAGMA table_info({table_name})")
#     existing_columns = [col[1] for col in cursor.fetchall()]
#
#     # Add missing columns if needed
#     if update_column not in existing_columns:
#             cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN '{update_column}' TEXT")
#
#     # Check if the string exists for the given ID
#     cursor.execute(f"""
#         SELECT 1 FROM {table_name}
#         WHERE {id_column} = ? AND {search_column} LIKE ?
#     """, (mainID, f"%{search_string}%"))
#
#     if cursor.fetchone():
#         # Update the target column
#         cursor.execute(f"""
#             UPDATE {table_name}
#             SET {update_column} = ?
#             WHERE {id_column} = ?
#         """, (update_string, mainID))
#

#### Create/update C2C database with CAS numbers from Excel files ####

# if READ_IN_CPS == True:
#     try:
#         ### SQL SET-UP
#         connection = sqlite3.connect(db_path)
#         print("Connected to SQLite database at", db_path)
#
#         # Create main C2C_DATABASE table if not existing
#         cursor = connection.cursor()
#         cursor.execute('''
#             CREATE TABLE IF NOT EXISTS C2C_DATABASE (
#                 ID TEXT PRIMARY KEY,
#                 LastUpdate TEXT,
#                 FileName TEXT,
#                 Comments TEXT
#             )
#         ''')
#
#         # Regex pattern to extract CAS number from filename
#         file_pattern = re.compile(r'CAS (.*?)\.xlsx')
#         cas_pattern = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})(.*?)\.xlsx', re.IGNORECASE)
#         cas_pattern_strict = re.compile(r'CAS (\d{2,7}[-‐‑–—]\d{2,3}[-‐‑–—]\d{1})', re.IGNORECASE)
#         ec_pattern = re.compile(r'EC (\d{2,7}[-‐‑–—]\d{3}[-‐‑–—]\d{1})')

        # # Loop through Excel files with CAS number and add their info from the template
        # for filename in os.listdir(C2Cfiles_path):
        #     full_path = os.path.join(C2Cfiles_path, filename)
        #     if os.path.isfile(full_path):
        #         match = file_pattern.search(filename)
        #         if match:
        #             print(filename)
        #             #### Update database: general info #####
        #             # Get last modification time and format it as DD/MM/YYYY
        #             mod_time = os.path.getmtime(full_path)
        #             last_update = datetime.fromtimestamp(mod_time).strftime("%d/%m/%Y")
        #
        #             # Extract CAS number or EC number if applicable
        #             match_inv = cas_pattern.search(filename)    # Check for CAS number
        #             comments = "There should be something here. Please check."
        #             if match_inv:     # If CAS
        #                 comments = "CAS"
        #                 if match_inv.group(2):  # If there is additional info after the CAS number, save it in comments
        #                     comments = "CAS, " +  match_inv.group(2)
        #
        #             else:   # if no CAS, then check for EC
        #                 match_inv = ec_pattern.search(filename)
        #                 comments = "EC"
        #             if match_inv:   # If a standard format is found, save for use in the database
        #                 inv_number = match_inv.group(1)
        #             else:   # Else print the file for which there is an issue
        #                 print(filename)


#                     #### Update database: extract info on CAS from file if new/more recent #####
#                     cursor.execute('SELECT 1 FROM C2C_DATABASE WHERE ID = ? AND LastUpdate > ?', (inv_number, last_update))
#                     exists = cursor.fetchone()
#                     if not exists:  # If there is no info or there is more recent info
#                         # Update general entry in the database
#                         cursor.execute(
#                             'INSERT OR REPLACE INTO C2C_DATABASE (ID, LastUpdate, FileName , Comments) VALUES (?,?,?,?)',
#                             (inv_number, last_update, filename, comments))
#
#                         # Open the Excel file
#                         CPS_wb_obj = openpyxl.load_workbook(full_path)
#                         CPSsheet = CPS_wb_obj.active
#
#                         # CHEMICAL NAME
#                         add_info_CPS_below(CPSsheet, ["Chemical name"],"C2C_DATABASE","CHEMICALNAMES",inv_number)
#
#                         # ASSESSOR
#                         add_info_CPS_below(CPSsheet, {"Name assessor":"Name assessor","Date created/updated" : "Date assessed"},"C2C_DATABASE","ASSESSORS",inv_number)
#
#                         # CHECKED status
#                         add_info_CPS_below(CPSsheet, ["Checked"],"C2C_DATABASE","C2C_DATABASE",inv_number)
#
#                         # Add various info
#                         for info in ["Common name","Organohalogens","Toxic metal", "Colourant", "Geological", "Biological", "Polymer"]:
#                             add_all_info_CPS_right_no_resource(CPSsheet,info,[2],[info],
#                                 "C2C_DATABASE","C2C_DATABASE",inv_number)
#                         # # # Add SMILES
#                         # for info in ["SMILES"]:
#                         #     add_info_CPS_right(CPSsheet,info,[1],[info],
#                         #         "C2C_DATABASE","C2C_DATABASE",inv_number)
#                         # # # # Molecular formula
#                         # add_info_CPS_below(CPSsheet, ["Molecular Formula"], "C2C_DATABASE", "C2C_DATABASE", inv_number)
#
#                         # CARCINOGENICITY
#                         for carc_type in ["Carcinogenicity Classified CLP", "Carcinogenicity Classified MAK","Carcinogenicity Classified IARC",
#                             "Carcinogenicity Classified TLV", "Carcinogenicity Comments"]:
#                             add_info_CPS_right(CPSsheet,carc_type,[1],[carc_type],
#                                 "C2C_DATABASE","CARCINOGENICITY",inv_number)
#
# #                         # ED
# #                         for ED_type in ["Endocrine Classified CLP", "Endocrine evidence", "Endocrine Disruption Comments"]:
# #                             add_info_CPS_right(CPSsheet,ED_type,[1],[ED_type],
# #                                 "C2C_DATABASE","ENDOCRINE",inv_number)
# #
# #                         # MUTAGENICITY/GENOTOXICITY
# #                         for muta_type in ["Mutagenicity Classified CLP", "Mutagenicity Classified MAK","Mutagenicity Comments"]:
# #                             add_all_info_CPS_right(CPSsheet,muta_type,[1],[muta_type],
# #                                 "C2C_DATABASE","MUTAGENICITY",inv_number)
# #                         # For the strings with multiple possible hits
# #                         for muta_type in ["OECD 471", "OECD 473", "OECD 474", "OECD 475", "OECD 476", "OECD 478",
# #                             "OECD 483", "OECD 485", "OECD 486", "OECD 487", "OECD 488", "OECD 489", "OECD 490", "No data"]:
# #                             add_all_info_CPS_right(CPSsheet,muta_type,[3],[muta_type],
# #                                 "C2C_DATABASE","MUTAGENICITY",inv_number)
# #
# #                         # REPRODUCTIVE TOXICITY
# #                         for repro_type in ["Reprotox Classified CLP", "Reprotox Classified MAK", "Oral NOAEL =",
# #                                            "Inhalation NOAEL =", "Reproductive Toxicity Comments"]:
# #                             add_info_CPS_right(CPSsheet,repro_type,[1],[repro_type],
# #                                 "C2C_DATABASE","REPROTOX",inv_number)
# #
# #                         # DEVELOPMENTAL TOXICITY
# #                         for devo_type in ["Developmental Classified CLP", "Developmental Classified MAK", "Oral NOAEL =",
# #                                            "Inhalation NOAEL =", "Developmental Toxicity Comments"]:
# #                             add_info_CPS_right(CPSsheet,devo_type,[1],[devo_type],
# #                                 "C2C_DATABASE","DEVELOPTOX",inv_number)
# #
# #                         # ORAL TOXICITY
# #                         for oral_type in ["Oral toxicity Acute Tox classified", "Oral toxicity STOT classified", "Oral Acute: LD50 =",
# #                             "Oral Chronic: LOAEL =", "Oral Toxicity Comments"]:
# #                             add_info_CPS_right(CPSsheet,oral_type,[1],[oral_type],
# #                                 "C2C_DATABASE","ORALTOX",inv_number)
# #
# #                         # INHALATIVE TOXICITY
# #                         for inhal_type in ["Inhalative toxicity Acute Tox classification", "Inhalative toxicity STOT classified",
# #                             "Acute: LC50 (gas) =", "Acute: LC50 (vapor) =", "Acute: LC50 (dust/mist/aerosol) =", "Chronic: LOAEL (gas) =",
# #                             "Chronic: LOAEL (vapor) =", "Chronic: LOAEL (dust/mist/aerosol) =", "Boiling Point", "Inhalative Toxicity Comments"]:
# #                             add_info_CPS_right(CPSsheet,inhal_type,[1],[inhal_type],
# #                                 "C2C_DATABASE","INHALTOX",inv_number)
# #
# #                         # DERMAL TOXICITY
# #                         for dermal_type in ["Dermal toxicity Acute Tox classified", "Dermal toxicity STOT classified",
# #                                             "Dermal Acute: LD50 =", "Dermal Chronic: LOAEL =", "Dermal Toxicity Comments"]:
# #                             add_info_CPS_right(CPSsheet,dermal_type,[1],[dermal_type],
# #                                 "C2C_DATABASE","DERMALTOX",inv_number)
# #
# #                         # NEUROTOXICITY
# #                         for neuro_type in ["Neurotox Classified CLP", "Neurotox on a list", "Neurotox scientific evidence?",
# #                             "Neurotox chronic LOAEL", "Neurtox STOT LOAEL", "Neurotox Comments"]:
# #                             add_info_CPS_right(CPSsheet,neuro_type,[1],[neuro_type],
# #                                 "C2C_DATABASE","NEUROTOX",inv_number)
# #
# #                         # SKIN/EYE IRRITATION/CORROSION
# #                         for irrit_type in ["Skin irritation classification", "Skin testing: conclusion", "Eye irritation classification",
# #                             "Eye testing conclusion", "Respiratory irritation classification", "Respiratory testing conclusion"]:
# #                             add_info_CPS_right(CPSsheet,irrit_type,[1],[irrit_type],
# #                                 "C2C_DATABASE","IRRITCOR",inv_number)
# #
# #                         # SENSITISATION
# #                         for sens_type in ["Skin sensitization classification", "Skin sensitization testing conclusion",
# #                             "Skin sensitization classified MAK", "Inhalation sensitization classification",
# #                             "Inhalation sensitization testing conclusion", "Inhalation sensitization classified MAK"]:
# #                             add_info_CPS_right(CPSsheet,sens_type,[1],[sens_type],
# #                                 "C2C_DATABASE","SENSITISATION",inv_number)
# #
# #                         # AQUATIC TOXICITY
# #                             # VERTEBRATE
# #                         for fish_type in ["Fish toxicity Acute: LC50 (96h) =", "Fish toxicity Chronic: NOEC ="]:
# #                             add_info_CPS_right(CPSsheet,fish_type,[1],[fish_type],
# #                                 "C2C_DATABASE","FISHTOX",inv_number)
# #                             # INVERTEBRATE
# #                         for inv_type in ["Invertebrate toxicity Acute: L(E)C50 (48h) =", "Invertebrae toxicity Chronic: NOEC ="]:
# #                             add_info_CPS_right(CPSsheet, inv_type, [1], [inv_type],
# #                                                "C2C_DATABASE", "INVTOX", inv_number)
# #                             # ALGAE
# #                         for algae_type in ["Algae toxicity Acute: L(E)C50 (72/96h) =", "Algae toxicity Chronic: NOEC ="]:
# #                             add_info_CPS_right(CPSsheet, algae_type, [1], [algae_type],
# #                                                "C2C_DATABASE", "ALGAETOX", inv_number)
# #                             # General aquatic tox
# #                             for aqtox_type in ["Aquatic toxicity CLP classification","Water solubility", "M factor: "]:
# #                                 add_info_CPS_right(CPSsheet, aqtox_type, [1], [aqtox_type],
# #                                                    "C2C_DATABASE", "AQUATOX", inv_number)
# #
# #                         # TERRESTRIAL TOXICITY
# #                         for tertox_type in ["Terrestial toxicity CLP classification"]:
# #                             add_info_CPS_right(CPSsheet, tertox_type, [1], [tertox_type],
# #                                                "C2C_DATABASE", "TERTOX", inv_number)
# #
# #                         # PERSISTENCE
# #                         for pers_type in ["% DOC biodegradation after 28 days", "% ThOD biodegradation after 28 days",
# #                             "Degradation halflife time in air", "Degradation halflife time in water", "soil or sediment", "QSAR prediction"]:
# #                             add_info_CPS_right(CPSsheet,pers_type,[1],[pers_type],
# #                                 "C2C_DATABASE","PERSISTENCE",inv_number)
# #
# #                         # BIOACCUMULATION
# #                         for bioac_type in ["BCF/BAF (QSAR)", "BCF/BAF (experimental)", "Log kow", "Molecular weight"]:
# #                             add_info_CPS_right(CPSsheet,bioac_type,[1],[bioac_type],
# #                                 "C2C_DATABASE","BIOACCUMULATION",inv_number)
# #
# #                         # CLIMATIC RELEVANCE
# #                         for clima_type in ["Climatic relevance/ozone depletion potential"]:
# #                             add_info_CPS_right(CPSsheet,clima_type,[2],[clima_type],
# #                                 "C2C_DATABASE","CLIMATICREL",inv_number)
# #
# #                         # PHYSICAL PROPERTIES
# #                         for physchem_type in ["VOC designation", "Molecular weight", "Boiling point", "Log kow (octanol-water partition coefficient)",
# #                             "Vapor pressure", "Water solubility"]:
# #                             add_info_CPS_right(CPSsheet,physchem_type,[1],[physchem_type],
# #                                 "C2C_DATABASE","PHYSCHEM",inv_number)
# #
#         connection.commit()
#         print("SQL updated")
#
#     except sqlite3.Error as e:
#         print("SQLite error", e, inv_number)
#
#     finally:
#         if connection:
#
#             # Create new database copy with date of today (master database is also updated)
#             # Define new filename
#             backup_path = db_path.replace(".db", f"_{today}.db")
#             # Create backup connection
#             backup_conn = sqlite3.connect(backup_path)
#             # Perform the backup
#             with backup_conn:
#                 connection.backup(backup_conn)
#                 print("Backup made: " + backup_path)
#             backup_conn.close()
#
#             connection.close()
#             print("Connection closed.")
#
#
#
# #### Save Database as Excel ####
# try:
#     ### SQL SET-UP
#     connection = sqlite3.connect(db_path)
#     print("Reconnected to database to export to Excel", db_path)
#
#     # Get a list of all tables in the database
#     tables_df = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';", connection)
#     tables = tables_df['name'].tolist()
#     tables = [t for t in tables if t not in ("C2C_DATABASE", "ECHACHEM_CL")] # Exclude the main table from the aggregation
#     # Also excludes ECHA-CHEM as that datatable is not connected through ref
#     cte_list = []
#     join_list = []
#     main_table = "C2C_DATABASE"
#
#     # AGGREGATE ALL DEPENDENT TABLES SO THAT WE CAN EASILY USE LEFTJOIN
#     for table in tables:
#         cursor = connection.cursor()
#         cursor.execute(f"PRAGMA table_info({table})")
#         columns_info = cursor.fetchall()
#         cols = [col[1] for col in columns_info if col[1] not in ('ref', 'ID')]
#
#         if not cols or 'ref' not in [col[1] for col in columns_info]:
#             continue  # skip tables without 'ref'
#
#         # Build GROUP_CONCAT for each column
#         group_concat_cols = ",\n       ".join([f"GROUP_CONCAT([{col}], CHAR(10)) AS [{col}]" for col in cols])
#
#         # Define CTE
#         cte_name = f"{table}Agg"
#         cte = f"{cte_name} AS (\n    SELECT ref,\n           {group_concat_cols}\n    FROM {table}\n    GROUP BY ref\n)"
#         cte_list.append(cte)
#
#         # Prepare LEFT JOIN
#         join_list.append(f"LEFT JOIN {cte_name} {table[:4].lower()} ON {main_table}.ID = {table[:4].lower()}.ref")
#
#     # Combine all CTEs
#     cte_sql = ",\n".join(cte_list)
#
#     # Select all columns from main_table + aggregated tables
#     cursor.execute(f"PRAGMA table_info({main_table})")
#     main_cols = [f"{main_table}.[{col[1]}]" for col in cursor.fetchall()]
#
#     select_cols = main_cols.copy()
#     for table in tables:
#         cursor.execute(f"PRAGMA table_info({table})")
#         for col in cursor.fetchall():
#             if col[1] not in ('ref',"ID"):
#                 select_cols.append(f"{table[:4].lower()}.[{col[1]}]")
#
#     select_sql = ",\n    ".join(select_cols)
#     join_sql = "\n".join(join_list)
#
#     final_query = f"WITH {cte_sql}\nSELECT\n    {select_sql}\nFROM {main_table}\n{join_sql};"
#
#     df = pd.read_sql_query(final_query, connection)
#
#     # Make sure everything is readable for Excel
#     def sanitize_excel_cells(df):
#         def fix_value(val):
#             if isinstance(val, str):
#                 # If it starts with '=', prefix with a single quote to force Excel to treat it as text
#                 if val.strip().startswith("="):
#                     return "'" + val
#                 # Remove control characters that break XML
#                 return ''.join(ch for ch in val if ch.isprintable())
#             return val
#
#         for col in df.columns:
#             df[col] = df[col].map(fix_value)
#         return df
#
#
#     # Clean both headers and cell values
#     # df.columns = clean_excel_headers(df.columns)  # from earlier
#     df = sanitize_excel_cells(df)
#
#     # Write away the database as Excel
#     with pd.ExcelWriter(C2Cpath + '/Database/C2Cdatabase ' + today + '.xlsx', engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name="C2C DATABASE", index=False)
#     print("Database exported to Excel")
#
#
# except sqlite3.Error as e:
#     # Catches SQLite-specific errors
#     print("SQLite error:", e)
#     traceback.print_exc()  # prints the full traceback
#
# except Exception as e:
#     # Catches other Python errors
#     print("General error:", e)
#     traceback.print_exc()
#
# finally:
#     # Always close connection if it was created
#     try:
#         connection.close()
#         print("Connection closed.")
#     except NameError:
#         pass  # connection was never created